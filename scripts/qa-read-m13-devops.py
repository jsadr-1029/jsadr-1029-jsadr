#!/usr/bin/env python3
"""Lee la hoja M13-Sync DevOps del Excel de plan de pruebas."""
from openpyxl import load_workbook

wb = load_workbook("/home/z/my-project/download/plan-pruebas-qa-jsadr-actualizado.xlsx", data_only=False)
target = None
for s in wb.sheetnames:
    if "M13" in s or "Sync" in s or "DevOps" in s:
        target = s
        break
print(f"Hoja objetivo: {target}")
if not target:
    raise SystemExit("No se encontro hoja M13-Sync DevOps")

ws = wb[target]
print(f"Dimensiones: {ws.max_row} filas x {ws.max_column} columnas\n")

print("=== Filas de datos ===")
for row in range(1, ws.max_row + 1):
    fila = []
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=row, column=col).value
        fila.append(str(v) if v is not None else "")
    if any(fila):
        print(f"Fila {row}: {' | '.join(fila)}")
