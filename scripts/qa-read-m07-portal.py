#!/usr/bin/env python3
"""Lee los TCs pendientes del módulo M07-Portal Cliente desde el Excel."""
import openpyxl
from pathlib import Path

EXCEL = Path("/home/z/my-project/download/plan-pruebas-qa-jsadr-actualizado.xlsx")

wb = openpyxl.load_workbook(EXCEL, data_only=False)
# Buscar la hoja de M07
hoja_nombre = None
for name in wb.sheetnames:
    if "M07" in name or "Portal Cliente" in name:
        hoja_nombre = name
        break
print(f"Hoja encontrada: {hoja_nombre}")
print(f"Todas las hojas: {wb.sheetnames}")
print()

ws = wb[hoja_nombre]
print(f"Dimensiones: {ws.dimensions}, max_row={ws.max_row}, max_col={ws.max_column}")
print()

# Imprimir headers (fila 4 usualmente)
print("=== HEADERS (fila 4) ===")
for col in range(1, ws.max_column + 1):
    v = ws.cell(row=4, column=col).value
    print(f"  col {col}: {v!r}")

print()
print("=== FILAS 5-19 (TCs) ===")
for row in range(5, min(ws.max_row + 1, 25)):
    fila = []
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=row, column=col).value
        fila.append(str(v) if v is not None else "")
    print(f"Fila {row}: {fila}")
