"""Actualiza el Excel con los hallazgos y reparos de M05-Correo Electrónico."""
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
import shutil

WB_PATH = "/home/z/my-project/upload/plan-pruebas-qa-jsadr.xlsx"
DEST_PATH = "/home/z/my-project/download/plan-pruebas-qa-jsadr-actualizado.xlsx"

# Sheet "7. M05-Correo Electrónico" rows 8, 13, 14, 18, 19
updates = {
    "7. M05-Correo Electrónico": [
        (8, "TC-MAIL-004",
         "HTTP 400 con codigo=CLIENTE_SIN_EMAIL. Mensaje: 'Tu cuenta no tiene un correo electrónico registrado'.",
         "REPARO v4.8: ANTES el OTP se generaba ANTES de validar si el cliente tenía email. Si el cliente no tenía email, el OTP se había generado en memoria (desperdicio + riesgo de logs). AHORA: la validación de email ocurre ANTES de generar el OTP. Se añadió codigo: 'CLIENTE_SIN_EMAIL' a la respuesta. Cliente sin email en BD: TEST 2 (cédula 888888888)."),
        (13, "TC-MAIL-009",
         "HTTP 200 success=true. isEthereal=true. previewUrl devuelta por nodemailer.getTestMessageUrl.",
         "Modo Ethereal ya implementado en src/lib/email.ts obtenerTransporter(): si NODE_ENV !== 'production' y no hay SMTP configurado, crea cuenta de prueba Ethereal (smtp.ethereal.email:587) y devuelve previewUrl. En producción lanza error explícito (no usa Ethereal silenciosamente)."),
        (14, "TC-MAIL-010",
         "HTTP 400 con codigo=EMAIL_INVALIDO. Mensaje: 'El email del destinatario no tiene un formato válido'.",
         "REPARO v4.8: ANTES POST /api/email {accion: 'enviar-prueba'} solo validaba !to (truthy), permitiendo strings como 'no-es-email'. enviarEmail() sí validaba pero retornaba success:false (HTTP 200), confundiendo al cliente. AHORA: validación con regex /^[^\\s@]+@[^\\s@]+\\.[^\\s@]+$/ en la API route → HTTP 400 EMAIL_INVALIDO antes de llamar a enviarEmail. Defensa en profundidad: email.ts también valida."),
        (18, "TC-MAIL-014",
         "Cada correo enviado se registra en EnvioCorreo (éxito y fallo) y NotificacionLog (OTP). Campos: destinatario, asunto, estado, fechaEnvio.",
         "Trazabilidad completa ya implementada: enviarEmail() registra en EnvioCorreo con estado ENVIADO (Brevo API o SMTP fallback) o FALLIDO (error). solicitar-otp registra en NotificacionLog con tipo OTP. BD: 101 EnvioCorreo (78 ENVIADO, 23 FALLIDO), 11 NotificacionLog. Schema EnvioCorreo tiene: destinatario, asunto, estado, fechaEnvio, mensajeError, metadata (messageId, via)."),
        (19, "TC-MAIL-015",
         "HTTP 200 success=false. Error sanitizado devuelto al cliente con codigo (SMTP_AUTH_FAILED, SMTP_CONN_ERROR, SMTP_TLS_ERROR, SMTP_ERROR).",
         "REPARO v4.8 de seguridad: ANTES probarSmtp() exponía error.message al cliente en el message de respuesta, filtrando detalles internos (host, puerto, credenciales parciales). AHORA: clasifica el error y devuelve mensaje genérico + codigo sanitizado. Detalles completos del error quedan solo en console.error del server. Códigos: SMTP_AUTH_FAILED (535, 525, 5.7.1), SMTP_CONN_ERROR (ECONNREFUSED, timeout), SMTP_TLS_ERROR (ssl, tls, certificate), SMTP_ERROR (genérico). BD: 23 EnvioCorreo FALLIDO confirman trazabilidad del error."),
    ],
}

green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
green_font = Font(color="006100", bold=True)
yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

wb = load_workbook(WB_PATH)
total_updated = 0

for sheet_name, items in updates.items():
    ws = wb[sheet_name]
    for row_num, tc_id, nuevo_riesgo, nuevo_hallazgo in items:
        cell_id = ws.cell(row=row_num, column=2).value
        if cell_id != tc_id:
            print(f"⚠ {sheet_name} fila {row_num}: esperado {tc_id}, encontrado {cell_id}")
            continue

        cell_state = ws.cell(row=row_num, column=13)
        cell_state.value = "Aprobado"
        cell_state.fill = green_fill
        cell_state.font = green_font
        cell_state.alignment = Alignment(horizontal="center", vertical="center")

        cell_riesgo = ws.cell(row=row_num, column=11)
        cell_riesgo.value = nuevo_riesgo
        cell_riesgo.fill = yellow_fill
        cell_riesgo.alignment = Alignment(wrap_text=True, vertical="top")

        cell_hallazgo = ws.cell(row=row_num, column=12)
        cell_hallazgo.value = nuevo_hallazgo
        cell_hallazgo.fill = yellow_fill
        cell_hallazgo.alignment = Alignment(wrap_text=True, vertical="top")

        print(f"✓ {sheet_name} fila {row_num} ({tc_id}): Estado=Aprobado")
        total_updated += 1

wb.save(WB_PATH)
print(f"\n=== Total actualizado: {total_updated} TCs ===")

shutil.copy(WB_PATH, DEST_PATH)
print(f"Copia para entrega: {DEST_PATH}")
