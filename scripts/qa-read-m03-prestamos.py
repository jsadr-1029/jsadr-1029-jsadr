"""Detalle completo de TCs de M03-Préstamos."""
from openpyxl import load_workbook
wb = load_workbook("/home/z/my-project/upload/plan-pruebas-qa-jsadr.xlsx", data_only=True)
ws = wb["5. M03-Préstamos"]
print(f"=== {ws.title}  max_row={ws.max_row} ===\n")

# Cabecera
for c in range(1, ws.max_column+1):
    h = ws.cell(row=3, column=c).value
    if h:
        print(f"  col {c:2d}: {h}")
print()

for r in range(4, ws.max_row + 1):
    tc_id = ws.cell(row=r, column=2).value
    if not tc_id or not str(tc_id).startswith('TC-PRE'):
        continue
    modulo = ws.cell(row=r, column=3).value
    func = ws.cell(row=r, column=4).value
    caso = ws.cell(row=r, column=5).value
    datos = ws.cell(row=r, column=6).value
    pasos = ws.cell(row=r, column=7).value
    esperado = ws.cell(row=r, column=8).value
    crit = ws.cell(row=r, column=9).value
    estado = ws.cell(row=r, column=13).value
    riesgo = ws.cell(row=r, column=11).value
    hallazgo = ws.cell(row=r, column=12).value

    print(f"━━━ row {r} | {tc_id} | estado={estado} ━━━")
    print(f"  Funcionalidad: {func}")
    print(f"  Caso:          {caso}")
    print(f"  Datos entrada: {str(datos or '')[:200]}")
    print(f"  Pasos:         {str(pasos or '')[:200]}")
    print(f"  Esperado:      {str(esperado or '')[:200]}")
    print(f"  Criterio Acept:{str(crit or '')[:200]}")
    if riesgo: print(f"  Riesgo:        {str(riesgo)[:200]}")
    if hallazgo: print(f"  Hallazgo:      {str(hallazgo)[:200]}")
    print()
