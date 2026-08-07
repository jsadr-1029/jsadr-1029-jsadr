"""Lee el detalle del caso de prueba TC-CLI-001 desde el Excel."""
from openpyxl import load_workbook

WB_PATH = "/home/z/my-project/upload/plan-pruebas-qa-jsadr.xlsx"
wb = load_workbook(WB_PATH, data_only=True)
ws = wb["4. M02-Clientes"]

# Header row 4: ['', 'ID', 'Módulo', 'Función', 'Caso de Prueba', 'Tipo', 'Prioridad', 
#                'Precondiciones', 'Pasos', 'Datos de Entrada', 'Resultado Esperado', 
#                'Criterios de Aceptación', 'Estado']
headers = [str(ws.cell(row=4, column=c).value or '').strip() for c in range(1, ws.max_column + 1)]
print("HEADERS:")
for i, h in enumerate(headers, 1):
    print(f"  col {i}: {h}")

# TC-CLI-001 está en fila 5
print("\n" + "="*80)
print("TC-CLI-001 - Crear cliente válido")
print("="*80)
for c, h in enumerate(headers, 1):
    if not h:
        continue
    val = ws.cell(row=5, column=c).value
    print(f"\n[{h}]:")
    print(f"  {val}")
