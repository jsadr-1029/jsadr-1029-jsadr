"""
Generación del Plan de Pruebas QA — JSADR Plataforma de Gestión de Préstamos
Excel con casos de prueba por módulo, estándar QA Tester.
"""
import sys, os
XLSX_SKILL_DIR = "/home/z/my-project/skills/xlsx"
for sub in [XLSX_SKILL_DIR, os.path.join(XLSX_SKILL_DIR, "templates")]:
    if sub not in sys.path:
        sys.path.insert(0, sub)

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from copy import copy
from datetime import datetime
import base as B

# Usar paleta professional (azul corporativo)
B.use_palette_explicit("professional")

OUTPUT = "/home/z/my-project/download/plan-pruebas-qa-jsadr.xlsx"

# ============================================================
# Definición de casos de prueba por módulo
# Formato: (ID, Módulo, Función, Caso de Prueba, Tipo, Prioridad, Precondiciones, Pasos, Datos de Entrada, Resultado Esperado, Criterios de Aceptación, Estado)
# ============================================================

PRIORIDADES = {"Alta": "C0392B", "Media": "D4820A", "Baja": "1B7D46"}
ESTADOS = {"Pendiente": "8C8A84", "Aprobado": "1B7D46", "Fallido": "C0392B", "Bloqueado": "D4820A", "En Progreso": "1B2A4A"}
TIPOS = {"Funcional", "No Funcional", "Integración", "Seguridad", "UI/UX", "Performance", "Smoke", "Regresión"}

def caso(cid, mod, func, caso, tipo, prio, pre, pasos, datos, esperado, criterios, estado="Pendiente"):
    return {
        "ID": cid, "Módulo": mod, "Función": func, "Caso de Prueba": caso,
        "Tipo": tipo, "Prioridad": prio, "Precondiciones": pre,
        "Pasos": pasos, "Datos de Entrada": datos,
        "Resultado Esperado": esperado, "Criterios de Aceptación": criterios,
        "Estado": estado
    }

# ============================================================
# M01 — Autenticación
# ============================================================
M01 = [
    caso("TC-AUTH-001", "Autenticación", "Login Sistema", "Login admin exitoso", "Funcional", "Alta",
         "Servidor corriendo en localhost:3000", "1. Navegar a /login\n2. Ingresar username=Adm-Jsadr\n3. Ingresar password=Js951029*\n4. Click en Ingresar",
         "username: Adm-Jsadr, password: Js951029*",
         "Sistema responde 200 con access_token + refresh_token; usuario redirigido a /admin",
         "HTTP 200 + access_token JWT válido por 15min", "Aprobado"),
    caso("TC-AUTH-002", "Autenticación", "Login Sistema", "Login gestor exitoso", "Funcional", "Alta",
         "Servidor corriendo", "1. Navegar a /login\n2. Ingresar gestor-jsadr / Js951029*\n3. Click Ingresar",
         "username: gestor-jsadr, password: Js951029*",
         "HTTP 200 + token; redirige a /gestor", "Token con rol=GESTOR", "Aprobado"),
    caso("TC-AUTH-003", "Autenticación", "Login Sistema", "Login consultor exitoso", "Funcional", "Alta",
         "Servidor corriendo", "Login con consultor-jsadr / Js951029*", "consultor-jsadr / Js951029*",
         "HTTP 200 + token rol=CONSULTOR", "Acceso solo lectura", "Aprobado"),
    caso("TC-AUTH-004", "Autenticación", "Login Sistema", "Login abogado exitoso", "Funcional", "Alta",
         "Servidor corriendo", "Login con abogado-jsadr / Js951029*", "abogado-jsadr / Js951029*",
         "HTTP 200 + token rol=ABOGADO", "Acceso al portal jurídico", "Aprobado"),
    caso("TC-AUTH-005", "Autenticación", "Login Sistema", "Login con credenciales inválidas", "Seguridad", "Alta",
         "Servidor corriendo", "1. Navegar a /login\n2. Ingresar usuario inexistente\n3. Ingresar password",
         "username: noexiste, password: 123456",
         "HTTP 401 con mensaje genérico 'Credenciales incorrectas'. No revelar si el usuario existe",
         "Respuesta uniforme anti-enumeración", "Aprobado"),
    caso("TC-AUTH-006", "Autenticación", "Login Sistema", "Login con password incorrecta", "Seguridad", "Alta",
         "Usuario válido existe", "1. Login con Adm-Jsadr\n2. Password incorrecta (xxx)",
         "username: Adm-Jsadr, password: incorrecta",
         "HTTP 401. Incrementa intentosFallidos. A los 5 intentos bloquea por 15min",
         "intentosFallidos++ en BD", "Aprobado"),
    caso("TC-AUTH-007", "Autenticación", "Login Sistema", "Bloqueo por intentos fallidos", "Seguridad", "Alta",
         "Usuario con 4 intentos fallidos previos", "1. Ingresar password incorrecta por 5ta vez",
         "Quinto intento incorrecto",
         "HTTP 403 con mensaje 'Cuenta bloqueada temporalmente'. bloqueadoHasta = now+15min",
         "Usuario bloqueado, no puede login hasta expirar bloqueo", "Aprobado"),
    caso("TC-AUTH-008", "Autenticación", "Login Sistema", "Rate limiting por IP", "Seguridad", "Media",
         "Múltiples intentos rápidos", "1. Hacer 11 peticiones POST /api/auth/login en menos de 1 minuto",
         "11 requests en 60s",
         "A partir de la petición 11: HTTP 429 'Demasiados intentos de login. Espera 1 minuto.'",
         "Límite de 10 req/min por IP", "Aprobado"),
    caso("TC-AUTH-009", "Autenticación", "Recuperar Clave", "Recuperar clave por email", "Funcional", "Alta",
         "Usuario con email registrado", "1. POST /api/auth/recuperar-clave\n2. Body: {identificador: 'admin@jsadr.co'}",
         "identificador: admin@jsadr.co",
         "HTTP 200. Se genera password temporal. Se envía por correo. mustChangePassword=true",
         "Password temporal de 12 chars (4 may + 4 min + 3 num + 1 símb)", "Aprobado"),
    caso("TC-AUTH-010", "Autenticación", "Recuperar Clave", "Rate limit recuperación (1/5min)", "Seguridad", "Media",
         "Recuperación previa hace <5min", "1. Solicitar recuperación 2 veces en menos de 5 min",
         "2 solicitudes en <5min",
         "Segunda solicitud: HTTP 429 'Demasiadas solicitudes. Intenta en 5 minuto(s).'",
         "Rate limit 1 req/5min por IP", "Aprobado"),
    caso("TC-AUTH-011", "Autenticación", "Recuperar Clave", "Recuperar por cédula de cliente", "Funcional", "Media",
         "Cliente con cédula registrada", "POST /api/auth/recuperar-clave {identificador: '1214731649'}",
         "identificador: 1214731649",
         "HTTP 200. Se envía password temporal al email del cliente",
         "Email enviado a jsadr23@gmail.com", "Aprobado"),
    caso("TC-AUTH-012", "Autenticación", "MFA/OTP", "Login con MFA TOTP", "Seguridad", "Media",
         "Usuario con mfaEnabled=true y mfaSecret configurado", "1. Login step 1: credenciales\n2. Recibir requiresMFA=true\n3. Login step 2: OTP TOTP de Google Authenticator",
         "step=1 + credenciales, luego step=2 + otp TOTP",
         "Step 1: requiresMFA=true + tempToken. Step 2: access_token + refresh_token",
         "OTP TOTP RFC 6238 válido por 30s", "Pendiente"),
    caso("TC-AUTH-013", "Autenticación", "MFA/OTP", "OTP por WhatsApp", "Funcional", "Media",
         "Usuario con OTP_WHATSAPP_<email> en Configuracion", "1. Solicitar OTP WhatsApp\n2. Ingresar OTP recibido",
         "OTP de 6 dígitos",
         "Sistema valida OTP y expira el registro. Login exitoso",
         "OTP expira en 5 min", "Pendiente"),
    caso("TC-AUTH-014", "Autenticación", "Login Sistema", "JWT expira a los 15 min", "Seguridad", "Media",
         "Login exitoso hace 16 min", "1. Hacer una petición autenticada con access_token expirado",
         "Token con iat hace 16 min",
         "HTTP 401 'Token expirado'. Frontend debe refrescar con refresh_token",
         "Access token expira en 15min, refresh en 7d", "Aprobado"),
    caso("TC-AUTH-015", "Autenticación", "Login Sistema", "Refresh token exitoso", "Seguridad", "Media",
         "Access token expirado, refresh token válido", "1. POST /api/auth/refresh con refresh_token",
         "refresh_token válido",
         "HTTP 200 con nuevo access_token",
         "Nuevo access_token con iat reciente", "Aprobado"),
]

# ============================================================
# M02 — Clientes
# ============================================================
M02 = [
    caso("TC-CLI-001", "Clientes", "Crear Cliente", "Crear cliente válido", "Funcional", "Alta",
         "Admin autenticado", "1. POST /api/clientes con datos válidos",
         "nombre: Juan, cedula: 1000000001, telefono: 3000000001, email: juan@test.com",
         "HTTP 201. Cliente creado con id generado. cedula única",
         "Cliente aparece en listado", "Pendiente"),
    caso("TC-CLI-002", "Clientes", "Crear Cliente", "Crear cliente con cédula duplicada", "Funcional", "Alta",
         "Cliente con cédula 1214731649 ya existe", "1. POST /api/clientes con cedula=1214731649",
         "cedula: 1214731649 (duplicada)",
         "HTTP 409 'La cédula ya está registrada'",
         "No se crea registro duplicado", "Pendiente"),
    caso("TC-CLI-003", "Clientes", "Crear Cliente", "Crear cliente con email inválido", "Validación", "Media",
         "Admin autenticado", "POST /api/clientes con email: 'no-es-email'",
         "email: no-es-email",
         "HTTP 400 con fieldErrors.email",
         "Validación Zod en backend", "Pendiente"),
    caso("TC-CLI-004", "Clientes", "Crear Cliente", "Crear cliente con teléfono vacío", "Validación", "Baja",
         "Admin autenticado", "POST /api/clientes con telefono: ''",
         "telefono: ''",
         "HTTP 400 'Teléfono es requerido'",
         "Validación campo requerido", "Pendiente"),
    caso("TC-CLI-005", "Clientes", "Buscar Cliente", "Búsqueda por cédula exacta", "Funcional", "Alta",
         "Múltiples clientes en BD", "1. GET /api/clientes?cedula=1214731649",
         "cedula: 1214731649",
         "HTTP 200 con 1 cliente: JOHAN ALVAREZ",
         "Búsqueda exacta por índice único", "Pendiente"),
    caso("TC-CLI-006", "Clientes", "Buscar Cliente", "Búsqueda por nombre parcial", "Funcional", "Media",
         "Clientes con 'ALVAREZ' en nombre", "1. GET /api/clientes?q=ALVAREZ",
         "q: ALVAREZ",
         "HTTP 200 con clientes: JOHAN ALVAREZ, CAROLINA ALVAREZ",
         "Búsqueda case-insensitive", "Pendiente"),
    caso("TC-CLI-007", "Clientes", "Gestión PIN", "Cliente sin PIN primer acceso", "Funcional", "Alta",
         "Cliente con pinHash=null", "1. POST /api/portal/login con cédula y PIN nuevo",
         "cedula: 123456789, pin: 1234",
         "Sistema crea pinHash (bcrypt rounds=10). Login exitoso. Retorna nuevoPin=true",
         "pinHash creado en BD", "Aprobado"),
    caso("TC-CLI-008", "Clientes", "Gestión PIN", "PIN incorrecto bloquea tras 5 intentos", "Seguridad", "Alta",
         "Cliente con pinHash", "1. Ingresar PIN incorrecto 5 veces",
         "5 intentos con PIN incorrecto",
         "pinBloqueadoHasta = now + 15min. Login bloqueado",
         "pinIntentos se reinicia tras bloqueo expira", "Pendiente"),
    caso("TC-CLI-009", "Clientes", "Gestión Clave", "Cliente con clave alfanumérica", "Funcional", "Alta",
         "Cliente con claveHash", "1. POST /api/portal/login con cédula + clave",
         "cedula: 1214731649, clave: 1234",
         "HTTP 200 + token. Sesión persistida 2h",
         "claveHash verificado con bcrypt compare", "Aprobado"),
    caso("TC-CLI-010", "Clientes", "Activar/Inactivar", "Inactivar cliente", "Funcional", "Media",
         "Cliente activo", "1. PATCH /api/clientes/<id> {activo: false}",
         "activo: false",
         "HTTP 200. Cliente.inactivo=true. No puede hacer login al portal",
         "Login del portal rechaza con 'Cuenta inactiva'", "Pendiente"),
    caso("TC-CLI-011", "Clientes", "Listar Clientes", "Listar con paginación", "Funcional", "Media",
         "8 clientes en BD", "1. GET /api/clientes?page=1&pageSize=5",
         "page: 1, pageSize: 5",
         "HTTP 200 con 5 clientes + total=8",
         "Paginación correcta", "Pendiente"),
    caso("TC-CLI-012", "Clientes", "Editar Cliente", "Actualizar email", "Funcional", "Media",
         "Cliente existe", "1. PATCH /api/clientes/<id> {email: 'nuevo@test.com'}",
         "email: nuevo@test.com",
         "HTTP 200 con email actualizado",
         "Email único en BD", "Pendiente"),
    caso("TC-CLI-013", "Clientes", "Eliminar Cliente", "Eliminar con préstamos asociados", "Reglas Negocio", "Alta",
         "Cliente con préstamos", "1. DELETE /api/clientes/<id>",
         "id de cliente con préstamos",
         "HTTP 409 'No se puede eliminar: tiene préstamos asociados'",
         "Integridad referencial", "Pendiente"),
    caso("TC-CLI-014", "Clientes", "Validación Email", "Email duplicado entre clientes", "Validación", "Baja",
         "Email compartido", "POST /api/clientes con email ya existente",
         "email: jsadr23@gmail.com",
         "HTTP 409 (email único) o HTTP 201 (email no único según modelo)",
         "Verificar constraint en schema.prisma", "Pendiente"),
    caso("TC-CLI-015", "Clientes", "Tasa Personalizada", "Asignar tasa personalizada", "Reglas Negocio", "Media",
         "Cliente activo", "PATCH /api/clientes/<id> {tieneTasaPersonalizada: true, tasaPersonalizada: 2.5}",
         "tasaPersonalizada: 2.5% mensual",
         "HTTP 200. Préstamos futuros usan esta tasa",
         "Cálculo de préstamo usa tasaPersonalizada si tieneTasaPersonalizada=true", "Pendiente"),
]

# ============================================================
# M03 — Préstamos
# ============================================================
M03 = [
    caso("TC-PRE-001", "Préstamos", "Crear Préstamo", "Crear préstamo válido", "Funcional", "Alta",
         "Cliente y gestor autenticado", "1. POST /api/prestamos con monto, plazo, tasa, frecuencia",
         "montoPrincipal: 1,000,000, plazoMeses: 12, tasa: 24%, frecuencia: MENSUAL",
         "HTTP 201 con código autogenerado (PR-YYYY-NNNN). Estado=ACTIVO",
         "Código único autogenerado", "Pendiente"),
    caso("TC-PRE-002", "Préstamos", "Crear Préstamo", "Monto mínimo válido", "Validación", "Media",
         "Gestor autenticado", "POST /api/prestamos monto: 50,000",
         "montoPrincipal: 50,000 (mínimo)",
         "HTTP 201 creado correctamente",
         "Monto mínimo respetado", "Pendiente"),
    caso("TC-PRE-003", "Préstamos", "Crear Préstamo", "Monto inferior al mínimo", "Validación", "Alta",
         "Gestor autenticado", "POST /api/prestamos monto: 10,000",
         "montoPrincipal: 10,000",
         "HTTP 400 'Monto debe ser ≥ 50,000'",
         "Validación backend", "Pendiente"),
    caso("TC-PRE-004", "Préstamos", "Crear Préstamo", "Plazo inválido (0 meses)", "Validación", "Media",
         "Gestor autenticado", "POST /api/prestamos plazoMeses: 0",
         "plazoMeses: 0",
         "HTTP 400 'Plazo debe ser ≥ 1'",
         "Validación Zod", "Pendiente"),
    caso("TC-PRE-005", "Préstamos", "Cálculo Intereses", "Cálculo interés compuesto mensual", "Funcional", "Alta",
         "Préstamo creado", "1. GET /api/prestamos/<id>/calculo",
         "monto: 1,000,000, tasa: 24% anual, plazo: 12m, frecuencia: MENSUAL",
         "Interés total = 264,683 (aprox). Cuota mensual = 105,390",
         "Fórmula interés compuesto verificada", "Pendiente"),
    caso("TC-PRE-006", "Préstamos", "Cálculo Mora", "Cálculo de mora compuesta", "Funcional", "Alta",
         "Préstamo con cuotas vencidas", "1. GET /api/prestamos/<id>/calculo",
         "5 días de mora, tasa mora 2% diario",
         "Mora calculada con interés compuesto sobre saldo pendiente",
         "Mora = saldo × (1+tasa)^días - saldo", "Pendiente"),
    caso("TC-PRE-007", "Préstamos", "Estados", "Transición ACTIVO → EN_MORA", "Reglas Negocio", "Alta",
         "Préstamo ACTIVO con cuota vencida", "1. Job diario evalúa mora\n2. GET /api/prestamos/<id>",
         "Cuota vencida hace 1 día",
         "Estado cambia a EN_MORA. diasMora=1",
         "Job cron diario verifica vencimientos", "Pendiente"),
    caso("TC-PRE-008", "Préstamos", "Estados", "Transición EN_MORA → PAGADO", "Reglas Negocio", "Alta",
         "Préstamo EN_MORA, pago total recibido", "1. Registrar pago por saldo total",
         "Pago = saldo pendiente",
         "Estado cambia a PAGADO. fechaCierre = now",
         "Cierre automático al saldar", "Pendiente"),
    caso("TC-PRE-009", "Préstamos", "Estados", "Transición ACTIVO → ANULADO", "Reglas Negocio", "Media",
         "Préstamo ACTIVO sin pagos", "1. POST /api/prestamos/<id>/anular {motivo}",
         "motivo: 'Error de creación'",
         "Estado=ANULADO. Solo ADMIN puede anular. AuditLog registrado",
         "Solo ADMIN; bitácoraPrestamo actualizada", "Pendiente"),
    caso("TC-PRE-010", "Préstamos", "Firma Electrónica", "Firma con OTP exitosa", "Funcional", "Alta",
         "Préstamo creado, cliente autenticado en portal", "1. POST /api/portal/solicitar-otp\n2. POST /api/portal/validar-otp",
         "firmaId + OTP 6 dígitos recibido por email",
         "firmaElectronica.estadoFirma=COMPLETADA. otpValidado=true. hashDocumento firmado",
         "OTP hasheado SHA-256, comparación constant-time", "Aprobado"),
    caso("TC-PRE-011", "Préstamos", "Firma Electrónica", "Firma con OTP expirado", "Seguridad", "Alta",
         "OTP solicitado hace 6 min", "1. POST /api/portal/validar-otp con OTP expirado",
         "OTP con expiración >5min",
         "HTTP 400 'OTP expirado'. Solicitar nuevo OTP",
         "OTP expira en 5 min", "Pendiente"),
    caso("TC-PRE-012", "Préstamos", "Firma Electrónica", "Máximo 5 intentos OTP", "Seguridad", "Alta",
         "4 intentos fallidos previos", "1. Ingresar OTP incorrecto por 5ta vez",
         "5to OTP incorrecto",
         "HTTP 429 'Máximo de intentos alcanzado'. Firma bloqueada",
         "intentosOTP >= maxIntentos", "Pendiente"),
    caso("TC-PRE-013", "Préstamos", "Aceptar TyC", "Aceptar TyC con OTP", "Funcional", "Alta",
         "Préstamo pendiente TyC", "1. POST /api/prestamos/<id>/aceptar-tyc-otp",
         "firmaId + OTP 6 dígitos",
         "HTTP 200. tycAceptado=true. Fecha aceptación registrada",
         "BitácoraPrestamo actualizada", "Aprobado"),
    caso("TC-PRE-014", "Préstamos", "Listar Préstamos", "Filtrar por estado", "Funcional", "Media",
         "30 préstamos en BD", "1. GET /api/prestamos?estado=EN_MORA",
         "estado: EN_MORA",
         "HTTP 200 con solo préstamos EN_MORA",
         "Filtro por campo enum", "Pendiente"),
    caso("TC-PRE-015", "Préstamos", "Reversar Pago", "Reversar pago aplicado", "Reglas Negocio", "Alta",
         "Préstamo con pago aplicado", "1. POST /api/pagos/<id>/reversar {motivo}",
         "motivo: 'Pago duplicado'",
         "HTTP 200. Pago marcado como REVERSADO. Saldo recalculado. AuditLog registrado",
         "Solo ADMIN; bitácora actualizada", "Pendiente"),
]

# ============================================================
# M04 — Pagos
# ============================================================
M04 = [
    caso("TC-PAG-001", "Pagos", "Registrar Pago", "Pago completo de cuota", "Funcional", "Alta",
         "Préstamo ACTIVO con cuota pendiente", "1. POST /api/pagos {prestamoId, monto, fecha}",
         "monto: cuota exacta",
         "HTTP 201. Pago registrado. Saldo disminuido. Cuota marcada PAGADA",
         "SaldoActual = saldoAnterior - monto", "Pendiente"),
    caso("TC-PAG-002", "Pagos", "Registrar Pago", "Pago parcial", "Funcional", "Alta",
         "Préstamo ACTIVO", "POST /api/pagos monto < cuota",
         "monto: 50% de cuota",
         "HTTP 201. Pago registrado. Saldo disminuido. Cuota sigue PENDIENTE",
         "Saldo actualizado; cuota pendiente con nuevo saldo", "Pendiente"),
    caso("TC-PAG-003", "Pagos", "Registrar Pago", "Pago mayor a saldo (sobrepago)", "Validación", "Media",
         "Préstamo con saldo $100,000", "POST /api/pagos monto: $200,000",
         "monto: 200,000 > saldo 100,000",
         "HTTP 400 'Monto no puede exceder saldo pendiente' o sistema permite y genera nota crédito",
         "Política de negocio documentada", "Pendiente"),
    caso("TC-PAG-004", "Pagos", "Registrar Pago", "Pago a préstamo ANULADO", "Reglas Negocio", "Alta",
         "Préstamo estado=ANULADO", "POST /api/pagos prestamoId=<anulado>",
         "prestamoId de préstamo anulado",
         "HTTP 409 'No se pueden registrar pagos a préstamos anulados'",
         "Validación de estado", "Pendiente"),
    caso("TC-PAG-005", "Pagos", "Registrar Pago", "Pago con fecha futura", "Validación", "Media",
         "Gestor autenticado", "POST /api/pagos fecha: '2030-12-31'",
         "fecha: 2030-12-31",
         "HTTP 400 'Fecha no puede ser futura'",
         "Validación fecha ≤ now", "Pendiente"),
    caso("TC-PAG-006", "Pagos", "Reversar Pago", "Reversar pago válido", "Funcional", "Alta",
         "Pago existente aplicado", "1. POST /api/pagos/<id>/reversar {motivo}",
         "motivo: 'Error de registro'",
         "HTTP 200. Pago.estado=REVERSADO. Saldo restituido. AuditLog registrado",
         "Solo ADMIN; bitácora actualizada", "Pendiente"),
    caso("TC-PAG-007", "Pagos", "Reversar Pago", "Reversar pago ya reversado", "Reglas Negocio", "Alta",
         "Pago estado=REVERSADO", "POST /api/pagos/<id>/reversar",
         "id de pago ya reversado",
         "HTTP 409 'El pago ya está reversado'",
         "Validación de estado", "Pendiente"),
    caso("TC-PAG-008", "Pagos", "Anular Pago", "Anular pago con motivo", "Funcional", "Alta",
         "Pago aplicado", "1. POST /api/pagos/<id>/anular {motivo}",
         "motivo: 'Doble aplicación'",
         "HTTP 200. Pago.estado=ANULADO. Saldo restituido. AuditLog. Bitácora actualizada",
         "Solo ADMIN; pagoAnuladoPor registrado", "Pendiente"),
    caso("TC-PAG-009", "Pagos", "Anular Pago", "Anular pago reversado", "Reglas Negocio", "Alta",
         "Pago estado=REVERSADO", "POST /api/pagos/<id>/anular",
         "id de pago ya reversado",
         "HTTP 409 'No se puede anular pago reversado'",
         "Validación de estado previo", "Pendiente"),
    caso("TC-PAG-010", "Pagos", "Conciliación", "Conciliar pagos bancarios", "Integración", "Media",
         "Archivo de extracto bancario", "1. Upload CSV\n2. Sistema cruza con pagos registrados",
         "Archivo CSV extracto bancario",
         "Pagos conciliados marcados. Discrepancias reportadas",
         "Archivo procesado en cola asíncrona", "Pendiente"),
    caso("TC-PAG-011", "Pagos", "Validación Monto", "Pago con monto negativo", "Validación", "Media",
         "Gestor autenticado", "POST /api/pagos monto: -100000",
         "monto: -100000",
         "HTTP 400 'Monto debe ser positivo'",
         "Validación Zod > 0", "Pendiente"),
    caso("TC-PAG-012", "Pagos", "Validación Monto", "Pago con monto 0", "Validación", "Baja",
         "Gestor autenticado", "POST /api/pagos monto: 0",
         "monto: 0",
         "HTTP 400 'Monto debe ser > 0'",
         "Validación Zod > 0", "Pendiente"),
    caso("TC-PAG-013", "Pagos", "Listar Pagos", "Listar por préstamo", "Funcional", "Media",
         "Préstamo con 5 pagos", "1. GET /api/prestamos/<id>/pagos",
         "prestamoId específico",
         "HTTP 200 con array de pagos ordenados por fecha",
         "Incluye pagos REVERSADOS y ANULADOS", "Pendiente"),
    caso("TC-PAG-014", "Pagos", "Cálculo Descuento", "Descuento por pago anticipado", "Reglas Negocio", "Baja",
         "Préstamo con descuento anticipado configurado", "1. POST /api/pagos con fecha anticipada",
         "Pago anticipado de 3 cuotas",
         "Sistema calcula descuento y aplica al saldo",
         "Política de descuento documentada", "Pendiente"),
    caso("TC-PAG-015", "Pagos", "Caja", "Registrar movimiento de caja", "Funcional", "Media",
         "Caja abierta", "POST /api/caja/movimientos {tipo, monto, concepto}",
         "tipo: INGRESO, monto: 50000",
         "HTTP 201. Movimiento registrado. Saldo caja actualizado",
         "SaldoCaja consistente", "Pendiente"),
]

# ============================================================
# M05 — Correo Electrónico
# ============================================================
M05 = [
    caso("TC-MAIL-001", "Correo", "SMTP Conexión", "Verificar conexión SMTP", "Smoke", "Alta",
         "BREVO_SMTP_KEY y API_ENCRYPTION_KEY configurados", "1. POST /api/email {accion: 'probar'} como admin",
         "Token admin válido",
         "HTTP 200 success=true. Message: 'Conexión SMTP verificada correctamente con smtp-relay.brevo.com:587'",
         "ConexionAPI.EMAIL_SMTP.probada=true", "Aprobado"),
    caso("TC-MAIL-002", "Correo", "Envío Prueba", "Enviar correo de prueba", "Smoke", "Alta",
         "SMTP configurado", "1. POST /api/email {accion: 'enviar-prueba', to: 'test@test.com'}",
         "to: test@test.com",
         "HTTP 200 success=true. messageId devuelto por Brevo. isEthereal=false",
         "Correo real enviado vía Brevo", "Aprobado"),
    caso("TC-MAIL-003", "Correo", "Envío OTP", "Enviar OTP a cliente", "Funcional", "Alta",
         "Cliente con email, firma pendiente", "1. POST /api/portal/solicitar-otp {firmaId}",
         "firmaId válido",
         "HTTP 200. otpGenerado=true. emailEnviado=true. emailEnmascarado=j***@gmail.com",
         "OTP 6 dígitos enviado por correo", "Aprobado"),
    caso("TC-MAIL-004", "Correo", "Envío OTP", "Enviar OTP a cliente sin email", "Reglas Negocio", "Alta",
         "Cliente con email=null", "POST /api/portal/solicitar-otp {firmaId}",
         "firmaId de cliente sin email",
         "HTTP 400 'Tu cuenta no tiene un correo electrónico registrado'",
         "Validación previa al envío", "Pendiente"),
    caso("TC-MAIL-005", "Correo", "Recuperar Clave", "Enviar password temporal", "Funcional", "Alta",
         "Usuario con email", "1. POST /api/auth/recuperar-clave {identificador}",
         "identificador: admin@jsadr.co",
         "HTTP 200. Email enviado con password temporal. mustChangePassword=true",
         "Password temporal 12 chars aleatorios", "Aprobado"),
    caso("TC-MAIL-006", "Correo", "Notificación Masiva", "Disparar recordatorios de pago", "Funcional", "Alta",
         "Préstamos activos con cuotas próximas a vencer", "1. POST /api/notificaciones {accion: 'recordatorios'} como admin",
         "Token admin",
         "HTTP 200 success=true. enviadas=N. fallidas=M. Detalles por préstamo",
         "Log de cada envío en NotificacionLog", "Aprobado"),
    caso("TC-MAIL-007", "Correo", "Notificación Masiva", "Disparar notificaciones de mora", "Funcional", "Alta",
         "Préstamos EN_MORA", "1. POST /api/notificaciones {accion: 'mora'} como admin",
         "Token admin",
         "HTTP 200. Notificaciones de mora enviadas a préstamos en mora",
         "DiasMora > 0 en préstamos notificados", "Aprobado"),
    caso("TC-MAIL-008", "Correo", "Fallback SMTP", "Fallback a SMTP cuando HTTPS API falla", "Integración", "Media",
         "BREVO_API_KEY inválida, BREVO_SMTP_KEY válida", "1. Disparar envío de correo",
         "API key inválida, SMTP key válida",
         "Sistema intenta HTTPS API (falla), luego SMTP fallback (éxito). Log 'via: BREVO_SMTP'",
         "Doble camino implementado en src/lib/email.ts", "Aprobado"),
    caso("TC-MAIL-009", "Correo", "Modo Ethereal", "Sin SMTP configurado → Ethereal", "Seguridad", "Media",
         "ConexionAPI inactiva, sin BREVO_*", "1. POST /api/email {accion: 'enviar-prueba'}",
         "Sin credenciales SMTP",
         "HTTP 200 success=true. isEthereal=true. previewUrl devuelta",
         "Modo desarrollo no envía correos reales", "Pendiente"),
    caso("TC-MAIL-010", "Correo", "Validación destinatario", "Email destinatario inválido", "Validación", "Media",
         "Admin autenticado", "POST /api/email {accion: 'enviar-prueba', to: 'no-es-email'}",
         "to: no-es-email",
         "HTTP 400 con validación de email",
         "Validación previa al envío", "Pendiente"),
    caso("TC-MAIL-011", "Correo", "Cifrado Credenciales", "API_ENCRYPTION_KEY protege credenciales", "Seguridad", "Alta",
         "BD con ConexionAPI cifrada", "1. Verificar ConexionAPI.EMAIL_SMTP.password en BD",
         "Lectura directa BD",
         "Password en BD es IV:ciphertext (AES-256-CBC). Solo con API_ENCRYPTION_KEY se desencripta",
         "Sin API_ENCRYPTION_KEY, no se puede desencriptar", "Aprobado"),
    caso("TC-MAIL-012", "Correo", "Rate Limit", "Rate limit recuperación (1/5min)", "Seguridad", "Media",
         "Recuperación previa <5min", "POST /api/auth/recuperar-clave 2 veces en 5min",
         "2 solicitudes <5min",
         "HTTP 429 'Demasiadas solicitudes. Intenta en 5 minuto(s).'",
         "Rate limit 1 req/5min por IP", "Aprobado"),
    caso("TC-MAIL-013", "Correo", "Plantilla HTML", "Correo con plantilla HTML profesional", "UI/UX", "Baja",
         "SMTP configurado", "POST /api/email {accion: 'enviar-prueba', to}",
         "to válido",
         "Correo recibido con plantilla HTML responsive, gradientes, sin texto plano",
         "Plantilla HTML aplicada", "Aprobado"),
    caso("TC-MAIL-014", "Correo", "Trazabilidad", "Log de cada correo enviado", "Seguridad", "Media",
         "Envío de correo", "1. Enviar correo\n2. Verificar NotificacionLog",
         "Correo de prueba",
         "Registro en NotificacionLog con destinatario, asunto, estado, fechaEnvio",
         "Trazabilidad completa en BD", "Pendiente"),
    caso("TC-MAIL-015", "Correo", "Error Handling", "Manejo de error SMTP (535 auth failed)", "Funcional", "Media",
         "Credenciales SMTP inválidas", "1. POST /api/email {accion: 'probar'}",
         "Credenciales inválidas",
         "HTTP 200 success=false. Error sanitizado devuelto al cliente",
         "Error interno no expone detalles al cliente", "Pendiente"),
]

# ============================================================
# M06 — Seguridad
# ============================================================
M06 = [
    caso("TC-SEC-001", "Seguridad", "RBAC", "ADMIN puede acceder a todo", "Seguridad", "Alta",
         "Admin autenticado", "1. Hacer peticiones a endpoints /api/admin/*",
         "Token rol=ADMIN",
         "HTTP 200 en todos los endpoints administrativos",
         "requireRole(['ADMIN']) pasa", "Aprobado"),
    caso("TC-SEC-002", "Seguridad", "RBAC", "CONSULTOR no puede mutar", "Seguridad", "Alta",
         "Consultor autenticado", "1. POST /api/clientes con token consultor",
         "Token rol=CONSULTOR",
         "HTTP 403 'No autorizado'. requireRole rechaza",
         "Solo lectura para CONSULTOR", "Pendiente"),
    caso("TC-SEC-003", "Seguridad", "RBAC", "GESTOR no puede anular pagos", "Seguridad", "Alta",
         "Gestor autenticado", "1. POST /api/pagos/<id>/reversar con token gestor",
         "Token rol=GESTOR",
         "HTTP 403 'Solo ADMIN puede reversar pagos'",
         "requireRole(['ADMIN']) rechaza GESTOR", "Pendiente"),
    caso("TC-SEC-004", "Seguridad", "Audit Log", "AuditLog es inmutable", "Seguridad", "Alta",
         "AuditLog existe en BD", "1. Intentar delete/update/deleteMany/updateMany sobre auditLog",
         "Cualquier mutación AuditLog",
         "Error 'AuditLog es inmutable: no se permite delete/update'",
         "Prisma $extends bloquea mutaciones", "Aprobado"),
    caso("TC-SEC-005", "Seguridad", "Cifrado", "AES-256-CBC para credenciales", "Seguridad", "Alta",
         "API_ENCRYPTION_KEY configurada", "1. encryptSensitive('test')\n2. decryptSensitive(resultado)",
         "Texto plano 'test'",
         "Roundtrip exitoso. Output es IV:ciphertext hex. Longitud varía",
         "AES-256-CBC con IV aleatorio", "Aprobado"),
    caso("TC-SEC-006", "Seguridad", "Hash Password", "Bcrypt rounds=12", "Seguridad", "Alta",
         "Password hasheada", "1. bcrypt.hash('test', 12)\n2. bcrypt.compare('test', hash)",
         "Password 'test'",
         "Hash con prefijo $2b$12$. Verificación OK",
         "Bcrypt rounds=12 (recomendado OWASP)", "Aprobado"),
    caso("TC-SEC-007", "Seguridad", "Hash PIN", "Bcrypt rounds=10 para PIN", "Seguridad", "Media",
         "PIN hasheado", "1. bcrypt.hash('1234', 10)\n2. bcrypt.compare('1234', hash)",
         "PIN '1234'",
         "Hash con prefijo $2b$10$. Verificación OK",
         "Rounds menor por ser PIN de 4 dígitos", "Aprobado"),
    caso("TC-SEC-008", "Seguridad", "JWT", "JWT firmado con JWT_SECRET", "Seguridad", "Alta",
         "Login exitoso", "1. Decodificar access_token JWT en jwt.io",
         "access_token JWT",
         "Payload tiene userId, username, rol, type=access, iat, exp. Firma HS256 válida",
         "JWT_SECRET > 32 chars en .env", "Aprobado"),
    caso("TC-SEC-009", "Seguridad", "Sanitización Errores", "Errores no exponen stack interno", "Seguridad", "Alta",
         "Provocar error 500", "1. Hacer petición que cause excepción en backend",
         "Petición con datos que causan excepción",
         "HTTP 500 con mensaje genérico. Stack trace NO expuesto al cliente",
         "sanitizeError() activo en todas las routes", "Pendiente"),
    caso("TC-SEC-010", "Seguridad", "Rate Limiting", "Rate limit por IP en login", "Seguridad", "Alta",
         "Múltiples intentos rápidos", "1. 11 peticiones POST /api/auth/login en 1 min",
         "11 requests / 60s",
         "HTTP 429 a partir de la 11ava. Mensaje 'Demasiados intentos'",
         "Límite 10 req/min por IP", "Aprobado"),
    caso("TC-SEC-011", "Seguridad", "Bloqueo Cuenta", "Bloqueo tras 5 intentos fallidos", "Seguridad", "Alta",
         "Usuario válido", "1. 5 intentos fallidos consecutivos",
         "5 passwords incorrectas",
         "Usuario.bloqueadoHasta = now + 15min. Login rechazado con 403",
         "intentosFallidos se reinicia tras éxito", "Aprobado"),
    caso("TC-SEC-012", "Seguridad", "OTP Hash", "OTP hasheado SHA-256 en BD", "Seguridad", "Alta",
         "OTP generado", "1. Solicitar OTP\n2. Verificar firmaElectronica.otpCodigo en BD",
         "OTP generado",
         "BD almacena SHA-256(otp), no texto plano. Comparación constant-time",
         "verificarOtp usa crypto.timingSafeEqual", "Aprobado"),
    caso("TC-SEC-013", "Seguridad", "TOTP", "TOTP RFC 6238 válido", "Seguridad", "Media",
         "mfaSecret configurado", "1. Generar TOTP con Google Authenticator\n2. Login step 2",
         "OTP TOTP 6 dígitos",
         "verifyTOTP valida con ventana de 30s",
         "TOTP propio implementado en src/lib/totp.ts", "Pendiente"),
    caso("TC-SEC-014", "Seguridad", "Eliminar Credenciales", "Eliminación con clave maestra", "Seguridad", "Alta",
         "Admin autenticado", "1. POST /api/seguridad/credenciales/eliminar {claveMaestra: 'Eliminar'}",
         "claveMaestra: 'Eliminar'",
         "HTTP 200. Credenciales Brevo eliminadas de BD y Vercel. Sistema sigue con SMTP fallback",
         "Doble eliminación BD + Vercel", "Pendiente"),
    caso("TC-SEC-015", "Seguridad", "CORS", "CORS estricto", "Seguridad", "Media",
         "Petición desde dominio no permitido", "1. Hacer CORS preflight desde evil.com",
         "Origin: evil.com",
         "HTTP 403 CORS. Headers Access-Control-Allow-Origin no devueltos",
         "Solo dominios JSADR permitidos", "Pendiente"),
]

# ============================================================
# M07 — Portal Cliente
# ============================================================
M07 = [
    caso("TC-PORT-001", "Portal Cliente", "Login PIN", "Login con cédula + PIN válido", "Funcional", "Alta",
         "Cliente con pinHash en BD", "1. POST /api/portal/login {cedula, pin}",
         "cedula: 1214731649, pin: 1234",
         "HTTP 200. Token sesión 2h. Datos cliente devueltos",
         "tokenSesion persistido en cliente.tokenSesion", "Aprobado"),
    caso("TC-PORT-002", "Portal Cliente", "Login PIN", "Login con cédula inexistente", "Seguridad", "Alta",
         "Cédula no registrada", "POST /api/portal/login {cedula: 9999999999, pin: 1234}",
         "cedula: 9999999999",
         "HTTP 404 'Cuenta no encontrada'",
         "No revela si cédula existe sino solo 'no encontrada'", "Aprobado"),
    caso("TC-PORT-003", "Portal Cliente", "Login PIN", "Login con PIN incorrecto", "Seguridad", "Alta",
         "Cliente con pinHash", "POST /api/portal/login {cedula, pin: '0000'}",
         "pin: 0000 (incorrecto)",
         "HTTP 401 'PIN incorrecto'. pinIntentos++. Bloqueo a los 5",
         "pinIntentos++ en BD", "Pendiente"),
    caso("TC-PORT-004", "Portal Cliente", "Login PIN", "Cliente inactivo no puede login", "Reglas Negocio", "Media",
         "Cliente activo=false", "POST /api/portal/login {cedula, pin}",
         "cedula de cliente inactivo",
         "HTTP 403 'Cuenta inactiva'",
         "Validación previa al check de PIN", "Pendiente"),
    caso("TC-PORT-005", "Portal Cliente", "Login Clave", "Login con cédula + clave alfanumérica", "Funcional", "Alta",
         "Cliente con claveHash", "POST /api/portal/login {cedula, clave}",
         "cedula: 1214731649, clave: 1234",
         "HTTP 200 + token. Sesión 2h",
         "claveHash verificada con bcrypt", "Aprobado"),
    caso("TC-PORT-006", "Portal Cliente", "Sesión", "Sesión expira a las 2h", "Seguridad", "Media",
         "Login exitoso hace 2h+", "1. Petición a /api/portal/* con token expirado",
         "Token con tokenExpira < now",
         "HTTP 401 'Sesión expirada'. Cliente debe reautenticarse",
         "tokenExpira verificado en cada request", "Pendiente"),
    caso("TC-PORT-007", "Portal Cliente", "Solicitar OTP", "Solicitar OTP para firma", "Funcional", "Alta",
         "Cliente autenticado en portal, firma pendiente", "POST /api/portal/solicitar-otp {firmaId}",
         "firmaId válido",
         "HTTP 200. OTP generado y enviado por email. emailEnviado=true",
         "OTP 6 dígitos, expira en 5 min", "Aprobado"),
    caso("TC-PORT-008", "Portal Cliente", "Validar OTP", "Validar OTP correcto", "Funcional", "Alta",
         "OTP solicitado, no expirado", "POST /api/portal/validar-otp {firmaId, otp}",
         "OTP correcto recibido por email",
         "HTTP 200. firma.estadoFirma=COMPLETADA. otpValidado=true",
         "Comparación constant-time SHA-256", "Aprobado"),
    caso("TC-PORT-009", "Portal Cliente", "Validar OTP", "Validar OTP incorrecto", "Seguridad", "Alta",
         "OTP solicitado", "POST /api/portal/validar-otp {firmaId, otp: '000000'}",
         "OTP incorrecto",
         "HTTP 401. intentosOTP++. A los 5: bloqueo de firma",
         "intentosOTP++ en BD", "Pendiente"),
    caso("TC-PORT-010", "Portal Cliente", "Clave Dinámica", "Solicitar clave dinámica simulador", "Funcional", "Alta",
         "Cliente autenticado, sesión válida", "POST /api/portal/clave-dinamica/solicitar {clienteId, token}",
         "token sesión válido",
         "HTTP 200. OTP generado y enviado por email. tipo=SOLICITUD_SIMULADOR",
         "Rate limit 1/60s por cliente", "Aprobado"),
    caso("TC-PORT-011", "Portal Cliente", "Clave Dinámica", "Validar clave dinámica", "Funcional", "Alta",
         "Clave dinámica generada", "POST /api/portal/clave-dinamica/validar",
         "OTP 6 dígitos",
         "HTTP 200 con codigoConfirmacion de un solo uso. Estado OTP=USADO",
         "codigoConfirmacion usado por /api/solicitudes-web", "Pendiente"),
    caso("TC-PORT-012", "Portal Cliente", "Estado Cuenta", "Ver estado de cuenta", "Funcional", "Media",
         "Cliente con préstamos", "1. GET /api/portal/mi-estado",
         "Token sesión válido",
         "HTTP 200 con préstamos, saldos, próximos vencimientos",
         "Solo datos del cliente autenticado", "Pendiente"),
    caso("TC-PORT-013", "Portal Cliente", "Solicitudes Web", "Crear solicitud de crédito", "Funcional", "Alta",
         "Cliente con codigoConfirmación", "1. POST /api/solicitudes-web {codigoConfirmacion, monto, plazo}",
         "Solicitud de crédito completa",
         "HTTP 201. Solicitud creada con estado PENDIENTE",
         "codigoConfirmación consumido", "Pendiente"),
    caso("TC-PORT-014", "Portal Cliente", "Logout", "Cerrar sesión", "Funcional", "Baja",
         "Cliente autenticado", "1. DELETE /api/portal/login con token",
         "Token sesión válido",
         "HTTP 200. tokenSesion=null. tokenExpira=null en BD",
         "Sesión invalidada en BD", "Pendiente"),
    caso("TC-PORT-015", "Portal Cliente", "Acceso Cross-Cliente", "Cliente A no puede ver datos de B", "Seguridad", "Alta",
         "Dos clientes autenticados", "1. Cliente A intenta GET /api/portal/cliente/<id_B>",
         "Token cliente A, id cliente B",
         "HTTP 403. No autorizado para ver datos de otro cliente",
         "Validación token vs clienteId", "Pendiente"),
]

# ============================================================
# M08 — Portal Jurídico
# ============================================================
M08 = [
    caso("TC-JUR-001", "Portal Jurídico", "Login", "Login con cédula numérica", "Funcional", "Alta",
         "Abogado con cédula 1234567890", "1. POST /api/juridico/portal/auth {cedula, clave}",
         "cedula: 1234567890, clave: Js951029*",
         "HTTP 200. Token sesión 8h. Rol ABOGADO verificado",
         "Usuario con rol IN [ABOGADO, GESTOR]", "Aprobado"),
    caso("TC-JUR-002", "Portal Jurídico", "Login", "Login con username (JD_jsadr)", "Funcional", "Alta",
         "Abogado con username JD_jsadr", "POST /api/juridico/portal/auth {cedula: 'JD_jsadr', clave}",
         "cedula: JD_jsadr, clave: 731649",
         "HTTP 200 + token. Login acepta cédula o username",
         "Búsqueda en cedula O username", "Aprobado"),
    caso("TC-JUR-003", "Portal Jurídico", "Login", "Login con clave incorrecta", "Seguridad", "Alta",
         "Abogado existe", "POST /api/juridico/portal/auth {cedula, clave: 'incorrecta'}",
         "clave incorrecta",
         "HTTP 401. Anti-enumeración: misma respuesta que si usuario no existe",
         "Respuesta uniforme", "Aprobado"),
    caso("TC-JUR-004", "Portal Jurídico", "Login", "Usuario sin rol ABOGADO/GESTOR", "Seguridad", "Alta",
         "Admin intenta login al portal jurídico", "POST /api/juridico/portal/auth {cedula: 'admin@jsadr.co', clave}",
         "Admin con rol ADMIN",
         "HTTP 401. Solo ABOGADO/GESTOR pueden acceder al portal jurídico",
         "Filtro rol IN [ABOGADO, GESTOR]", "Aprobado"),
    caso("TC-JUR-005", "Portal Jurídico", "Sesión", "Sesión expira a las 8h", "Seguridad", "Media",
         "Login exitoso hace 8h+", "1. Petición con token expirado",
         "Token con expiración >8h",
         "HTTP 401 'Sesión expirada'",
         "SESSION_EXPIRY_HOURS=8", "Pendiente"),
    caso("TC-JUR-006", "Portal Jurídico", "Logout", "Cerrar sesión portal jurídico", "Funcional", "Baja",
         "Abogado autenticado", "1. DELETE /api/juridico/portal/auth",
         "Token sesión válido",
         "HTTP 200. tokenSesion=null. Logout exitoso",
         "Sesión invalidada en BD", "Pendiente"),
    caso("TC-JUR-007", "Portal Jurídico", "Casos", "Listar casos asignados", "Funcional", "Media",
         "Abogado autenticado", "1. GET /api/juridico/casos",
         "Token sesión válido",
         "HTTP 200 con casos asignados al abogado",
         "Solo casos del abogado autenticado", "Pendiente"),
    caso("TC-JUR-008", "Portal Jurídico", "Casos", "Ver detalle de caso", "Funcional", "Media",
         "Caso asignado al abogado", "1. GET /api/juridico/casos/<id>",
         "Token sesión válido",
         "HTTP 200 con detalle completo del caso",
         "Solo si caso está asignado al abogado", "Pendiente"),
    caso("TC-JUR-009", "Portal Jurídico", "Documentos", "Subir documento legal", "Funcional", "Media",
         "Caso asignado", "1. POST /api/juridico/casos/<id>/documentos (multipart)",
         "Archivo PDF",
         "HTTP 201. Documento guardado. Trazabilidad en BD",
         "Archivo almacenado en /uploads/juridico/", "Pendiente"),
    caso("TC-JUR-010", "Portal Jurídico", "Casos", "Abogado no puede ver caso ajeno", "Seguridad", "Alta",
         "Dos abogados autenticados", "1. Abogado A intenta GET caso de abogado B",
         "Token abogado A, caso de abogado B",
         "HTTP 403. No autorizado",
         "Validación de asignación", "Pendiente"),
    caso("TC-JUR-011", "Portal Jurídico", "Notas Internas", "Agregar nota interna", "Funcional", "Media",
         "Caso asignado", "1. POST /api/juridico/casos/<id>/notas",
         "Texto de nota",
         "HTTP 201. Nota creada con autor=abogado actual",
         "Notas internas visibles solo al equipo jurídico", "Pendiente"),
    caso("TC-JUR-012", "Portal Jurídico", "Bitácora", "Ver bitácora del caso", "Funcional", "Media",
         "Caso con movimientos", "1. GET /api/juridico/casos/<id>/bitacora",
         "Token sesión válido",
         "HTTP 200 con movimientos ordenados por fecha",
         "Solo bitácora del caso asignado", "Pendiente"),
    caso("TC-JUR-013", "Portal Jurídico", "Rate Limit", "Rate limit login jurídico (20/min)", "Seguridad", "Media",
         "Múltiples intentos rápidos", "1. 21 peticiones POST /api/juridico/portal/auth en 1 min",
         "21 requests / 60s",
         "HTTP 429 a partir de la 21ava 'Demasiadas solicitudes'",
         "Límite 20 req/min por IP", "Pendiente"),
    caso("TC-JUR-014", "Portal Jurídico", "Token", "Token de sesión es randomBytes seguro", "Seguridad", "Alta",
         "Login exitoso", "1. Verificar tokenSesion en BD",
         "Token generado",
         "tokenSesion es crypto.randomBytes(32).toString('hex'). 64 chars hex",
         "Token aleatorio criptográficamente seguro", "Aprobado"),
    caso("TC-JUR-015", "Portal Jurídico", "Auditoría", "Cada acción registrada en AuditLog", "Seguridad", "Media",
         "Abogado autenticado", "1. Realizar acción\n2. Verificar AuditLog",
         "Cualquier acción del portal",
         "Registro en AuditLog con usuarioId, accion, ip, userAgent, fecha",
         "registrarAuditLog() en cada endpoint", "Pendiente"),
]

# ============================================================
# M09 — Notificaciones
# ============================================================
M09 = [
    caso("TC-NOT-001", "Notificaciones", "Recordatorios", "Disparar recordatorios manualmente", "Funcional", "Alta",
         "Préstamos activos con cuotas próximas", "1. POST /api/notificaciones {accion: 'recordatorios'}",
         "Token admin",
         "HTTP 200. enviadas=N. fallidas=M. Detalle por préstamo",
         "NotificacionLog registrado por envío", "Aprobado"),
    caso("TC-NOT-002", "Notificaciones", "Mora", "Disparar notificaciones de mora", "Funcional", "Alta",
         "Préstamos EN_MORA", "1. POST /api/notificaciones {accion: 'mora'}",
         "Token admin",
         "HTTP 200. Notificaciones enviadas a préstamos en mora",
         "DiasMora > 0", "Aprobado"),
    caso("TC-NOT-003", "Notificaciones", "WhatsApp", "Enviar WhatsApp Cloud API", "Integración", "Media",
         "WHATSAPP_TOKEN configurado", "1. enviarWhatsApp({destino, mensaje})",
         "Número destino, mensaje",
         "Mensaje enviado vía Meta WhatsApp Cloud API. wamid devuelto",
         "API de Meta integrada", "Pendiente"),
    caso("TC-NOT-004", "Notificaciones", "WhatsApp", "Mensaje recordatorio pago", "Funcional", "Media",
         "Préstamo con pago próximo", "1. mensajeRecordatorioPago(prestamo)",
         "Préstamo válido",
         "Mensaje formateado: 'Hola {nombre}, tu pago de ${monto} vence el {fecha}'",
         "Plantilla de mensaje aplicada", "Pendiente"),
    caso("TC-NOT-005", "Notificaciones", "WhatsApp", "Mensaje mora", "Funcional", "Media",
         "Préstamo EN_MORA", "1. mensajeMora(prestamo)",
         "Préstamo en mora",
         "Mensaje formateado: 'Tu préstamo está en mora por {dias} días'",
         "Incluye días de mora y monto", "Pendiente"),
    caso("TC-NOT-006", "Notificaciones", "Email", "Notificación por email", "Integración", "Alta",
         "Cliente con email", "1. enviarEmail({to, subject, html})",
         "Email destino, asunto, html",
         "Correo enviado vía Brevo. messageId devuelto",
         "Usa src/lib/email.ts enviarEmail()", "Aprobado"),
    caso("TC-NOT-007", "Notificaciones", "Log", "Guardar notificación en NotificacionLog", "Funcional", "Media",
         "Notificación enviada", "1. guardarNotificacion({prestamoId, tipo, estado, mensaje})",
         "Datos de notificación",
         "Registro en NotificacionLog con fecha, tipo, estado",
         "Trazabilidad completa", "Pendiente"),
    caso("TC-NOT-008", "Notificaciones", "Listar", "Listar notificaciones con filtros", "Funcional", "Media",
         "NotificacionLog con registros", "1. GET /api/notificaciones?tipo=recordatorio&estado=enviado",
         "Filtros tipo y estado",
         "HTTP 200 con notificaciones filtradas. Take=100",
         "Paginación por take=100", "Pendiente"),
    caso("TC-NOT-009", "Notificaciones", "Reenviar", "Reenviar notificación fallida", "Funcional", "Media",
         "NotificacionLog con estado=fallido", "1. POST /api/notificaciones/<id>/enviar",
         "id de notificación fallida",
         "HTTP 200. Notificación reenviada. Estado actualizado",
         "Estado cambia a 'enviado' si éxito", "Pendiente"),
    caso("TC-NOT-010", "Notificaciones", "Programación", "Job cron de notificaciones automáticas", "Funcional", "Media",
         "Sistema en producción", "1. Verificar ejecución del job cron diario",
         "Job programado",
         "Job corre diariamente y envía recordatorios/mora automáticamente",
         "Cron configurado en Vercel/CI", "Pendiente"),
    caso("TC-NOT-011", "Notificaciones", "Plantillas", "Plantilla HTML de recordatorio", "UI/UX", "Baja",
         "Cliente con email", "1. Enviar notificación recordatorio",
         "Cliente válido",
         "Email con plantilla HTML profesional: header con gradiente, body, CTA al portal",
         "Plantilla responsive", "Aprobado"),
    caso("TC-NOT-012", "Notificaciones", "Deduplicación", "No enviar duplicados en 24h", "Reglas Negocio", "Media",
         "Notificación ya enviada hoy", "1. Disparar misma notificación 2 veces en 24h",
         "Misma notificación 2 veces",
         "Sistema detecta duplicado y no reenvía. Log de skip",
         "Deduplicación por tipo+prestamoId+fecha", "Pendiente"),
    caso("TC-NOT-013", "Notificaciones", "RBAC", "Solo ADMIN dispara masivas", "Seguridad", "Alta",
         "Gestor autenticado", "1. POST /api/notificaciones {accion} con token gestor",
         "Token rol=GESTOR",
         "HTTP 403 'Solo ADMIN puede disparar notificaciones masivas'",
         "requireRole(['ADMIN'])", "Aprobado"),
    caso("TC-NOT-014", "Notificaciones", "Fallback", "Fallback WhatsApp → Email", "Integración", "Media",
         "Cliente sin WhatsApp válido", "1. Enviar notificación a cliente sin WhatsApp",
         "Cliente sin teléfono",
         "Sistema intenta WhatsApp (falla), luego Email",
         "Doble canal implementado", "Pendiente"),
    caso("TC-NOT-015", "Notificaciones", "Opt-out", "Cliente desuscribe notificaciones", "Reglas Negocio", "Baja",
         "Cliente con opt-out=true", "1. Enviar notificación a cliente con opt-out",
         "Cliente desuscrito",
         "Sistema respeta opt-out y no envía",
         "Campo optOutNotificaciones en Cliente", "Pendiente"),
]

# ============================================================
# M10 — Reportes
# ============================================================
M10 = [
    caso("TC-REP-001", "Reportes", "Cartera", "Generar reporte de cartera", "Funcional", "Alta",
         "Préstamos en BD", "1. GET /api/reportes/cartera",
         "Token admin",
         "HTTP 200 con resumen: total cartera, en mora, al día, porcentaje mora",
         "Cálculos financieros correctos", "Pendiente"),
    caso("TC-REP-002", "Reportes", "Morosidad", "Reporte de morosidad por rango", "Funcional", "Alta",
         "Préstamos EN_MORA", "1. GET /api/reportes/morosidad?desde=2026-01-01&hasta=2026-08-31",
         "Rango de fechas",
         "HTTP 200 con morosidad por día/semana/mes en el rango",
         "Filtros por fecha aplicados", "Pendiente"),
    caso("TC-REP-003", "Reportes", "Estados Financieros", "Balance de cartera", "Funcional", "Alta",
         "Datos completos", "1. GET /api/reportes/balance",
         "Token admin",
         "HTTP 200 con balance: capital prestado, intereses, mora, pagos recibidos",
         "Totales financieros correctos", "Pendiente"),
    caso("TC-REP-004", "Reportes", "Pagos", "Reporte de pagos por período", "Funcional", "Media",
         "Pagos registrados", "1. GET /api/reportes/pagos?desde=2026-01-01",
         "Desde fecha",
         "HTTP 200 con pagos del período, total, count",
         "Incluye pagos reversados/anulados por separado", "Pendiente"),
    caso("TC-REP-005", "Reportes", "Clientes", "Reporte de clientes activos", "Funcional", "Media",
         "Clientes en BD", "1. GET /api/reportes/clientes-activos",
         "Token admin",
         "HTTP 200 con clientes activos, # préstamos, # pagos",
         "Join Cliente → Préstamo → Pago", "Pendiente"),
    caso("TC-REP-006", "Reportes", "Exportación", "Exportar a Excel", "Funcional", "Media",
         "Reporte generado", "1. GET /api/reportes/cartera?format=xlsx",
         "format: xlsx",
         "HTTP 200 con Content-Type: application/vnd.openxmlformats. Descarga .xlsx",
         "Archivo Excel descargable", "Pendiente"),
    caso("TC-REP-007", "Reportes", "Exportación", "Exportar a PDF", "Funcional", "Media",
         "Reporte generado", "1. GET /api/reportes/cartera?format=pdf",
         "format: pdf",
         "HTTP 200 con Content-Type: application/pdf. Descarga .pdf",
         "Archivo PDF descargable", "Pendiente"),
    caso("TC-REP-008", "Reportes", "Filtros", "Filtrar por gestor", "Funcional", "Media",
         "Préstamos con gestor asignado", "1. GET /api/reportes/cartera?gestorId=<id>",
         "gestorId específico",
         "HTTP 200 con cartera del gestor específico",
         "Filtro por gestorId", "Pendiente"),
    caso("TC-REP-009", "Reportes", "Filtros", "Filtrar por período", "Funcional", "Media",
         "Préstamos por fecha", "1. GET /api/reportes/cartera?desde=2026-01-01&hasta=2026-06-30",
         "Rango fechas 6 meses",
         "HTTP 200 con cartera en el rango",
         "Filtros por fecha aplicados", "Pendiente"),
    caso("TC-REP-010", "Reportes", "Gráficos", "Datos para gráfico de morosidad", "Funcional", "Baja",
         "Datos morosidad", "1. GET /api/reportes/morosidad-grafico",
         "Token admin",
         "HTTP 200 con datos estructurados para gráfico (labels + data)",
         "Formato compatible con Chart.js", "Pendiente"),
    caso("TC-REP-011", "Reportes", "RBAC", "CONSULTOR solo lectura reportes", "Seguridad", "Alta",
         "Consultor autenticado", "1. GET /api/reportes con token consultor",
         "Token rol=CONSULTOR",
         "HTTP 200 (lectura permitida). POST/PUT/PATCH: HTTP 403",
         "requireRole(['ADMIN', 'CONSULTOR']) en GET", "Pendiente"),
    caso("TC-REP-012", "Reportes", "Performance", "Reporte grande > 10s", "Performance", "Media",
         "BD con 10,000+ préstamos", "1. GET /api/reportes/cartera",
         "BD grande",
         "HTTP 200 en <5s. Si >10s: optimizar con índices o paginación",
         "Tiempo de respuesta <5s", "Pendiente"),
    caso("TC-REP-013", "Reportes", "Caja", "Reporte de caja diario", "Funcional", "Media",
         "Movimientos de caja", "1. GET /api/reportes/caja?fecha=2026-08-06",
         "Fecha específica",
         "HTTP 200 con movimientos de caja del día, saldo inicial, final",
         "Conciliación con cierres de caja", "Pendiente"),
    caso("TC-REP-014", "Reportes", "Categorías", "Reporte por categoría", "Funcional", "Baja",
         "Categorías configuradas", "1. GET /api/reportes/categorias",
         "Token admin",
         "HTTP 200 con montos por categoría de préstamo",
         "Agrupación por categoriaId", "Pendiente"),
    caso("TC-REP-015", "Reportes", "Auditoría", "Reporte de AuditLog", "Seguridad", "Media",
         "AuditLog con registros", "1. GET /api/reportes/auditoria?usuarioId=<id>",
         "usuarioId filtro",
         "HTTP 200 con logs del usuario. Solo ADMIN",
         "AuditLog inmutable; solo lectura", "Pendiente"),
]

# ============================================================
# M11 — Integraciones
# ============================================================
M11 = [
    caso("TC-INT-001", "Integraciones", "Brevo SMTP", "Conexión SMTP Brevo", "Smoke", "Alta",
         "BREVO_SMTP_KEY válida", "1. POST /api/email {accion: 'probar'}",
         "Token admin",
         "HTTP 200 success=true. Mensaje 'Conexión SMTP verificada'",
         "ConexionAPI.EMAIL_SMTP.probada=true", "Aprobado"),
    caso("TC-INT-002", "Integraciones", "Brevo HTTPS API", "Llamada a /v3/account", "Integración", "Alta",
         "BREVO_API_KEY válida", "1. fetch https://api.brevo.com/v3/account",
         "API key válida",
         "HTTP 200. Datos de cuenta Brevo devueltos (email, plan)",
         "API key desencriptada de BD", "Aprobado"),
    caso("TC-INT-003", "Integraciones", "WhatsApp Cloud", "Enviar mensaje WhatsApp", "Integración", "Media",
         "WHATSAPP_TOKEN y PHONE_NUMBER_ID configurados", "1. enviarWhatsApp({destino, mensaje})",
         "Número destino, mensaje",
         "HTTP 200 de Meta Cloud API. wamid devuelto",
         "Meta Cloud API integrada", "Pendiente"),
    caso("TC-INT-004", "Integraciones", "Bancolombia", "Botón Bancolombia - iniciar", "Integración", "Media",
         "Integración configurada", "1. POST /api/bancolombia/probar",
         "Token admin",
         "HTTP 200. Botón Bancolombia responde",
         "Integración activa", "Pendiente"),
    caso("TC-INT-005", "Integraciones", "Bancolombia", "Webhook de confirmación", "Integración", "Media",
         "Bancolombia configurado", "1. POST /api/bancolombia/webhook",
         "Payload webhook Bancolombia",
         "HTTP 200. Pago conciliado automáticamente",
         "Webhook con firma válida", "Pendiente"),
    caso("TC-INT-006", "Integraciones", "Vercel", "Deploy automático vía GitHub Actions", "Integración", "Alta",
         "Push a main", "1. git push origin main",
         "Push a rama main",
         "GitHub Actions workflow deploy-vercel.yml corre. Deploy a Vercel producción",
         "Workflow disparado por push", "Aprobado"),
    caso("TC-INT-007", "Integraciones", "Vercel", "Env vars sincronizadas", "Integración", "Alta",
         "VERCEL_TOKEN válido", "1. Listar envs en Vercel",
         "Token Vercel",
         "BREVO_API_KEY, BREVO_SMTP_KEY, API_ENCRYPTION_KEY presentes en Vercel",
         "Variables marcadas como encrypted", "Pendiente"),
    caso("TC-INT-008", "Integraciones", "Neon", "Conexión Neon Postgres", "Smoke", "Alta",
         "DATABASE_URL válida", "1. SELECT 1",
         "Query simple",
         "HTTP 200. Query ejecutada. Pool de conexiones OK",
         "Pooler -pooled de Neon activo", "Aprobado"),
    caso("TC-INT-009", "Integraciones", "Neon", "Conexión con SSL", "Seguridad", "Alta",
         "sslmode=require en URL", "1. Verificar conexión SSL",
         "URL con sslmode=require",
         "Conexión establecida solo con SSL. Sin SSL: rechazada",
         "sslmode=require obligatorio", "Aprobado"),
    caso("TC-INT-010", "Integraciones", "GitHub", "Push a repositorio", "Integración", "Alta",
         "Cambios locales committeados", "1. git push origin main",
         "Commits locales",
         "Push exitoso. Refs actualizadas en GitHub",
         "origin/main = HEAD local", "Aprobado"),
    caso("TC-INT-011", "Integraciones", "GitHub", "Secret scanning activo", "Seguridad", "Alta",
         "Repo con GitHub Advanced Security", "1. git push con secreto en commit",
         "Commit con xkeysib-...",
         "Push rechazado por secret scanning. URL para desbloquear generada",
         "Push declined por repo rule violations", "Aprobado"),
    caso("TC-INT-012", "Integraciones", "GitHub", "Workflow deploy-vercel.yml", "Integración", "Media",
         "Push a main", "1. Verificar GitHub Actions tab",
         "Push disparó workflow",
         "Workflow corre: checkout, setup node, install, build, vercel deploy --prod",
         "Workflow exitoso con VERCEL_TOKEN secret", "Pendiente"),
    caso("TC-INT-013", "Integraciones", "Vercel", "Build de producción exitoso", "Integración", "Alta",
         "Código en main", "1. Trigger deploy\n2. Verificar logs Vercel",
         "Deploy production",
         "Build OK. next build && cp -r .next/static .next/standalone/.next/",
         "Output standalone generado", "Pendiente"),
    caso("TC-INT-014", "Integraciones", "Vercel", "Runtime variables disponibles", "Integración", "Alta",
         "Deploy exitoso", "1. Verificar process.env en runtime",
         "Deploy production",
         "Todas las env vars (BREVO_*, API_ENCRYPTION_KEY, DATABASE_URL) accesibles",
         "Variables encrypted en Vercel", "Pendiente"),
    caso("TC-INT-015", "Integraciones", "Sync DevOps", "Push GitHub + Deploy Vercel + Sync Neon", "Integración", "Alta",
         "Cambios locales listos", "1. git push\n2. Workflow deploy\n3. Migrate prisma db push",
         "Cambios en schema.prisma",
         "GitHub actualizado → Vercel deploy → Neon schema sincronizado",
         "Pipeline DevOps completo", "Aprobado"),
]

# ============================================================
# M12 — UI/UX Mobile/Desktop
# ============================================================
M12 = [
    caso("TC-UI-001", "UI/UX", "Responsive", "Layout desktop 1920px", "UI/UX", "Media",
         "Servidor corriendo", "1. Abrir /login en 1920px width",
         "Viewport 1920x1080",
         "Layout se ve correcto, sin scroll horizontal, componentes centrados",
         "Desktop layout optimizado", "Pendiente"),
    caso("TC-UI-002", "UI/UX", "Responsive", "Layout tablet 768px", "UI/UX", "Media",
         "Servidor corriendo", "1. Abrir /login en 768px width",
         "Viewport 768x1024",
         "Layout se adapta, sidebar colapsa, navegación optimizada",
         "Responsive en breakpoint md", "Pendiente"),
    caso("TC-UI-003", "UI/UX", "Responsive", "Layout mobile 375px", "UI/UX", "Alta",
         "Servidor corriendo", "1. Abrir /login en 375px width",
         "Viewport 375x812 (iPhone X)",
         "Layout mobile, botones táctiles ≥44px, sin scroll horizontal",
         "Mobile-first responsive", "Pendiente"),
    caso("TC-UI-004", "UI/UX", "Navegación", "Sidebar colapsable mobile", "UI/UX", "Media",
         "Mobile 375px", "1. Click en hamburguesa",
         "Click en botón menú",
         "Sidebar se colapsa/despliega. Overlay en mobile",
         "Animación smooth", "Pendiente"),
    caso("TC-UI-005", "UI/UX", "Formularios", "Validación frontend en vivo", "UI/UX", "Media",
         "Form de cliente", "1. Escribir email inválido en input",
         "Email inválido",
         "Mensaje de error en vivo debajo del input. Submit deshabilitado",
         "Validación Zod en frontend (react-hook-form)", "Pendiente"),
    caso("TC-UI-006", "UI/UX", "Formularios", "Mensajes de error accesibles", "UI/UX", "Media",
         "Form con error", "1. Submit form con campos inválidos",
         "Datos inválidos",
         "Errores anunciados con aria-live. Focus en primer campo con error",
         "ARIA labels correctos", "Pendiente"),
    caso("TC-UI-007", "UI/UX", "Accesibilidad", "Navegación por teclado", "UI/UX", "Alta",
         "Página con form", "1. Tab a través de inputs",
         "Solo teclado",
         "Focus visible en todos los elementos interactivos. Orden lógico",
         "Tabindex correcto, focus rings", "Pendiente"),
    caso("TC-UI-008", "UI/UX", "Accesibilidad", "Contraste WCAG AA", "UI/UX", "Media",
         "Cualquier página", "1. Verificar contraste con axe DevTools",
         "Análisis axe",
         "Contraste texto/fondo ≥4.5:1 (texto normal), ≥3:1 (texto grande)",
         "Cumple WCAG 2.1 AA", "Pendiente"),
    caso("TC-UI-009", "UI/UX", "Componentes", "Modal dialog accesible", "UI/UX", "Media",
         "Página con modal", "1. Abrir modal\n2. Tab dentro",
         "Modal abierto",
         "Focus atrapado dentro del modal. Esc cierra. Aria-modal=true",
         "Radix Dialog accesible", "Pendiente"),
    caso("TC-UI-010", "UI/UX", "Componentes", "Toast notifications", "UI/UX", "Baja",
         "Acción que dispara toast", "1. Acción exitosa (ej: guardar cliente)",
         "Acción exitosa",
         "Toast aparece con mensaje. Auto-desaparece en 5s. Botón cerrar",
         "Sonner o similar", "Pendiente"),
    caso("TC-UI-011", "UI/UX", "Componentes", "Tabla con sorting", "UI/UX", "Media",
         "Tabla de clientes", "1. Click en header de columna",
         "Click header",
         "Tabla ordena ascendente/descendente por columna",
         "Icono de orden visible", "Pendiente"),
    caso("TC-UI-012", "UI/UX", "Componentes", "Tabla con paginación", "UI/UX", "Media",
         "Tabla con 8+ filas", "1. Click en página 2",
         "Paginación",
         "Página 2 se carga. Paginación con anterior/siguiente",
         "Estado de página actual visible", "Pendiente"),
    caso("TC-UI-013", "UI/UX", "Loading States", "Skeleton loading", "UI/UX", "Baja",
         "Página con datos async", "1. Cargar página lenta",
         "Carga lenta",
         "Skeleton loaders aparecen mientras cargan datos",
         "shadcn/ui Skeleton", "Pendiente"),
    caso("TC-UI-014", "UI/UX", "Loading States", "Botón con loading state", "UI/UX", "Baja",
         "Botón de submit", "1. Click en submit\n2. Esperar respuesta",
         "Submit form",
         "Botón muestra spinner y se deshabilita. No doble submit",
         "Estado loading visible", "Pendiente"),
    caso("TC-UI-015", "UI/UX", "Cross-Browser", "Funciona en Chrome, Firefox, Safari", "UI/UX", "Media",
         "Múltiples navegadores", "1. Probar login en cada navegador",
         "Chrome 120+, Firefox 120+, Safari 17+",
         "Login funciona en los 3 navegadores. Sin errores en consola",
         "Cross-browser compatible", "Pendiente"),
]

# ============================================================
# M13 — Sync DevOps
# ============================================================
M13 = [
    caso("TC-DEV-001", "Sync DevOps", "GitHub Push", "Push exitoso a main", "Integración", "Alta",
         "Commits locales válidos", "1. git push origin main",
         "git push",
         "Push exitoso. refs/heads/main actualizado en GitHub",
         "origin/main = HEAD local", "Aprobado"),
    caso("TC-DEV-002", "Sync DevOps", "GitHub Push", "Push rechazado por secret scanning", "Seguridad", "Alta",
         "Commit con secreto Brevo", "1. git push origin main",
         "Commit con xkeysib-...",
         "Push declined. Mensaje con URL para desbloquear",
         "Secret scanning protege repo", "Aprobado"),
    caso("TC-DEV-003", "Sync DevOps", "GitHub Actions", "Workflow deploy-vercel disparado", "Integración", "Alta",
         "Push a main", "1. Verificar Actions tab en GitHub",
         "Push a main",
         "Workflow deploy-vercel.yml corre automáticamente",
         "Trigger on push to main", "Aprobado"),
    caso("TC-DEV-004", "Sync DevOps", "GitHub Actions", "Workflow exitoso", "Integración", "Alta",
         "Workflow corriendo", "1. Verificar logs del workflow",
         "Workflow en ejecución",
         "Steps: checkout, setup-node, install, build, vercel deploy --prod exitosos",
         "Tiempo total <20 min", "Pendiente"),
    caso("TC-DEV-005", "Sync DevOps", "Vercel Deploy", "Deploy a producción exitoso", "Integración", "Alta",
         "Workflow completado", "1. Verificar deployment en Vercel dashboard",
         "Deploy production",
         "Vercel deploy completado. URL de producción accesible",
         "vercel deploy --prod --yes exitoso", "Pendiente"),
    caso("TC-DEV-006", "Sync DevOps", "Vercel Env Vars", "Variables sincronizadas", "Integración", "Alta",
         "VERCEL_TOKEN válido", "1. Listar envs en Vercel project",
         "API call a Vercel",
         "BREVO_API_KEY, BREVO_SMTP_KEY, API_ENCRYPTION_KEY, DATABASE_URL presentes",
         "Variables encrypted en Vercel", "Pendiente"),
    caso("TC-DEV-007", "Sync DevOps", "Neon DB", "Schema sincronizado con prisma db push", "Integración", "Alta",
         "schema.prisma actualizado", "1. npx prisma db push",
         "Schema actualizado",
         "Neon DB schema actualizado. Sin pérdida de datos en prod",
         "prisma db push acepta data-loss solo con --accept-data-loss", "Aprobado"),
    caso("TC-DEV-008", "Sync DevOps", "Neon DB", "Conexión pooled vs directa", "Performance", "Media",
         "DATABASE_URL con -pooler", "1. Verificar tipo de conexión",
         "URL con -pooler.c-4.us-east-2.aws.neon.tech",
         "Pooler connection (PgBouncer). Hasta 5 conexiones simultáneas",
         "Pooler Neon para serverless", "Aprobado"),
    caso("TC-DEV-009", "Sync DevOps", "PlataformaSync", "GITHUB sincronizado", "Integración", "Media",
         "PlataformaSync.GITHUB", "1. Verificar en BD",
         "BD Neon",
         "PlataformaSync.GITHUB.sincronizado=true, ultimoEstado=OK",
         "Estado OK en BD", "Aprobado"),
    caso("TC-DEV-010", "Sync DevOps", "PlataformaSync", "VERCEL sincronizado", "Integración", "Media",
         "PlataformaSync.VERCEL", "1. Verificar en BD",
         "BD Neon",
         "PlataformaSync.VERCEL.sincronizado=true, ultimoEstado=OK",
         "Estado OK en BD", "Aprobado"),
    caso("TC-DEV-011", "Sync DevOps", "PlataformaSync", "NEON sincronizado", "Integración", "Media",
         "PlataformaSync.NEON", "1. Verificar en BD",
         "BD Neon",
         "PlataformaSync.NEON.sincronizado=true. ultimoState puede ser ERROR (auto-monitor)",
         "Auto-monitor del estado de Neon", "Pendiente"),
    caso("TC-DEV-012", "Sync DevOps", "ConexionAPI", "EMAIL_SMTP probada y activa", "Integración", "Alta",
         "ConexionAPI.EMAIL_SMTP", "1. Verificar en BD",
         "BD Neon",
         "ConexionAPI.EMAIL_SMTP.activa=true, probada=true, fechaUltimaPrueba reciente",
         "probada=true tras /api/email/probar", "Aprobado"),
    caso("TC-DEV-013", "Sync DevOps", "CorreoInstitucional", "Email principal activo", "Integración", "Media",
         "CorreoInstitucional.jsa@jsadr.com.co", "1. Verificar en BD",
         "BD Neon",
         "CorreoInstitucional.estado=activo, esPrincipal=true, smtpHost=smtp-relay.brevo.com",
         "Email principal configurado", "Aprobado"),
    caso("TC-DEV-014", "Sync DevOps", "Audit Log", "Cada sync registrado en AuditLog", "Seguridad", "Media",
         "Sync DevOps ejecutado", "1. Verificar AuditLog",
         "Sync disparado",
         "Registro en AuditLog con accion='SYNC_GITHUB' / 'SYNC_VERCEL' / 'SYNC_NEON'",
         "Trazabilidad completa de sync", "Pendiente"),
    caso("TC-DEV-015", "Sync DevOps", "Rollback", "Rollback deploy Vercel", "Integración", "Media",
         "Deploy fallido en producción", "1. vercel rollback <deployment-url>",
         "Deployment anterior",
         "Vercel rollback al deployment anterior. Producción vuelve a versión previa",
         "Rollback disponible en dashboard Vercel", "Pendiente"),
]

# ============================================================
# TODOS LOS MÓDULOS
# ============================================================
MODULOS = [
    ("M01", "Autenticación", M01),
    ("M02", "Clientes", M02),
    ("M03", "Préstamos", M03),
    ("M04", "Pagos", M04),
    ("M05", "Correo Electrónico", M05),
    ("M06", "Seguridad", M06),
    ("M07", "Portal Cliente", M07),
    ("M08", "Portal Jurídico", M08),
    ("M09", "Notificaciones", M09),
    ("M10", "Reportes", M10),
    ("M11", "Integraciones", M11),
    ("M12", "UI/UX Mobile/Desktop", M12),
    ("M13", "Sync DevOps", M13),
]

# ============================================================
# Construcción del workbook
# ============================================================
wb = Workbook()
wb.properties.creator = "Z.ai"
wb.properties.title = "Plan de Pruebas QA - JSADR Plataforma"
wb.properties.subject = "Casos de prueba por módulo - Estándar QA Tester"

# Remover hoja por defecto
default_sheet = wb.active
wb.remove(default_sheet)

# -----------------------------------------------------------
# HOJA 1: Portada
# -----------------------------------------------------------
ws1 = wb.create_sheet("1. Portada")
B.setup_sheet(ws1, "Plan de Pruebas QA — JSADR Plataforma de Gestión de Préstamos", 8)

# Columnas
ws1.column_dimensions["A"].width = 4
ws1.column_dimensions["B"].width = 30
ws1.column_dimensions["C"].width = 50
ws1.column_dimensions["D"].width = 30

# Metadata
metadata = [
    ("Versión", "1.0", ""),
    ("Fecha", datetime.now().strftime("%Y-%m-%d"), ""),
    ("Responsable", "Equipo QA Tester", ""),
    ("Sistema", "JSADR Plataforma de Gestión de Préstamos", ""),
    ("Stack", "Next.js 16 + TypeScript + Prisma + Neon Postgres + Brevo SMTP", ""),
    ("URL Producción", "https://jsadr.com.co (Vercel)", ""),
    ("URL Desarrollo", "http://localhost:3000", ""),
    ("", "", ""),
    ("ALCANCE", "", ""),
    ("Cobertura", "13 módulos funcionales", ""),
    ("Casos totales", str(sum(len(m[2]) for m in MODULOS)), ""),
    ("Prioridad Alta", str(sum(1 for m in MODULOS for c in m[2] if c["Prioridad"] == "Alta")), ""),
    ("Prioridad Media", str(sum(1 for m in MODULOS for c in m[2] if c["Prioridad"] == "Media")), ""),
    ("Prioridad Baja", str(sum(1 for m in MODULOS for c in m[2] if c["Prioridad"] == "Baja")), ""),
    ("", "", ""),
    ("OBJETIVOS", "", ""),
    ("Obj 1", "Validar que TODAS las funciones críticas del sistema operen correctamente", ""),
    ("Obj 2", "Verificar integraciones con servicios externos (Brevo, Vercel, Neon, GitHub, Bancolombia, WhatsApp)", ""),
    ("Obj 3", "Asegurar cumplimiento de requisitos de seguridad (RBAC, AES-256, bcrypt, JWT, rate limiting)", ""),
    ("Obj 4", "Confirmar que el envío de correos (OTP, reset, notificaciones) funcione al 100%", ""),
    ("", "", ""),
    ("LEYENDA - PRIORIDAD", "", ""),
    ("Alta", "Funcionalidad crítica, bloquea release", "🔴"),
    ("Media", "Funcionalidad importante, no bloquea release", "🟡"),
    ("Baja", "Mejoras de UX/performance", "🟢"),
    ("", "", ""),
    ("LEYENDA - ESTADO", "", ""),
    ("Pendiente", "Caso no ejecutado", "⚪"),
    ("Aprobado", "Caso ejecutado y exitoso", "🟢"),
    ("Fallido", "Caso ejecutado y fallido (bug encontrado)", "🔴"),
    ("Bloqueado", "Caso no se puede ejecutar por dependencia", "🟡"),
    ("En Progreso", "Caso en ejecución", "🔵"),
]

for i, (k, v, extra) in enumerate(metadata, start=4):
    ws1.cell(row=i, column=2, value=k).font = B.font_subheader() if k and not k.startswith(("Obj", "Alta", "Media", "Baja", "Pendiente", "Aprobado", "Fallido", "Bloqueado", "En Progreso")) else B.font_body()
    ws1.cell(row=i, column=3, value=v).font = B.font_body()
    ws1.cell(row=i, column=4, value=extra).font = B.font_body()
    ws1.cell(row=i, column=2).alignment = B.align_text()
    ws1.cell(row=i, column=3).alignment = B.align_text()

# -----------------------------------------------------------
# HOJA 2: Índice de Módulos
# -----------------------------------------------------------
ws2 = wb.create_sheet("2. Índice de Módulos")
B.setup_sheet(ws2, "Índice de Módulos — Plan de Pruebas QA", 8)

headers_idx = ["ID", "Módulo", "Descripción", "Casos Totales", "Alta", "Media", "Baja", "Aprobados"]
for col, h in enumerate(headers_idx, 2):
    ws2.cell(row=4, column=col, value=h)

descripciones = {
    "M01": "Login sistema, MFA/OTP, recuperación de clave, rate limiting, bloqueo",
    "M02": "CRUD cliente, búsqueda, gestión PIN/clave, tasas personalizadas",
    "M03": "Creación, cálculo intereses/mora, estados, firma electrónica, TyC",
    "M04": "Registro pago, reverso, anulación, conciliación, validaciones",
    "M05": "SMTP Brevo, OTP, recuperación, notificaciones masivas, Ethereal fallback",
    "M06": "RBAC, audit log, AES-256, bcrypt, JWT, rate limiting, sanitización",
    "M07": "Login cédula+PIN/clave, OTP firma, clave dinámica, estado cuenta",
    "M08": "Login abogado (cédula/username+clave), casos, sesión 8h, notas",
    "M09": "Recordatorios, mora, WhatsApp Cloud API, Email, logs, deduplicación",
    "M10": "Cartera, morosidad, estados financieros, exportación Excel/PDF",
    "M11": "Brevo, WhatsApp, Bancolombia, Vercel, Neon, GitHub",
    "M12": "Responsive, navegación, accesibilidad, formularios, cross-browser",
    "M13": "GitHub push, Vercel deploy, Neon sync, env vars, secret scanning",
}

for i, (mid, mname, casos) in enumerate(MODULOS):
    row = 5 + i
    alta = sum(1 for c in casos if c["Prioridad"] == "Alta")
    media = sum(1 for c in casos if c["Prioridad"] == "Media")
    baja = sum(1 for c in casos if c["Prioridad"] == "Baja")
    aprobados = sum(1 for c in casos if c["Estado"] == "Aprobado")
    ws2.cell(row=row, column=2, value=mid)
    ws2.cell(row=row, column=3, value=mname)
    ws2.cell(row=row, column=4, value=descripciones[mid])
    ws2.cell(row=row, column=5, value=len(casos))
    ws2.cell(row=row, column=6, value=alta)
    ws2.cell(row=row, column=7, value=media)
    ws2.cell(row=row, column=8, value=baja)
    ws2.cell(row=row, column=9, value=aprobados)
    B.style_data_row(ws2, row, 2, 9, i)

# Aplicar estilo a header
B.style_header_row(ws2, 4, 2, 9)

# Total
total_row = 5 + len(MODULOS)
ws2.cell(row=total_row, column=2, value="TOTAL")
ws2.cell(row=total_row, column=5, value=f"=SUM(E5:E{total_row-1})")
ws2.cell(row=total_row, column=6, value=f"=SUM(F5:F{total_row-1})")
ws2.cell(row=total_row, column=7, value=f"=SUM(G5:G{total_row-1})")
ws2.cell(row=total_row, column=8, value=f"=SUM(H5:H{total_row-1})")
ws2.cell(row=total_row, column=9, value=f"=SUM(I5:I{total_row-1})")
B.style_total_row(ws2, total_row, 2, 9)

# Anchos de columna
for col_letter, width in zip("BCDEFGHI", [10, 25, 50, 14, 10, 10, 10, 12]):
    ws2.column_dimensions[col_letter].width = width

# Freeze
ws2.freeze_panes = "C5"

# -----------------------------------------------------------
# Hojas 3-15: Una por módulo
# -----------------------------------------------------------
HEADERS = ["ID", "Módulo", "Función", "Caso de Prueba", "Tipo", "Prioridad",
           "Precondiciones", "Pasos", "Datos de Entrada", "Resultado Esperado",
           "Criterios de Aceptación", "Estado"]

# Anchos por columna
COL_WIDTHS = [12, 16, 22, 35, 14, 12, 35, 45, 35, 50, 40, 14]

def build_modulo_sheet(ws, modulo_id, modulo_name, casos):
    B.setup_sheet(ws, f"{modulo_id} — {modulo_name}", 13)
    
    # Headers en fila 4
    for col, h in enumerate(HEADERS, 2):
        ws.cell(row=4, column=col, value=h)
    B.style_header_row(ws, 4, 2, 13)
    
    # Datos
    for i, c in enumerate(casos):
        row = 5 + i
        values = [c["ID"], c["Módulo"], c["Función"], c["Caso de Prueba"], c["Tipo"], c["Prioridad"],
                  c["Precondiciones"], c["Pasos"], c["Datos de Entrada"], c["Resultado Esperado"],
                  c["Criterios de Aceptación"], c["Estado"]]
        for col, v in enumerate(values, 2):
            cell = ws.cell(row=row, column=col, value=v)
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        B.style_data_row(ws, row, 2, 13, i)
        
        # Aplicar color de prioridad
        prio_cell = ws.cell(row=row, column=7)  # Prioridad
        prio_color = PRIORIDADES.get(c["Prioridad"], "000000")
        prio_cell.font = Font(name=B.FONT_NAME, size=11, bold=True, color=prio_color)
        
        # Aplicar color de estado
        estado_cell = ws.cell(row=row, column=13)  # Estado
        estado_color = ESTADOS.get(c["Estado"], "000000")
        estado_cell.font = Font(name=B.FONT_NAME, size=11, bold=True, color=estado_color)
        
        # Altura de fila dinámica según contenido
        max_lines = max(
            str(c["Pasos"]).count("\n") + 1,
            str(c["Resultado Esperado"]).count("\n") + 1,
            str(c["Criterios de Aceptación"]).count("\n") + 1,
            3
        )
        ws.row_dimensions[row].height = max(22, 18 * max_lines)
    
    # Anchos de columna
    for col_letter, width in zip("BCDEFGHIJKLM", COL_WIDTHS):
        ws.column_dimensions[col_letter].width = width
    
    # Freeze panes
    ws.freeze_panes = "F5"
    
    # Filtros
    ws.auto_filter.ref = f"B4:M{4+len(casos)}"

for idx, (mid, mname, casos) in enumerate(MODULOS, start=3):
    sheet_name = f"{idx}. {mid}-{mname}"[:31].replace("/", "-")  # Excel limit 31 chars + no '/'
    ws = wb.create_sheet(sheet_name)
    build_modulo_sheet(ws, mid, mname, casos)

# -----------------------------------------------------------
# Hoja final: Resumen Ejecutivo
# ------------------------------------------------===========
ws_resumen = wb.create_sheet(f"{len(MODULOS)+3}. Resumen Ejecutivo")
B.setup_sheet(ws_resumen, "Resumen Ejecutivo — Cobertura del Plan QA", 8)

headers_res = ["Módulo", "Casos", "Alta", "Media", "Baja", "Aprobados", "Pendientes", "% Aprobación"]
for col, h in enumerate(headers_res, 2):
    ws_resumen.cell(row=4, column=col, value=h)
B.style_header_row(ws_resumen, 4, 2, 9)

total_casos = 0
total_alta = 0
total_media = 0
total_baja = 0
total_aprobados = 0

for i, (mid, mname, casos) in enumerate(MODULOS):
    row = 5 + i
    alta = sum(1 for c in casos if c["Prioridad"] == "Alta")
    media = sum(1 for c in casos if c["Prioridad"] == "Media")
    baja = sum(1 for c in casos if c["Prioridad"] == "Baja")
    aprobados = sum(1 for c in casos if c["Estado"] == "Aprobado")
    pendientes = len(casos) - aprobados
    pct = (aprobados / len(casos) * 100) if casos else 0
    
    total_casos += len(casos)
    total_alta += alta
    total_media += media
    total_baja += baja
    total_aprobados += aprobados
    
    ws_resumen.cell(row=row, column=2, value=f"{mid} — {mname}")
    ws_resumen.cell(row=row, column=3, value=len(casos))
    ws_resumen.cell(row=row, column=4, value=alta)
    ws_resumen.cell(row=row, column=5, value=media)
    ws_resumen.cell(row=row, column=6, value=baja)
    ws_resumen.cell(row=row, column=7, value=aprobados)
    ws_resumen.cell(row=row, column=8, value=pendientes)
    pct_cell = ws_resumen.cell(row=row, column=9, value=pct/100)
    pct_cell.number_format = "0.0%"
    B.style_data_row(ws_resumen, row, 2, 9, i)

# Total
total_row = 5 + len(MODULOS)
ws_resumen.cell(row=total_row, column=2, value="TOTAL")
ws_resumen.cell(row=total_row, column=3, value=total_casos)
ws_resumen.cell(row=total_row, column=4, value=total_alta)
ws_resumen.cell(row=total_row, column=5, value=total_media)
ws_resumen.cell(row=total_row, column=6, value=total_baja)
ws_resumen.cell(row=total_row, column=7, value=total_aprobados)
ws_resumen.cell(row=total_row, column=8, value=total_casos - total_aprobados)
pct_total = (total_aprobados / total_casos) if total_casos else 0
pct_cell = ws_resumen.cell(row=total_row, column=9, value=pct_total)
pct_cell.number_format = "0.0%"
B.style_total_row(ws_resumen, total_row, 2, 9)

# Anchos
for col_letter, width in zip("BCDEFGHI", [35, 10, 10, 10, 10, 12, 12, 14]):
    ws_resumen.column_dimensions[col_letter].width = width

ws_resumen.freeze_panes = "C5"

# Notas
nota_row = total_row + 3
ws_resumen.cell(row=nota_row, column=2, value="NOTAS:").font = B.font_subheader()
notas = [
    f"• Total de casos de prueba: {total_casos} distribuidos en {len(MODULOS)} módulos funcionales",
    f"• Casos de prioridad Alta: {total_alta} (deben aprobarse antes de release)",
    f"• Casos aprobados hasta ahora: {total_aprobados} ({pct_total*100:.1f}%)",
    f"• Casos pendientes: {total_casos - total_aprobados}",
    "• Casos marcados como 'Aprobado' ya fueron ejecutados exitosamente en pruebas previas",
    "• Casos 'Pendiente' deben ejecutarse en próxima iteración de QA",
    "• Modificar Estado a 'Fallido' si el caso falla, adjuntando evidencia en columnas adicionales",
]
for i, n in enumerate(notas):
    ws_resumen.cell(row=nota_row+1+i, column=2, value=n).font = B.font_body()

# ============================================================
# Guardar
# ============================================================
os.makedirs(os.path.dirname(OUTPUT), exist_ok=True)
wb.save(OUTPUT)

print(f"\n✅ Excel QA generado: {OUTPUT}")
print(f"   Hojas: {len(wb.sheetnames)}")
print(f"   Módulos: {len(MODULOS)}")
print(f"   Casos totales: {total_casos}")
print(f"   Casos aprobados: {total_aprobados} ({pct_total*100:.1f}%)")
