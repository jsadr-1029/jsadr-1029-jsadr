"""Actualiza el archivo Excel marcando items como Aprobados."""
import sys
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment

WB_PATH = "/home/z/my-project/upload/plan-pruebas-qa-jsadr.xlsx"

# Mapeo: sheet_name → [(row_num, tc_id), ...]
updates = {
    "4. M02-Clientes": [
        (5, "TC-CLI-001"),
    ],
}

# Estilos
green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
green_font = Font(color="006100", bold=True)

wb = load_workbook(WB_PATH)
total_updated = 0

for sheet_name, items in updates.items():
    ws = wb[sheet_name]
    for row_num, tc_id in items:
        # Verificar que el ID coincide (columna B = índice 1)
        cell_id = ws.cell(row=row_num, column=2).value
        if cell_id != tc_id:
            print(f"⚠ {sheet_name} fila {row_num}: esperado {tc_id}, encontrado {cell_id}")
            continue
        # Columna Estado = M (índice 13)
        cell_state = ws.cell(row=row_num, column=13)
        old = cell_state.value
        cell_state.value = "Aprobado"
        cell_state.fill = green_fill
        cell_state.font = green_font
        cell_state.alignment = Alignment(horizontal="center", vertical="center")
        print(f"✓ {sheet_name} fila {row_num} ({tc_id}): '{old}' → 'Aprobado'")
        total_updated += 1

wb.save(WB_PATH)
print(f"\n=== Total actualizado: {total_updated} items ===")
print(f"Archivo guardado: {WB_PATH}")
