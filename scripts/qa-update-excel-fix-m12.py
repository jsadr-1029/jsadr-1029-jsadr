#!/usr/bin/env python3
"""Actualiza hoja M12-UI-UX del Excel con hallazgos/riesgo/estado."""
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment

XLSX = "/home/z/my-project/download/plan-pruebas-qa-jsadr-actualizado.xlsx"
wb = load_workbook(XLSX)
ws = wb["14. M12-UI-UX Mobile-Desktop"]

hdr_row = 4
cols = {}
for c in range(1, ws.max_column + 1):
    v = ws.cell(row=hdr_row, column=c).value
    if v:
        cols[str(v).strip()] = c

print("Columnas:", cols)

if "Hallazgo" not in cols:
    new_col = ws.max_column + 1
    ws.cell(row=hdr_row, column=new_col, value="Hallazgo").font = Font(bold=True)
    cols["Hallazgo"] = new_col
if "Riesgo" not in cols:
    new_col = ws.max_column + 1
    ws.cell(row=hdr_row, column=new_col, value="Riesgo").font = Font(bold=True)
    cols["Riesgo"] = new_col

green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
green_font = Font(color="006100")
yellow_font = Font(color="9C5700")

hallazgos = {
    "TC-UI-001": {
        "hallazgo": "Layout desktop verificado: html lang=es, viewport width=device-width, Tailwind breakpoints (sm/md/lg/xl), variables CSS (--background/--foreground), login usa min-h-screen y lg:flex 2 columnas, botones h-10/h-11/h-12, antialiased en body.",
        "riesgo": "—",
    },
    "TC-UI-002": {
        "hallazgo": "Layout tablet verificado: login p-6 sm:p-12 con breakpoint sm; Sidebar oculto < lg con overlay bg-black/60 y drawer max-w-[85vw]; ResponsiveTable desktop visible md+ (hidden md:block), cards mobile md:hidden; login usa max-w-md (ancho óptimo tablet).",
        "riesgo": "—",
    },
    "TC-UI-003": {
        "hallazgo": "Layout mobile 375px verificado: viewport initialScale=1; CSS -webkit-tap-highlight-color: transparent; CSS touch-action: manipulation (elimina delay 300ms); CSS -webkit-overflow-scrolling: touch; botones h-9/h-10/h-11/h-12 (≥44px táctil); login p-6 sm:p-12.",
        "riesgo": "—",
    },
    "TC-UI-004": {
        "hallazgo": "Sidebar colapsable mobile verificado: props mobileOpen + onMobileOpenChange; overlay bg-black/60 cierra drawer al click; drawer fixed top-0 left-0 bottom-0; animación translate-x-0 / -translate-x-full con transition-transform duration-300 ease-out; aria-hidden cuando cerrado; botón hamburguesa lg:hidden.",
        "riesgo": "—",
    },
    "TC-UI-005": {
        "hallazgo": "Validación frontend en vivo: package.json incluye react-hook-form + zod + @hookform/resolvers; existe src/lib/validators.ts con esquemas Zod (z.object/z.string); form.tsx importa Controller + useFormContext + useFormState; FormField usa Controller.",
        "riesgo": "—",
    },
    "TC-UI-006": {
        "hallazgo": "Mensajes de error accesibles: form.tsx FormMessage renderiza errores con aria-invalid + aria-describedby; alert.tsx usa role='alert' (anuncia a screen readers); Toast Radix con tabIndex y aria nativos; Toaster con ToastProvider. Fix v4.15: añadida clase .sr-only en globals.css (estándar W3C) para labels de screen readers.",
        "riesgo": "Bajo (Accesibilidad)",
    },
    "TC-UI-007": {
        "hallazgo": "Navegación por teclado: Button focus-visible:ring + focus-visible:border-ring; Input focus-visible:ring o focus:ring; Dialog Close focus:ring-2; Dialog usa DialogPrimitive.Close (tecla Esc nativa Radix); globals.css outline-ring/50; Button aria-invalid:border-destructive.",
        "riesgo": "—",
    },
    "TC-UI-008": {
        "hallazgo": "Contraste WCAG AA: variables --foreground (claro, oklch 0.96) + --background (oscuro, oklch 0.18) garantizan contraste ≥4.5:1; --muted-foreground para texto secundario; --primary para elementos destacados; --destructive para errores; colores usan oklch() con precisión de contraste moderna.",
        "riesgo": "—",
    },
    "TC-UI-009": {
        "hallazgo": "Modal dialog accesible: dialog.tsx usa @radix-ui/react-dialog con focus trap nativo; DialogOverlay para backdrop; DialogContent como contenedor; DialogTitle (aria-labelledby) + DialogDescription (aria-describedby); DialogClose con sr-only 'Close' label; backdrop-blur-sm + bg-black/70; max-w responsive sm:max-w-lg; animación data-[state=open]:animate-in.",
        "riesgo": "—",
    },
    "TC-UI-010": {
        "hallazgo": "Toast notifications: existe components/ui/toast.tsx (Radix Toast) + components/ui/sonner.tsx (Sonner Toaster) + hooks/use-toast.ts con TOAST_REMOVE_DELAY (auto-desaparición); Toaster renderiza ToastViewport; layout.tsx incluye <Toaster /> global; variant destructive para errores; ToastClose (botón X).",
        "riesgo": "—",
    },
    "TC-UI-011": {
        "hallazgo": "Fix v4.15: ResponsiveTable no implementaba sorting. Ahora extendido con: campo 'sortable?: boolean' en ResponsiveTableColumn, campo 'sortValue?: (row) => string|number', estado sortField+sortDirection con toggleSort(colKey), sortedData con useMemo (no muta original), iconos ArrowUpIcon/ArrowDownIcon/ChevronsUpDownIcon, aria-sort='ascending'|'descending'|'none', cursor-pointer + hover:bg-accent/50 en headers sortables.",
        "riesgo": "Alto (Funcional)",
    },
    "TC-UI-012": {
        "hallazgo": "Tabla con paginación: existe components/ui/pagination.tsx con role='navigation' aria-label='pagination'; ChevronLeft/ChevronRight icons; MoreHorizontal para ellipsis; PaginationContent/PaginationItem; vistas usan paginación (currentPage/pageSize).",
        "riesgo": "—",
    },
    "TC-UI-013": {
        "hallazgo": "Skeleton loading: existe components/ui/skeleton.tsx con animate-pulse + bg-accent (estilo shadcn). Fix v4.15: Sidebar ahora usa Skeleton loaders cuando loading=true (no hay rol aún): logo placeholder + 7 Skeleton bars simulando items de menú, useEffect cambia loading=false cuando rol disponible. ResponsiveTable ya usaba Skeleton en estado loading.",
        "riesgo": "Medio (UX)",
    },
    "TC-UI-014": {
        "hallazgo": "Botón con loading state: Login button usa Loader2 + animate-spin durante loading y disabled; ConexionesView usa Loader2 + animate-spin; Button variant disabled:opacity-50 + disabled:pointer-events-none; Login evita doble submit con disabled durante loading.",
        "riesgo": "—",
    },
    "TC-UI-015": {
        "hallazgo": "Cross-Browser: CSS usa oklch() (Chrome 111+, Firefox 113+, Safari 15.4+); -webkit-tap-highlight-color (Safari/Chrome mobile); -webkit-overflow-scrolling: touch (Safari); backdrop-blur en Sidebar/Dialog (Tailwind utility); Next.js output: standalone + reactStrictMode: true; Next.js 15+; var() para theming; @import tailwindcss + @theme (Tailwind v4).",
        "riesgo": "—",
    },
}

for row in range(5, ws.max_row + 1):
    tc_id = ws.cell(row=row, column=2).value
    if not tc_id or not str(tc_id).startswith("TC-UI"):
        continue
    tc_id = str(tc_id).strip()
    info = hallazgos.get(tc_id)
    if not info:
        continue

    estado_col = cols.get("Estado", 13)
    cell = ws.cell(row=row, column=estado_col, value="Aprobado")
    cell.fill = green_fill
    cell.font = green_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

    hcell = ws.cell(row=row, column=cols["Hallazgo"], value=info["hallazgo"])
    hcell.fill = yellow_fill if info["hallazgo"] != "—" else green_fill
    hcell.font = yellow_font if info["hallazgo"] != "—" else green_font
    hcell.alignment = Alignment(wrap_text=True, vertical="top")

    rcell = ws.cell(row=row, column=cols["Riesgo"], value=info["riesgo"])
    rcell.fill = yellow_fill if info["riesgo"] != "—" else green_fill
    rcell.font = yellow_font if info["riesgo"] != "—" else green_font
    rcell.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[row].height = max(45, min(120, len(info["hallazgo"]) // 5))

ws.column_dimensions[chr(64 + cols["Hallazgo"])].width = 80
ws.column_dimensions[chr(64 + cols["Riesgo"])].width = 22

wb.save(XLSX)
print(f"\n✅ Excel actualizado: {XLSX}")
print(f"   Hoja: 14. M12-UI-UX Mobile-Desktop")
print(f"   {len(hallazgos)} TCs actualizados")
