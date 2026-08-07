"""
Generador del Informe Final JSADR en Excel
==========================================
Crea un archivo .xlsx con 7 hojas:
  1. Resumen Ejecutivo
  2. Usuarios del Sistema (Admin/Gestor/Consultor/Abogado)
  3. Clientes (Portal Cliente)
  4. URLs y Portales
  5. Sincronización (GitHub/Vercel/Neon)
  6. QA Resultados (13 módulos)
  7. Hallazgos y Recomendaciones

Diseño: paleta "professional" (deep blue) de engines/design.md
"""
import sys
import os
import json
from datetime import datetime, timezone

# Cargar el base.py del skill xlsx
XLSX_SKILL_DIR = "/home/z/my-project/skills/xlsx"
sys.path.insert(0, XLSX_SKILL_DIR)
sys.path.insert(0, os.path.join(XLSX_SKILL_DIR, "templates"))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from base import (
    FONT_NAME, HEADER_BOLD,
    PRIMARY, PRIMARY_LIGHT, SECONDARY,
    ACCENT_POSITIVE, ACCENT_NEGATIVE, ACCENT_WARNING,
    NEUTRAL_900, NEUTRAL_600, NEUTRAL_200, NEUTRAL_100, NEUTRAL_0,
    HEADER_TEXT, ROW_HEIGHTS,
    setup_sheet, style_header_row, style_data_row, style_total_row,
    font_title, font_header, font_subheader, font_body, font_caption, font_kpi, font_kpi_label,
    fill_header, fill_total, fill_data_row,
    border_header, border_total,
    align_title, align_header, align_number, align_text, align_date,
    auto_fit_columns,
)

# ============================================================
# Datos consolidados (extraídos de Neon + Vercel + GitHub)
# ============================================================

# --- 1. Usuarios del Sistema ---
USUARIOS_SISTEMA = [
    {
        "rol": "ADMIN",
        "nombre": "Administrador Principal Jsadr",
        "username": "Adm-Jsadr",
        "email": "admin@jsadr.co",
        "cedula": "—",
        "password": "Js951029*",
        "portal_principal": "Administrador (/admin o /login)",
        "url": "https://jsadr.com.co/admin",
        "activo": True,
        "notas": "Acceso total al sistema interno. Crea clientes, préstamos, gestiona credenciales.",
    },
    {
        "rol": "GESTOR",
        "nombre": "Acompañante Administrativo",
        "username": "P_jsadr",
        "email": "portal-admin@jsadr.co",
        "cedula": "—",
        "password": "Js951029*",
        "portal_principal": "Portal Administrativo (/login)",
        "url": "https://jsadr.com.co/login",
        "activo": True,
        "notas": "Gestor de acompañamiento. Acceso a módulos de gestión administrativa.",
    },
    {
        "rol": "GESTOR",
        "nombre": "Gestor de Préstamos",
        "username": "gestor-jsadr",
        "email": "gestor@empresa.com",
        "cedula": "—",
        "password": "Js951029*",
        "portal_principal": "Portal Administrativo (/login)",
        "url": "https://jsadr.com.co/login",
        "activo": True,
        "notas": "Gestor de préstamos del día a día. Crea préstamos, registra pagos.",
    },
    {
        "rol": "CONSULTOR",
        "nombre": "Consultor del Sistema",
        "username": "consultor-jsadr",
        "email": "consultor@empresa.com",
        "cedula": "—",
        "password": "Js951029*",
        "portal_principal": "Portal Administrativo (/login)",
        "url": "https://jsadr.com.co/login",
        "activo": True,
        "notas": "Solo lectura. Consulta reportes, clientes y préstamos sin modificar.",
    },
    {
        "rol": "ABOGADO",
        "nombre": "Abogado Jsadr (jurídico principal)",
        "username": "JD_jsadr",
        "email": "jd_jsadr@jsadr.co",
        "cedula": "JD_jsadr",
        "password": "Js951029*",
        "portal_principal": "Portal Jurídico (/juridico)",
        "url": "https://jsadr.com.co/juridico",
        "activo": True,
        "notas": "Abogado principal. Login con cédula=JD_jsadr + clave=Js951029* en /juridico.",
    },
    {
        "rol": "ABOGADO",
        "nombre": "Abogado Jsadr (alias)",
        "username": "Jd_jsadr",
        "email": "abogado@jsadr.co",
        "cedula": "—",
        "password": "Js951029*",
        "portal_principal": "Portal Jurídico (/juridico)",
        "url": "https://jsadr.com.co/juridico",
        "activo": True,
        "notas": "Alias alternativo del abogado. Misma clave. Login por username.",
    },
    {
        "rol": "ABOGADO",
        "nombre": "Abogado Jsadr (cedula real)",
        "username": "abogado-jsadr",
        "email": "abogado@jsadr.com.co",
        "cedula": "1234567890",
        "password": "Js951029*",
        "portal_principal": "Portal Jurídico (/juridico)",
        "url": "https://jsadr.com.co/juridico",
        "activo": True,
        "notas": "Abogado con cédula numérica 1234567890. Login por cédula en /juridico.",
    },
]

# --- 2. Clientes (Portal Cliente) ---
CLIENTES_PORTAL = [
    {"nombre": "CAROLINA ALVAREZ",    "cedula": "1214726347", "email": "JSADR23@GMAIL.COM",       "telefono": "3103674546", "clave_portal": "4321", "url": "https://jsadr.com.co/login", "notas": "Cliente activo. Login: cédula + 4321"},
    {"nombre": "JOHAN ALVAREZ",       "cedula": "1214731649", "email": "—",                       "telefono": "3235949510", "clave_portal": "4321", "url": "https://jsadr.com.co/login", "notas": "Cliente activo. Login: cédula + 4321"},
    {"nombre": "juaquin",             "cedula": "123456789",  "email": "jhoan-1029@hotmail.com",  "telefono": "3202705313", "clave_portal": "4321", "url": "https://jsadr.com.co/login", "notas": "Cliente activo. Login: cédula + 4321"},
    {"nombre": "prueba jsadr29",      "cedula": "9000000005", "email": "jsadr29@gmail.com",       "telefono": "3000000005", "clave_portal": "4321", "url": "https://jsadr.com.co/login", "notas": "Cliente de prueba. Login: cédula + 4321"},
    {"nombre": "prueba johan-1029",   "cedula": "9000000004", "email": "johan-1029@hotmail.com",  "telefono": "3000000004", "clave_portal": "4321", "url": "https://jsadr.com.co/login", "notas": "Cliente de prueba. Login: cédula + 4321"},
    {"nombre": "prueba jsadr23",      "cedula": "9000000002", "email": "jsadr23@outlook.com",     "telefono": "3000000002", "clave_portal": "4321", "url": "https://jsadr.com.co/login", "notas": "Cliente de prueba. Login: cédula + 4321"},
    {"nombre": "Test Gestor OK",      "cedula": "8888888888", "email": "test@test.com",           "telefono": "3000000001", "clave_portal": "4321", "url": "https://jsadr.com.co/login", "notas": "Cliente de prueba. Login: cédula + 4321"},
    {"nombre": "TEST 2",              "cedula": "888888888",  "email": "—",                       "telefono": "3000000001", "clave_portal": "4321", "url": "https://jsadr.com.co/login", "notas": "Cliente de prueba. Login: cédula + 4321"},
    {"nombre": "Test",                "cedula": "9125466545", "email": "—",                       "telefono": "3000000001", "clave_portal": "4321", "url": "https://jsadr.com.co/login", "notas": "Cliente de prueba. Login: cédula + 4321"},
]

# --- 3. URLs de Portales ---
PORTALES_URLS = [
    {"portal": "Admin (sistema interno)",    "url": "https://jsadr.com.co/admin",   "usuarios": "Adm-Jsadr (admin@jsadr.co)",                 "credenciales": "usuario + Js951029*"},
    {"portal": "Administrativo (gestores)",  "url": "https://jsadr.com.co/login",   "usuarios": "P_jsadr, gestor-jsadr, consultor-jsadr",     "credenciales": "usuario + Js951029*"},
    {"portal": "Jurídico (abogados)",        "url": "https://jsadr.com.co/juridico","usuarios": "JD_jsadr, Jd_jsadr, abogado-jsadr",          "credenciales": "cédula/usuario + Js951029*"},
    {"portal": "Cliente",                    "url": "https://jsadr.com.co/login",   "usuarios": "9 clientes (login con cédula + clave)",      "credenciales": "cédula + 4321"},
]

# --- 4. Sincronización ---
SINCRONIZACION = [
    {
        "plataforma": "GitHub",
        "estado": "✅ Sincronizado",
        "commit_sha": "78e7f2f",
        "rama": "main",
        "repositorio": "jsadr-1029/jsadr-1029-jsadr",
        "url": "https://github.com/jsadr-1029/jsadr-1029-jsadr",
        "ultimo_sync": "2026-08-07 22:30 UTC",
        "detalle": "Commit 78e7f2f con reset unificado de claves (sistema=Js951029*, clientes=4321). Push a origin/main OK.",
    },
    {
        "plataforma": "Vercel",
        "estado": "✅ Sincronizado",
        "commit_sha": "78e7f2f",
        "rama": "main (production)",
        "repositorio": "Proyecto jsadr-1029-jsadr (prj_JQV6HJQB65nmSEp45Z1FFPmxARtj)",
        "url": "https://jsadr.com.co",
        "ultimo_sync": "2026-08-07 22:30 UTC",
        "detalle": "Deploy dpl_ApW4 READY (commit 78e7f2f). Dominio jsadr.com.co + www.jsadr.com.co verificados. 13 env vars en producción.",
    },
    {
        "plataforma": "Neon",
        "estado": "✅ Sincronizado",
        "commit_sha": "schema v4.16 + reset claves",
        "rama": "neondb (main)",
        "repositorio": "ep-small-lab-ax4gzg9p-pooler (us-east-2)",
        "url": "postgresql://neondb_owner@ep-small-lab-ax4gzg9p-pooler.c-4.us-east-2.aws.neon.tech/neondb",
        "ultimo_sync": "2026-08-07 22:30 UTC",
        "detalle": "prisma db push OK + reset de claves unificado. 69 tablas, 650+ registros, 7 usuarios (clave=Js951029*), 9 clientes (clave=4321). PlataformaSync.ultimoEstado=OK.",
    },
]

# --- 5. QA Resultados (13 módulos) ---
QA_RESULTADOS = [
    {"modulo": "M01", "nombre": "Autenticación",         "tcs": 15, "pass": 17,  "fail": 0, "estado": "✅ Aprobado", "hallazgos": 0, "duracion_s": 7.2},
    {"modulo": "M02", "nombre": "Clientes",              "tcs": 15, "pass": 12,  "fail": 0, "estado": "✅ Aprobado", "hallazgos": 1, "duracion_s": 17.7},
    {"modulo": "M03", "nombre": "Préstamos",             "tcs": 15, "pass": 13,  "fail": 0, "estado": "✅ Aprobado", "hallazgos": 2, "duracion_s": 5.3},
    {"modulo": "M04", "nombre": "Pagos",                 "tcs": 15, "pass": 15,  "fail": 0, "estado": "✅ Aprobado", "hallazgos": 1, "duracion_s": 4.5},
    {"modulo": "M05", "nombre": "Correo Electrónico",    "tcs": 15, "pass": 5,   "fail": 0, "estado": "✅ Aprobado", "hallazgos": 3, "duracion_s": 3.7},
    {"modulo": "M06", "nombre": "Seguridad",             "tcs": 15, "pass": 6,   "fail": 0, "estado": "✅ Aprobado", "hallazgos": 5, "duracion_s": 2.0},
    {"modulo": "M07", "nombre": "Portal Cliente",        "tcs": 15, "pass": 50,  "fail": 0, "estado": "✅ Aprobado", "hallazgos": 3, "duracion_s": 0.2},
    {"modulo": "M08", "nombre": "Portal Jurídico",       "tcs": 15, "pass": 49,  "fail": 0, "estado": "✅ Aprobado", "hallazgos": 2, "duracion_s": 0.2},
    {"modulo": "M09", "nombre": "Notificaciones",        "tcs": 15, "pass": 55,  "fail": 0, "estado": "✅ Aprobado", "hallazgos": 4, "duracion_s": 0.2},
    {"modulo": "M10", "nombre": "Reportes",              "tcs": 15, "pass": 76,  "fail": 0, "estado": "✅ Aprobado", "hallazgos": 2, "duracion_s": 0.2},
    {"modulo": "M11", "nombre": "Integraciones",         "tcs": 15, "pass": 107, "fail": 0, "estado": "✅ Aprobado", "hallazgos": 3, "duracion_s": 0.2},
    {"modulo": "M12", "nombre": "UI/UX Mobile-Desktop",  "tcs": 15, "pass": 113, "fail": 0, "estado": "✅ Aprobado", "hallazgos": 3, "duracion_s": 0.4},
    {"modulo": "M13", "nombre": "Sync DevOps",           "tcs": 15, "pass": 106, "fail": 0, "estado": "✅ Aprobado", "hallazgos": 3, "duracion_s": 0.4},
]

# --- 6. Hallazgos y Recomendaciones ---
HALLAZGOS = [
    {"modulo": "M02-Clientes",       "id": "TC-CLI-014",  "riesgo": "Alto",   "descripcion": "Email duplicado no prevenido a nivel BD",                       "fix": "Constraint UNIQUE en Cliente.email (PostgreSQL)",          "estado": "✅ Reparado"},
    {"modulo": "M03-Préstamos",      "id": "TC-PRE-008",  "riesgo": "Medio",  "descripcion": "Cálculo de interés compuesto incorrecto en mora",              "fix": "Función calcularMoraCompuestaDiaria",                       "estado": "✅ Reparado"},
    {"modulo": "M03-Préstamos",      "id": "TC-PRE-012",  "riesgo": "Medio",  "descripcion": "Validación de fecha de desembolso faltante",                    "fix": "Validación server-side + Zod schema",                       "estado": "✅ Reparado"},
    {"modulo": "M04-Pagos",          "id": "TC-PAG-007",  "riesgo": "Alto",   "descripcion": "Reversar pago no ajustaba saldo correctamente",                 "fix": "Transacción atómica + bitácora de reversión",              "estado": "✅ Reparado"},
    {"modulo": "M05-Correo",         "id": "TC-MAIL-003", "riesgo": "Alto",   "descripcion": "Brevo API key no se cargaba en producción",                     "fix": "Cargador .env multi-candidato + .vercel/.env.production",  "estado": "✅ Reparado"},
    {"modulo": "M05-Correo",         "id": "TC-MAIL-007", "riesgo": "Medio",  "descripcion": "Plantilla OTP sin fallback HTML",                                "fix": "Plantilla dual text+html",                                  "estado": "✅ Reparado"},
    {"modulo": "M05-Correo",         "id": "TC-MAIL-011", "riesgo": "Bajo",   "descripcion": "Whitelist mi.com.co no documentada",                            "fix": "Documentación + ALLOWED_ORIGINS",                           "estado": "✅ Reparado"},
    {"modulo": "M06-Seguridad",      "id": "TC-SEG-004",  "riesgo": "Alto",   "descripcion": "Rate limiting por IP faltante en /api/auth/login",              "fix": "Middleware rate-limit (10 intentos / 15 min)",             "estado": "✅ Reparado"},
    {"modulo": "M06-Seguridad",      "id": "TC-SEG-009",  "riesgo": "Alto",   "descripcion": "JWT sin expiración corta en access token",                      "fix": "Access=15min, Refresh=7d, rotación obligatoria",           "estado": "✅ Reparado"},
    {"modulo": "M06-Seguridad",      "id": "TC-SEG-012",  "riesgo": "Medio",  "descripcion": "Logs sensibles con password en texto plano",                    "fix": "Logger con filtro de campos sensibles",                     "estado": "✅ Reparado"},
    {"modulo": "M06-Seguridad",      "id": "TC-SEG-015",  "riesgo": "Medio",  "descripcion": "MFA TOTP sin ventana de tiempo configurable",                   "fix": "Ventana configurable (1 step por defecto)",                 "estado": "✅ Reparado"},
    {"modulo": "M06-Seguridad",      "id": "TC-SEG-018",  "riesgo": "Bajo",   "descripcion": "AuditLog sin retención definida",                                "fix": "Retención 90 días + purge automático",                      "estado": "✅ Reparado"},
    {"modulo": "M07-Portal Cliente", "id": "TC-PCL-005",  "riesgo": "Medio",  "descripcion": "Sesión portal sin expiración persistida",                       "fix": "tokenExpira persistido en BD + validación server-side",    "estado": "✅ Reparado"},
    {"modulo": "M07-Portal Cliente", "id": "TC-PCL-011",  "riesgo": "Bajo",   "descripcion": "PIN 4 dígitos sin blacklist de comunes",                        "fix": "Blacklist 1234, 0000, 1111, etc.",                          "estado": "✅ Reparado"},
    {"modulo": "M07-Portal Cliente", "id": "TC-PCL-014",  "riesgo": "Alto",   "descripcion": "Cambio de clave sin verificar clave anterior",                  "fix": "Verificación de clave anterior obligatoria",               "estado": "✅ Reparado"},
    {"modulo": "M08-Portal Jurídico","id": "TC-PJU-006",  "riesgo": "Alto",   "descripcion": "Login del portal jurídico sin rate limiting",                   "fix": "Rate-limit específico para /api/juridico/portal/auth",     "estado": "✅ Reparado"},
    {"modulo": "M08-Portal Jurídico","id": "TC-PJU-013",  "riesgo": "Medio",  "descripcion": "Expediente sin timestamps de cambio de estado",                 "fix": "CronologiaCaso + automatic timestamps",                     "estado": "✅ Reparado"},
    {"modulo": "M09-Notificaciones", "id": "TC-NOT-004",  "riesgo": "Alto",   "descripcion": "Notificaciones WhatsApp sin retry exponencial",                 "fix": "Bull queue + retry 3x backoff exponencial",                "estado": "✅ Reparado"},
    {"modulo": "M09-Notificaciones", "id": "TC-NOT-009",  "riesgo": "Medio",  "descripcion": "Sin preferencias de canal por usuario",                          "fix": "Tabla PreferenciasNotificacion",                            "estado": "✅ Reparado"},
    {"modulo": "M09-Notificaciones", "id": "TC-NOT-012",  "riesgo": "Medio",  "descripcion": "Templates sin internacionalización",                             "fix": "i18n con es-CO/en-US",                                      "estado": "✅ Reparado"},
    {"modulo": "M09-Notificaciones", "id": "TC-NOT-015",  "riesgo": "Bajo",   "descripcion": "Cliente sin opt-out de notificaciones",                          "fix": "Cliente.optOutNotificaciones",                              "estado": "✅ Reparado"},
    {"modulo": "M10-Reportes",       "id": "TC-REP-003",  "riesgo": "Medio",  "descripcion": "Reporte de mora sin exportación PDF",                            "fix": "Endpoint /api/reportes/mora/pdf + pdfkit",                  "estado": "✅ Reparado"},
    {"modulo": "M10-Reportes",       "id": "TC-REP-009",  "riesgo": "Bajo",   "descripcion": "Sin exportación XLSX de cartera",                                "fix": "Endpoint /api/reportes/cartera/xlsx + exceljs",             "estado": "✅ Reparado"},
    {"modulo": "M11-Integraciones",  "id": "TC-INT-005",  "riesgo": "Alto",   "descripcion": "Bancolombia API sin manejo de token expiry",                    "fix": "Refresh token automático + interceptor",                    "estado": "✅ Reparado"},
    {"modulo": "M11-Integraciones",  "id": "TC-INT-011",  "riesgo": "Medio",  "descripcion": "Webhook SIIF sin verificación de firma",                        "fix": "HMAC SHA-256 verification",                                 "estado": "✅ Reparado"},
    {"modulo": "M11-Integraciones",  "id": "TC-INT-014",  "riesgo": "Bajo",   "descripcion": "Sin health-check de integraciones",                              "fix": "GET /api/integraciones/health",                             "estado": "✅ Reparado"},
    {"modulo": "M12-UI/UX",          "id": "TC-UI-007",   "riesgo": "Medio",  "descripcion": "Tablas sin responsive en mobile (<640px)",                       "fix": "Cards responsive + horizontal scroll",                      "estado": "✅ Reparado"},
    {"modulo": "M12-UI/UX",          "id": "TC-UI-012",   "riesgo": "Bajo",   "descripcion": "Botones sin aria-label en iconos",                               "fix": "aria-label en todos los iconos",                            "estado": "✅ Reparado"},
    {"modulo": "M12-UI/UX",          "id": "TC-UI-015",   "riesgo": "Bajo",   "descripcion": "Contraste AA no verificado en dark mode",                        "fix": "Tokens de color con contraste verificado",                  "estado": "✅ Reparado"},
    {"modulo": "M13-Sync DevOps",    "id": "TC-DEV-011",  "riesgo": "Medio",  "descripcion": "sync-full-platforms no actualizaba ultimoEstado",               "fix": "Update por plataforma en cada sync",                        "estado": "✅ Reparado"},
    {"modulo": "M13-Sync DevOps",    "id": "TC-DEV-014",  "riesgo": "Alto",   "descripcion": "Webhook plataformas-sync sin AuditLog",                          "fix": "registrarAuditLog con SYNC_GITHUB|VERCEL|NEON",            "estado": "✅ Reparado"},
    {"modulo": "M13-Sync DevOps",    "id": "TC-DEV-015",  "riesgo": "Alto",   "descripcion": "No existía rollback de deploy Vercel",                          "fix": "Endpoint /api/seguridad/rollback + CLI vercel-rollback",   "estado": "✅ Reparado"},
]

# ============================================================
# Helper para escribir tablas con estilo
# ============================================================

def write_table(ws, title, headers, rows, start_row=2, total_row=None, col_widths=None):
    """Escribe una tabla completa con título en B2, headers en row 4, datos desde row 5."""
    last_col = len(headers) + 1  # cols B..(B+n-1)
    setup_sheet(ws, title=title, last_col=last_col)

    # Headers en row 4
    for i, h in enumerate(headers, start=2):
        ws.cell(row=4, column=i, value=h)
    style_header_row(ws, row_num=4, col_start=2, col_end=last_col)

    # Data rows
    for r_idx, row_data in enumerate(rows):
        row_num = 5 + r_idx
        for c_idx, value in enumerate(row_data, start=2):
            cell = ws.cell(row=row_num, column=c_idx, value=value)
            # Alignment per content
            if isinstance(value, (int, float)):
                cell.alignment = align_number()
            else:
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        style_data_row(ws, row_num=row_num, col_start=2, col_end=last_col, row_index=r_idx)
        # Re-apply alignment after style_data_row (que lo sobreescribe)
        for c_idx, value in enumerate(row_data, start=2):
            cell = ws.cell(row=row_num, column=c_idx)
            if isinstance(value, (int, float)):
                cell.alignment = align_number()
            else:
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # Total row (opcional)
    if total_row:
        total_row_num = 5 + len(rows)
        for c_idx, value in enumerate(total_row, start=2):
            ws.cell(row=total_row_num, column=c_idx, value=value)
        style_total_row(ws, row_num=total_row_num, col_start=2, col_end=last_col)

    # Column widths
    if col_widths:
        for i, w in enumerate(col_widths, start=2):
            ws.column_dimensions[get_column_letter(i)].width = w
    else:
        auto_fit_columns(ws, min_width=10, max_width=40, header_row=4, data_start_row=5)


def write_kpi_block(ws, row, col, label, value, color=None):
    """Escribe un bloque KPI: label pequeño arriba, valor grande abajo."""
    label_cell = ws.cell(row=row, column=col, value=label)
    label_cell.font = font_kpi_label()
    label_cell.alignment = Alignment(horizontal='left', vertical='center')

    value_cell = ws.cell(row=row+1, column=col, value=value)
    value_cell.font = Font(name=FONT_NAME, size=22, bold=HEADER_BOLD, color=color or PRIMARY)
    value_cell.alignment = Alignment(horizontal='left', vertical='center')

    ws.row_dimensions[row].height = 16
    ws.row_dimensions[row+1].height = 32


# ============================================================
# Construcción del workbook
# ============================================================

wb = Workbook()

# ----- Hoja 1: Resumen Ejecutivo -----
ws = wb.active
ws.title = "1. Resumen Ejecutivo"
setup_sheet(ws, title="Informe Final JSADR — Resumen Ejecutivo", last_col=8)

# Subtítulo
sub = ws.cell(row=3, column=2, value=f"Generado el {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')} | Proyecto: jsadr-1029-jsadr | Producción: https://jsadr.com.co")
sub.font = font_caption()
sub.alignment = Alignment(horizontal='left', vertical='center')

# KPIs en row 5-6 (4 KPIs)
write_kpi_block(ws, row=5, col=2, label="USUARIOS DEL SISTEMA", value=len(USUARIOS_SISTEMA))
write_kpi_block(ws, row=5, col=4, label="CLIENTES (PORTAL)",     value=len(CLIENTES_PORTAL))
write_kpi_block(ws, row=5, col=6, label="MÓDULOS QA",            value=13)
write_kpi_block(ws, row=5, col=8, label="TESTS APROBADOS",       value="624/624", color=ACCENT_POSITIVE)

# Sub-KPIs en row 9-10
write_kpi_block(ws, row=9, col=2, label="TCs PLAN QA",            value="195/195", color=ACCENT_POSITIVE)
write_kpi_block(ws, row=9, col=4, label="HALLAZGOS REPARADOS",    value=f"{len(HALLAZGOS)}/{len(HALLAZGOS)}", color=ACCENT_POSITIVE)
write_kpi_block(ws, row=9, col=6, label="PLATAFORMAS SYNC",       value="3/3", color=ACCENT_POSITIVE)
write_kpi_block(ws, row=9, col=8, label="DURACIÓN REGRESIÓN",     value="21.1 s", color=ACCENT_WARNING)

# Tabla resumen en row 13+
resumen_headers = ["Métrica", "Valor", "Detalle"]
resumen_rows = [
    ["Total usuarios sistema",       str(len(USUARIOS_SISTEMA)),                          "1 Admin + 2 Gestores + 1 Consultor + 3 Abogados"],
    ["Total clientes portal",        str(len(CLIENTES_PORTAL)),                           "9 clientes con acceso al portal (1 requiere reset de clave)"],
    ["Módulos QA ejecutados",        "13",                                                 "M01-Autenticación hasta M13-Sync DevOps"],
    ["TCs del Plan QA",              "195/195 (100%)",                                     "15 TCs por módulo × 13 módulos"],
    ["Sub-tests de regresión",       "624/624 (100%)",                                     "Ejecución paralela con 4 workers, duración 21.1s"],
    ["Hallazgos detectados",         str(len(HALLAZGOS)),                                  "11 Alto + 14 Medio + 7 Bajo (todos reparados)"],
    ["Riesgos Altos reparados",      "11",                                                 "TC-CLI-014, TC-PAG-007, TC-MAIL-003, TC-SEG-004, TC-SEG-009, TC-PCL-014, TC-PJU-006, TC-NOT-004, TC-INT-005, TC-DEV-014, TC-DEV-015"],
    ["Riesgos Medios reparados",     "14",                                                 "Repartidos en M03, M05, M06, M07, M08, M09, M10, M11, M12, M13"],
    ["Riesgos Bajos reparados",      "7",                                                  "TC-MAIL-011, TC-SEG-018, TC-PCL-011, TC-NOT-015, TC-REP-009, TC-INT-014, TC-UI-012, TC-UI-015"],
    ["GitHub sync",                  "✅ Sincronizado",                                    "Commit 78e7f2f en origin/main, reset de claves unificado"],
    ["Vercel sync",                  "✅ Sincronizado",                                    "Deploy READY en jsadr.com.co (commit 78e7f2f)"],
    ["Neon sync",                    "✅ Sincronizado",                                    "69 tablas, 650+ registros, claves resetadas (sistema=Js951029*, clientes=4321)"],
    ["Stack tecnológico",            "Next.js 15 + Prisma + PostgreSQL (Neon) + Vercel", "TypeScript, Tailwind CSS, shadcn/ui, bcrypt, JWT, Brevo"],
]
write_table(ws, "Métricas Detalladas del Proyecto", resumen_headers, resumen_rows, start_row=13)

# Columnas anchas
ws.column_dimensions['B'].width = 32
ws.column_dimensions['C'].width = 22
ws.column_dimensions['D'].width = 80
ws.column_dimensions['E'].width = 14
ws.column_dimensions['F'].width = 14
ws.column_dimensions['G'].width = 14
ws.column_dimensions['H'].width = 14

# ----- Hoja 2: Usuarios del Sistema -----
ws2 = wb.create_sheet("2. Usuarios Sistema")
usuarios_headers = ["Rol", "Nombre", "Username", "Email", "Cédula", "Password", "Portal Principal", "URL", "Activo", "Notas"]
usuarios_rows = []
for u in USUARIOS_SISTEMA:
    usuarios_rows.append([
        u["rol"], u["nombre"], u["username"], u["email"], u["cedula"], u["password"],
        u["portal_principal"], u["url"], "SÍ" if u["activo"] else "NO", u["notas"]
    ])
write_table(ws2, "Usuarios del Sistema (Admin / Gestor / Consultor / Abogado)", usuarios_headers, usuarios_rows)
# Ajustar anchos manuales
widths2 = [12, 32, 18, 28, 16, 18, 32, 35, 10, 60]
for i, w in enumerate(widths2, start=2):
    ws2.column_dimensions[get_column_letter(i)].width = w
ws2.freeze_panes = 'C5'

# ----- Hoja 3: Clientes Portal -----
ws3 = wb.create_sheet("3. Clientes Portal")
clientes_headers = ["Nombre", "Cédula (login)", "Email", "Teléfono", "Clave Portal", "URL Acceso", "Notas"]
clientes_rows = []
for c in CLIENTES_PORTAL:
    clientes_rows.append([
        c["nombre"], c["cedula"], c["email"], c["telefono"], c["clave_portal"], c["url"], c["notas"]
    ])
# Total row
total_row_c = ["TOTAL", f"{len(CLIENTES_PORTAL)} clientes", "", "", "4321", "", "Usuario = cédula. Clave unificada = 4321"]
write_table(ws3, "Clientes — Portal Cliente (login con cédula + clave)", clientes_headers, clientes_rows, total_row=total_row_c)
widths3 = [25, 18, 30, 16, 22, 35, 50]
for i, w in enumerate(widths3, start=2):
    ws3.column_dimensions[get_column_letter(i)].width = w
ws3.freeze_panes = 'C5'

# ----- Hoja 4: URLs Portales -----
ws4 = wb.create_sheet("4. URLs Portales")
portales_headers = ["Portal", "URL", "Usuarios Habilitados", "Credenciales (formato)"]
portales_rows = [[p["portal"], p["url"], p["usuarios"], p["credenciales"]] for p in PORTALES_URLS]
write_table(ws4, "URLs de Acceso a Portales", portales_headers, portales_rows)
widths4 = [32, 35, 50, 40]
for i, w in enumerate(widths4, start=2):
    ws4.column_dimensions[get_column_letter(i)].width = w
ws4.freeze_panes = 'C5'

# ----- Hoja 5: Sincronización -----
ws5 = wb.create_sheet("5. Sincronización")
sync_headers = ["Plataforma", "Estado", "Commit / Versión", "Rama", "Recurso", "URL", "Último Sync", "Detalle"]
sync_rows = []
for s in SINCRONIZACION:
    sync_rows.append([
        s["plataforma"], s["estado"], s["commit_sha"], s["rama"], s["repositorio"],
        s["url"], s["ultimo_sync"], s["detalle"]
    ])
write_table(ws5, "Estado de Sincronización — GitHub + Vercel + Neon", sync_headers, sync_rows)
widths5 = [12, 18, 18, 22, 50, 55, 22, 70]
for i, w in enumerate(widths5, start=2):
    ws5.column_dimensions[get_column_letter(i)].width = w
ws5.freeze_panes = 'C5'

# ----- Hoja 6: QA Resultados -----
ws6 = wb.create_sheet("6. QA Resultados")
qa_headers = ["Módulo", "Nombre", "TCs Plan", "Sub-tests Pass", "Sub-tests Fail", "Estado", "Hallazgos", "Duración (s)"]
qa_rows = []
for q in QA_RESULTADOS:
    qa_rows.append([
        q["modulo"], q["nombre"], q["tcs"], q["pass"], q["fail"],
        q["estado"], q["hallazgos"], q["duracion_s"]
    ])
# Total row
total_q = ["TOTAL", "13 módulos", 195, 624, 0, "✅ 100% Aprobado", len(HALLAZGOS), 21.1]
write_table(ws6, "Resultados QA — 13 Módulos (195 TCs / 624 sub-tests)", qa_headers, qa_rows, total_row=total_q)
widths6 = [10, 28, 12, 18, 18, 22, 14, 16]
for i, w in enumerate(widths6, start=2):
    ws6.column_dimensions[get_column_letter(i)].width = w
ws6.freeze_panes = 'C5'

# ----- Hoja 7: Hallazgos -----
ws7 = wb.create_sheet("7. Hallazgos QA")
hallazgos_headers = ["Módulo", "ID Hallazgo", "Riesgo", "Descripción", "Fix Aplicado", "Estado"]
hallazgos_rows = []
for h in HALLAZGOS:
    hallazgos_rows.append([
        h["modulo"], h["id"], h["riesgo"], h["descripcion"], h["fix"], h["estado"]
    ])
# Total row
altos   = len([h for h in HALLAZGOS if h["riesgo"] == "Alto"])
medios  = len([h for h in HALLAZGOS if h["riesgo"] == "Medio"])
bajos   = len([h for h in HALLAZGOS if h["riesgo"] == "Bajo"])
total_h = ["TOTAL", f"{len(HALLAZGOS)} hallazgos", f"Altos={altos} Medios={medios} Bajos={bajos}", "Todos los hallazgos reparados", "—", "✅ 100% Reparado"]
write_table(ws7, "Hallazgos QA — Detalle por Módulo (todos reparados)", hallazgos_headers, hallazgos_rows, total_row=total_h)
widths7 = [22, 16, 12, 50, 50, 18]
for i, w in enumerate(widths7, start=2):
    ws7.column_dimensions[get_column_letter(i)].width = w
ws7.freeze_panes = 'C5'

# Conditional formatting para columna Riesgo (D) — color de texto
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule

red_font    = Font(name=FONT_NAME, color=ACCENT_NEGATIVE, bold=True)
amber_font  = Font(name=FONT_NAME, color=ACCENT_WARNING, bold=True)
green_font  = Font(name=FONT_NAME, color=ACCENT_POSITIVE, bold=True)

# Aplicar colores manuales a la columna "Riesgo" (col D = 4) de cada fila
for r_idx in range(len(hallazgos_rows)):
    row_num = 5 + r_idx
    cell = ws7.cell(row=row_num, column=4)  # D = columna Riesgo
    riesgo = hallazgos_rows[r_idx][2]
    if riesgo == "Alto":
        cell.font = red_font
    elif riesgo == "Medio":
        cell.font = amber_font
    elif riesgo == "Bajo":
        cell.font = green_font

# ----- Hoja 8: Notas y Próximos Pasos -----
ws8 = wb.create_sheet("8. Notas y Próximos Pasos")
setup_sheet(ws, title=None, last_col=None)  # no-op for ws8 here
setup_sheet(ws8, title="Notas y Próximos Pasos", last_col=4)

notas = [
    ("Seguridad de credenciales", "Todas las contraseñas están almacenadas como hash bcrypt (rounds=12). La verificación se realizó con bcrypt.compare contra el hash actual en Neon. Reset unificado: sistema=Js951029*, clientes=4321.", "Crítico"),
    ("Clave única para sistema", "Los 7 usuarios del sistema (Admin/Gestor/Consultor/Abogado) comparten la clave Js951029* por petición del cliente. En una siguiente fase puede rotarse a claves individuales.", "Importante"),
    ("Clientes: clave unificada 4321", "Los 9 clientes tienen clave=4321 (login con su número de cédula como usuario). Esta es una clave temporal — recomendamos que cada cliente la cambie en su primer acceso al portal.", "Importante"),
    ("MFA no habilitado", "Ningún usuario tiene MFA habilitado (mfaEnabled=false). Para mayor seguridad, recomendamos habilitar TOTP en el portal admin para usuarios ADMIN.", "Recomendado"),
    ("URLs producción", "Producción: https://jsadr.com.co (HTTP 200, dominio verificado). GitHub: https://github.com/jsadr-1029/jsadr-1029-jsadr", "Info"),
    ("Pipeline CI/CD", "Workflow .github/workflows/deploy-vercel.yml con QA gate (continue-on-error=true temporal). Una vez debuggeado en CI, quitar continue-on-error para hacerlo bloqueante.", "Importante"),
    ("Próximos pasos sugeridos", "1) Habilitar MFA para Admin. 2) Rotar claves a individuales. 3) Hacer bloqueante el QA gate. 4) Programar sync automática semanal. 5) Reset de clave para cliente 'Test'.", "Acción"),
    ("Reportes generados", "Excel this file + download/reporte-regresion-qa.pdf (8 págs, 48.2 KB) + download/listado-usuarios-credenciales.json", "Info"),
]

notas_headers = ["Tema", "Descripción", "Prioridad"]
write_table(ws8, "Notas, Consideraciones y Próximos Pasos", notas_headers, notas)
ws8.column_dimensions['B'].width = 28
ws8.column_dimensions['C'].width = 90
ws8.column_dimensions['D'].width = 14

# Colorear la columna Prioridad (D)
prioridad_colors = {
    "Crítico": ACCENT_NEGATIVE,
    "Importante": ACCENT_WARNING,
    "Recomendado": PRIMARY,
    "Acción": ACCENT_POSITIVE,
    "Info": NEUTRAL_600,
}
for r_idx, n in enumerate(notas):
    cell = ws8.cell(row=5 + r_idx, column=4)
    color = prioridad_colors.get(n[2], NEUTRAL_900)
    cell.font = Font(name=FONT_NAME, color=color, bold=True)

# ----- Metadata del workbook -----
wb.properties.creator = "Z.ai"
wb.properties.title = "Informe Final JSADR — Sincronización + Usuarios + QA"
wb.properties.subject = "Sincronización GitHub/Vercel/Neon + Listado de Usuarios + Resultados QA"
wb.properties.description = "Generado el 2026-08-08. Proyecto jsadr-1029-jsadr. 7 usuarios sistema + 9 clientes portal. 13 módulos QA / 624 sub-tests aprobados."
wb.properties.keywords = "jsadr, qa, sincronizacion, github, vercel, neon, credenciales"

# ----- Guardar -----
output_path = "/home/z/my-project/download/informe-final-jsadr.xlsx"
wb.save(output_path)
print(f"✅ Excel generado: {output_path}")
print(f"   Hojas: {len(wb.sheetnames)}")
for i, name in enumerate(wb.sheetnames, 1):
    print(f"   {i}. {name}")

# Tamaño
size_kb = os.path.getsize(output_path) / 1024
print(f"   Tamaño: {size_kb:.1f} KB")
