"""Actualiza el Excel con los hallazgos y reparos de M06-Seguridad."""
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
import shutil

WB_PATH = "/home/z/my-project/upload/plan-pruebas-qa-jsadr.xlsx"
DEST_PATH = "/home/z/my-project/download/plan-pruebas-qa-jsadr-actualizado.xlsx"

# Sheet "8. M06-Seguridad" rows 6, 7, 13, 17, 18, 19
updates = {
    "8. M06-Seguridad": [
        (6, "TC-SEC-002",
         "HTTP 403 FORBIDDEN con code=FORBIDDEN. CONSULTOR rechazado al mutar clientes.",
         "REPARO v4.9 CRÍTICO DE SEGURIDAD: ANTES POST /api/clientes NO tenía requireRole. Cualquier usuario autenticado (incluso CONSULTOR) podía crear clientes. Lo mismo para PUT y PATCH en /api/clientes/[id]. AHORA: GET permite ADMIN/GESTOR/CONSULTOR (lectura), POST y PUT solo ADMIN/GESTOR, PATCH (activar/desactivar) solo ADMIN. CONSULTOR recibe HTTP 403 FORBIDDEN al intentar mutar."),
        (7, "TC-SEC-003",
         "HTTP 403 FORBIDDEN. GESTOR rechazado al reversar pagos (solo ADMIN).",
         "Ya reparado en v4.6 (TC-PRE-015): requireRole(req, ['ADMIN']) en /api/pagos/[id]/reversar. GESTOR recibe 403 FORBIDDEN."),
        (13, "TC-SEC-009",
         "HTTP 500 con mensaje genérico 'Ocurrió un error procesando la solicitud'. Stack trace NO expuesto. internalDetails solo en logs.",
         "sanitizeError ya implementado en src/lib/error-handler.ts: mapea errores Prisma (P2002 DUPLICATE_ENTRY, P2025 NOT_FOUND, etc.) y errores de aplicación a mensajes seguros. Error genérico retorna code=INTERNAL_ERROR con mensaje 'Ocurrió un error procesando la solicitud'. internalDetails se separa del message (no se envía al cliente). logError() y errorResponse() helpers para logs internos."),
        (17, "TC-SEC-013",
         "verifyTOTP valida token TOTP de 6 dígitos con ventana de ±30s. Comparación timing-safe (crypto.timingSafeEqual).",
         "TOTP RFC 6238 propio implementado en src/lib/totp.ts (sin dependencias externas): generateSecret (160 bits, base32 RFC 4648), generateTOTP (HMAC-SHA1, 6 dígitos, step 30s), verifyTOTP (ventana ±1 step, timing-safe), generateURI (otpauth:// compatible Google Authenticator, Authy, 1Password). Prueba funcional: secret generado, token generado y verificado correctamente."),
        (18, "TC-SEC-014",
         "HTTP 200. Credenciales eliminadas de BD (ConexionAPI.password/apiKey null) y de Vercel env vars. Audit log registrado.",
         "POST /api/seguridad/credenciales/eliminar ya implementado. Auth requireRole(['ADMIN']). Clave maestra 'Eliminar' (constante del backend). Plataformas soportadas: BREVO_SMTP (limpia ConexionAPI.password + BREVO_SMTP_KEY en Vercel), BREVO_API (limpia ConexionAPI.apiKey + BREVO_API_KEY en Vercel), VERCEL, GITHUB, NEON. Intento fallido (clave incorrecta) registra audit log CREDENCIAL_ELIMINAR_INTENTO_FALLIDO. Eliminación exitosa registra CREDENCIAL_ELIMINADA. BD: 3 PlataformaSync configuradas."),
        (19, "TC-SEC-015",
         "HTTP 403 CORS_ORIGEN_NO_PERMITIDO. Origin no permitido rechazado explícitamente. Headers Access-Control-Allow-Origin NO devueltos.",
         "REPARO v4.9: ANTES CORS preflight desde dominio no permitido retornaba HTTP 204 (éxito) sin headers CORS, lo cual era confuso (el navegador lo interpretaba como éxito pero bloqueaba la petición real). AHORA: retorna HTTP 403 Forbidden explícito con codigo CORS_ORIGEN_NO_PERMITIDO y mensaje claro indicando qué origin fue rechazado. Whitelist: localhost, *.space-z.ai, *.vercel.app (overridable con ALLOWED_ORIGINS env var). CSRF check también valida Origin en mutaciones (CSRF_DENIED)."),
    ],
}

green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
green_font = Font(color="006100", bold=True)
yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

wb = load_workbook(WB_PATH)
total_updated = 0

for sheet_name, items in updates.items():
    ws = wb[sheet_name]
    for row_num, tc_id, nuevo_riesgo, nuevo_hallazgo in items:
        cell_id = ws.cell(row=row_num, column=2).value
        if cell_id != tc_id:
            print(f"⚠ {sheet_name} fila {row_num}: esperado {tc_id}, encontrado {cell_id}")
            continue

        cell_state = ws.cell(row=row_num, column=13)
        cell_state.value = "Aprobado"
        cell_state.fill = green_fill
        cell_state.font = green_font
        cell_state.alignment = Alignment(horizontal="center", vertical="center")

        cell_riesgo = ws.cell(row=row_num, column=11)
        cell_riesgo.value = nuevo_riesgo
        cell_riesgo.fill = yellow_fill
        cell_riesgo.alignment = Alignment(wrap_text=True, vertical="top")

        cell_hallazgo = ws.cell(row=row_num, column=12)
        cell_hallazgo.value = nuevo_hallazgo
        cell_hallazgo.fill = yellow_fill
        cell_hallazgo.alignment = Alignment(wrap_text=True, vertical="top")

        print(f"✓ {sheet_name} fila {row_num} ({tc_id}): Estado=Aprobado")
        total_updated += 1

wb.save(WB_PATH)
print(f"\n=== Total actualizado: {total_updated} TCs ===")

shutil.copy(WB_PATH, DEST_PATH)
print(f"Copia para entrega: {DEST_PATH}")
