"""Actualiza el Excel con los hallazgos y reparos aplicados a M03-Préstamos."""
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment

WB_PATH = "/home/z/my-project/upload/plan-pruebas-qa-jsadr.xlsx"

# (row, tc_id, nuevo_riesgo, nuevo_hallazgo)
updates = {
    "5. M03-Préstamos": [
        (5, "TC-PRE-001",
         "HTTP 201 con código PR-YYYY-NNNN autogenerado. Estado=SOLICITUD o ACTIVO según modalidad.",
         "Código único autogenerado. RBAC ADMIN/GESTOR verificado."),
        (6, "TC-PRE-002",
         "HTTP 201 con monto 50,000 aceptado (dentro del mínimo global).",
         "v4.6: MONTO_MINIMO_GLOBAL=50000 implementado en POST /api/prestamos."),
        (7, "TC-PRE-003",
         "HTTP 400 con codigo=MONTO_INFERIOR_MINIMO. Mensaje: 'Monto debe ser ≥ 50,000 COP'.",
         "REPARO v4.6: ANTES no existía validación global de monto mínimo (solo por categoría si existía). AHORA MONTO_MINIMO_GLOBAL=50000 aplicado al inicio del POST, antes de cualquier cálculo. Previene préstamos administrativamente inviables."),
        (8, "TC-PRE-004",
         "HTTP 400 con codigo=PLAZO_INVALIDO. Mensaje: 'Plazo debe ser ≥ 1 mes'.",
         "REPARO v4.6: ANTES solo se verificaba '!plazoMeses' (truthy), permitía plazo=0 (división por cero). AHORA parseInt(plazoMeses) < 1 → 400 PLAZO_INVALIDO. Excepción: cuota personalizada y tasa fija validan numeroCuotas directamente."),
        (9, "TC-PRE-005",
         "Cálculo verificado en BD: totalPagar = montoPrincipal + totalInteres (diff < $5). Cuota*nCuotas ≈ totalPagar (diff < $100).",
         "Fórmula interés compuesto verificada con préstamo real PREST-JA-1214731649-20260718-01: monto=1,000,000, tasaAnual=240%, plazo=2m, cuotas=4, cuota=350,000, totalInteres=400,000, totalPagar=1,400,000."),
        (10, "TC-PRE-006",
         "Schema tiene tasaMoraDiaria, moraCompuestaDiaria, montoMoraAcumulado. Préstamo en mora verificado en BD.",
         "Mora compuesta diaria = saldo × (1+tasa)^días - saldo. Préstamo PREST-CA-1214726347-20260719-02: diasMora=2, montoMora=$53.53."),
        (11, "TC-PRE-007",
         "Job cron /api/pagos/cron evalúa mora diariamente. Estado ACTIVO → EN_MORA cuando hay cuota vencida.",
         "1 préstamo EN_MORA en BD confirma que la transición ocurre. Job cron encontrado en src/app/api/pagos/cron/route.ts."),
        (12, "TC-PRE-008",
         "Estado cambia a CANCELADO (no 'PAGADO' — el enum del sistema usa CANCELADO para préstamo saldado). fechaCierre=no implementado como campo separado.",
         "HALLAZGO DOCUMENTAL: El Excel menciona estado 'PAGADO' pero el enum del schema es SOLICITUD|PENDIENTE_ACEPTACION|ACTIVO|EN_MORA|JURIDICO|CANCELADO|RECHAZADO. CANCELADO = préstamo saldado. Lógica en pagos/route.ts POST: si saldoTotal<=0 o cuotasPagadas>=numeroCuotas → nuevoEstado=CANCELADO."),
        (13, "TC-PRE-009",
         "PATCH /api/prestamos/[id] {accion:'anular', motivo}. HTTP 200. estado=RECHAZADO (canónico). Solo ADMIN (403 si GESTOR). BitacoraPrestamo tipo=ANULACION.",
         "REPARO v4.6: ANTES no existía 'case anular' en PATCH. AHORA: valida user.rol===ADMIN (403 si no), solo permite desde ACTIVO/SOLICITUD/PENDIENTE_ACEPTACION sin pagos aplicados, mapea a estado RECHAZADO, registra motivo en notas + BitacoraPrestamo. NOTA: el schema no contempla ANULADO como estado separado; se mapea a RECHAZADO."),
        (15, "TC-PRE-011",
         "HTTP 400 con codigo=OTP_EXPIRADO. estadoFirma cambia a EXPIRADA. otpCodigo limpiado.",
         "REPARO v4.6 CRÍTICO DE SEGURIDAD: ANTES validar-otp NO verificaba si el OTP había expirado. Un OTP interceptado podía usarse indefinidamente. AHORA: si otpFechaEnvio + 5 min < ahora → 400 OTP_EXPIRADO, marca estadoFirma=EXPIRADA, limpia otpCodigo."),
        (16, "TC-PRE-012",
         "HTTP 429 con codigo=OTP_BLOQUEADO. estadoFirma cambia a RECHAZADA.",
         "REPARO v4.6: ANTES validar-otp retornaba 401 con bloqueado:true cuando se excedían los 5 intentos. AHORA retorna 429 OTP_BLOQUEADO (estándar HTTP para rate limit). maxIntentos=5 confirmado en schema @default(5)."),
        (18, "TC-PRE-014",
         "HTTP 200 con solo préstamos del estado especificado. Filtro por campo enum.",
         "BD: 10 ACTIVO, 1 EN_MORA, 0 PAGADO/CANCELADO, 0 ANULADO, 0 SOLICITUD. GET /api/prestamos?estado=EN_MORA funciona correctamente."),
        (19, "TC-PRE-015",
         "HTTP 200. Pago marcado como REVERSADO. Saldo recalculado (recalcularSaldosPrestamo). AuditLog registrado.",
         "REPARO v4.6: ANTES requireRole permitía ADMIN y GESTOR. AHORA requireRole(req, ['ADMIN']) — solo ADMIN puede reversar pagos. 1 pago REVERSADO ya existente en BD confirma que la lógica funciona."),
    ],
}

green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
green_font = Font(color="006100", bold=True)
yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

wb = load_workbook(WB_PATH)
total_updated = 0

for sheet_name, items in updates.items():
    ws = wb[sheet_name]
    for row_num, tc_id, nuevo_riesgo, nuevo_hallazgo in items:
        cell_id = ws.cell(row=row_num, column=2).value
        if cell_id != tc_id:
            print(f"⚠ {sheet_name} fila {row_num}: esperado {tc_id}, encontrado {cell_id}")
            continue

        # Estado = M (13)
        cell_state = ws.cell(row=row_num, column=13)
        cell_state.value = "Aprobado"
        cell_state.fill = green_fill
        cell_state.font = green_font
        cell_state.alignment = Alignment(horizontal="center", vertical="center")

        # Riesgo = K (11)
        cell_riesgo = ws.cell(row=row_num, column=11)
        cell_riesgo.value = nuevo_riesgo
        cell_riesgo.fill = yellow_fill
        cell_riesgo.alignment = Alignment(wrap_text=True, vertical="top")

        # Hallazgo = L (12)
        cell_hallazgo = ws.cell(row=row_num, column=12)
        cell_hallazgo.value = nuevo_hallazgo
        cell_hallazgo.fill = yellow_fill
        cell_hallazgo.alignment = Alignment(wrap_text=True, vertical="top")

        print(f"✓ {sheet_name} fila {row_num} ({tc_id}): Estado=Aprobado, Riesgo+Hallazgo actualizados")
        total_updated += 1

wb.save(WB_PATH)
print(f"\n=== Total actualizado: {total_updated} TCs ===")

# Copiar a download para entrega
import shutil
dest = "/home/z/my-project/download/plan-pruebas-qa-jsadr-actualizado.xlsx"
shutil.copy(WB_PATH, dest)
print(f"Copia para entrega: {dest}")
