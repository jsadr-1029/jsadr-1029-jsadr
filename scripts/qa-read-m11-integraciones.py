#!/usr/bin/env python3
"""Lee la hoja M11-Integraciones del Excel de plan de pruebas."""
from openpyxl import load_workbook

wb = load_workbook("/home/z/my-project/download/plan-pruebas-qa-jsadr-actualizado.xlsx", data_only=False)
print("Hojas disponibles:")
for s in wb.sheetnames:
    print(f"  - {s}")

# Buscar hoja M11
target = None
for s in wb.sheetnames:
    if "M11" in s or "Integrac" in s.lower():
        target = s
        break
print(f"\nHoja objetivo: {target}")
if not target:
    raise SystemExit("No se encontro hoja M11-Integraciones")

ws = wb[target]
print(f"Dimensiones: {ws.max_row} filas x {ws.max_column} columnas\n")

# Imprimir encabezados (fila 1)
print("=== Encabezados (fila 1) ===")
for col in range(1, ws.max_column + 1):
    val = ws.cell(row=1, column=col).value
    print(f"  Col {col}: {val}")

print("\n=== Filas de datos ===")
for row in range(2, ws.max_row + 1):
    fila = []
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=row, column=col).value
        fila.append(str(v) if v is not None else "")
    if any(fila):
        print(f"Fila {row}: {' | '.join(fila)}")
