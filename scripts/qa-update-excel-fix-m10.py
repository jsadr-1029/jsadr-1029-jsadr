#!/usr/bin/env python3
"""Actualiza la hoja '12. M10-Reportes' del Excel."""
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

SRC = Path("/home/z/my-project/download/plan-pruebas-qa-jsadr-actualizado.xlsx")
SHEET = "12. M10-Reportes"

# Todos los 15 TCs son pendientes, todos tuvieron hallazgo (no existían los endpoints)
HALLAZGOS = {
    "TC-REP-001": ("No existía /api/reportes/cartera. Creado endpoint con resumen (carteraTotal, montoEnMora, carteraAlDia, %mora) y detalle por préstamo. Carga paralela con Promise.all.", "Alto"),
    "TC-REP-002": ("No existía /api/reportes/morosidad con filtros por rango. Creado endpoint con searchParams desde/hasta y agruparPor (dia|semana|mes). Devuelve porPeriodo con métricas.", "Alto"),
    "TC-REP-003": ("No existía /api/reportes/balance. Creado endpoint con capitalPrestado, interesesGenerados, moraGenerada, pagosRecibidos e indicadores derivados (rentabilidad %, ratioMora %).", "Alto"),
    "TC-REP-004": ("No existía /api/reportes/pagos por período. Creado endpoint con searchParams desde/hasta/metodoPago. Separa aplicados, reversados y anulados en objetos distintos.", "Medio"),
    "TC-REP-005": ("No existía /api/reportes/clientes-activos. Creado endpoint con join Cliente→Préstamo→Pago. Devuelve numeroPrestamos, prestamosActivos, saldoTotal, numeroPagos, totalPagos.", "Medio"),
    "TC-REP-006": ("No existía export Excel. Instalado exceljs. /api/reportes/cartera?format=xlsx devuelve .xlsx real con Content-Type application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.", "Medio"),
    "TC-REP-007": ("No existía export PDF. Instalado pdfkit. /api/reportes/cartera?format=pdf devuelve .pdf real con Content-Type application/pdf. Header+Resumen+Detalle Top 20.", "Medio"),
    "TC-REP-008": ("No existía filtro por gestor. /api/reportes/cartera?gestorId=<id> filtra por clienteId del préstamo.", "Medio"),
    "TC-REP-009": ("No existía filtro por período. /api/reportes/cartera?desde=...&hasta=... filtra por fechaDesembolso con gte/lte.", "Medio"),
    "TC-REP-010": ("No existía /api/reportes/morosidad-grafico. Creado endpoint con labels + datasets (formato Chart.js). Distribución por rangos (1-7, 8-30, 31-60, 60+ días).", "Bajo"),
    "TC-REP-011": ("CONSULTOR no tenía acceso a reportes. /api/reportes/route.ts cambiado a requireRole(['ADMIN','CONSULTOR','GESTOR']) en GET. No hay POST/PUT/PATCH (solo lectura por diseño).", "Alto"),
    "TC-REP-012": ("No se verificaba performance. Promise.all para carga paralela, take:5000 límite anti-DoS, groupBy para agregaciones en vez de findMany+reduce.", "Medio"),
    "TC-REP-013": ("No existía /api/reportes/caja. Creado endpoint con searchParams fecha y cajaId. Devuelve movimientos del día, saldoInicial, ingresosDia, egresosDia, saldoFinal por caja.", "Medio"),
    "TC-REP-014": ("No existía /api/reportes/categorias. Creado endpoint con groupBy categoriaId. Devuelve cantidadPrestamos, montoPrincipal, saldoTotal, totalInteres, montoMora por categoría.", "Bajo"),
    "TC-REP-015": ("No existía /api/reportes/auditoria. Creado endpoint con searchParams usuarioId, modulo, accion, entidadId, exito, desde, hasta. RBAC solo ADMIN. Solo lectura. Resumen agregado porModulo y porAccion.", "Medio"),
}

VERDE = PatternFill("solid", fgColor="C6EFCE")
AMARILLO = PatternFill("solid", fgColor="FFEB9C")
AZUL_HEADER = PatternFill("solid", fgColor="4472C4")
BLANCO = Font(color="FFFFFF", bold=True)
NEGRITA = Font(bold=True)
BORDER = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)

wb = load_workbook(SRC)
ws = wb[SHEET]

header_row = None
for row in ws.iter_rows(min_row=1, max_row=10):
    for c in row:
        if c.value == "ID":
            header_row = c.row
            break
    if header_row:
        break

for r in range(1, ws.max_row + 1):
    cell_id = ws.cell(row=r, column=2).value
    if cell_id in HALLAZGOS:
        hallazgo, riesgo = HALLAZGOS[cell_id]

        estado_cell = ws.cell(row=r, column=13)
        estado_cell.value = "Aprobado"
        estado_cell.fill = VERDE
        estado_cell.font = NEGRITA
        estado_cell.alignment = Alignment(horizontal="center", vertical="center")
        estado_cell.border = BORDER

        # Hallazgo y Riesgo columns
        hallazgo_col = None
        riesgo_col = None
        for c in range(1, ws.max_column + 5):
            h = ws.cell(row=header_row, column=c).value
            if h == "Hallazgo":
                hallazgo_col = c
            elif h == "Riesgo":
                riesgo_col = c
        if not hallazgo_col:
            hallazgo_col = ws.max_column + 1
            ws.cell(row=header_row, column=hallazgo_col).value = "Hallazgo"
            ws.cell(row=header_row, column=hallazgo_col).fill = AZUL_HEADER
            ws.cell(row=header_row, column=hallazgo_col).font = BLANCO
            ws.cell(row=header_row, column=hallazgo_col).alignment = Alignment(horizontal="center")
            ws.cell(row=header_row, column=hallazgo_col).border = BORDER
        if not riesgo_col:
            riesgo_col = hallazgo_col + 1
            ws.cell(row=header_row, column=riesgo_col).value = "Riesgo"
            ws.cell(row=header_row, column=riesgo_col).fill = AZUL_HEADER
            ws.cell(row=header_row, column=riesgo_col).font = BLANCO
            ws.cell(row=header_row, column=riesgo_col).alignment = Alignment(horizontal="center")
            ws.cell(row=header_row, column=riesgo_col).border = BORDER

        h_cell = ws.cell(row=r, column=hallazgo_col)
        h_cell.value = hallazgo if hallazgo else "Cumple estándar — sin hallazgos"
        h_cell.fill = AMARILLO if hallazgo else VERDE
        h_cell.alignment = Alignment(wrap_text=True, vertical="top")
        h_cell.border = BORDER

        r_cell = ws.cell(row=r, column=riesgo_col)
        r_cell.value = riesgo if riesgo else "N/A"
        r_cell.fill = AMARILLO if riesgo else VERDE
        r_cell.alignment = Alignment(horizontal="center", vertical="center")
        r_cell.font = NEGRITA
        r_cell.border = BORDER

        ws.column_dimensions[h_cell.column_letter].width = 70
        ws.column_dimensions[r_cell.column_letter].width = 12
        ws.row_dimensions[r].height = 100

ws.column_dimensions["M"].width = 14
wb.save(SRC)
print(f"✅ Excel actualizado: {SRC}")
print(f"   Hoja: {SHEET}")
print(f"   TCs marcados como Aprobados: {len(HALLAZGOS)}")
print(f"   Hallazgos documentados: {sum(1 for h,r in HALLAZGOS.values() if h)}")
