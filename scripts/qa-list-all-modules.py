"""Lista todas las hojas del Excel y los TCs pendientes (estado != Aprobado)."""
from openpyxl import load_workbook
wb = load_workbook("/home/z/my-project/upload/plan-pruebas-qa-jsadr.xlsx", data_only=True)
print(f"Hojas en el Excel: {wb.sheetnames}")
print()

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"=== {sheet_name} (max_row={ws.max_row}) ===")
    pendientes = []
    aprobados = 0
    total_tcs = 0
    for r in range(1, ws.max_row + 1):
        tc_id = ws.cell(row=r, column=2).value
        estado = ws.cell(row=r, column=13).value
        func = ws.cell(row=r, column=4).value
        caso = ws.cell(row=r, column=5).value
        if tc_id and isinstance(tc_id, str) and (tc_id.startswith('TC-') or tc_id.startswith('TC_')):
            total_tcs += 1
            if estado and str(estado).strip().lower() == 'aprobado':
                aprobados += 1
            else:
                pendientes.append((r, tc_id, estado, func, caso))
    print(f"  TCs totales: {total_tcs} | Aprobados: {aprobados} | Pendientes: {len(pendientes)}")
    for r, tc, est, func, caso in pendientes[:30]:
        print(f"    row {r:3d} | {tc:18s} | estado={str(est or '(vacío)'):15s} | {str(func or '')[:25]:25s} | {str(caso or '')[:55]}")
    print()
