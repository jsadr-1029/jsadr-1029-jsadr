"""Actualiza el Excel con los hallazgos y reparos aplicados a M02-Clientes.

Cambios:
- TC-CLI-014: actualizar hallazgo y riesgo con la nueva información del fix v4.5
- TC-CLI-012: ampliar hallazgo (PUT también valida email único)
- Marcar todos como Aprobados (ya lo estaban, pero refrescar fill+font)
"""
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment

WB_PATH = "/home/z/my-project/upload/plan-pruebas-qa-jsadr.xlsx"

# (row, tc_id, nuevo_riesgo, nuevo_hallazgo) — solo actualizar si hay cambio
updates = {
    "4. M02-Clientes": [
        # TC-CLI-012 — Editar email
        (16, "TC-CLI-012",
         "HTTP 200 con email actualizado. Email único validado en PUT (v4.5)",
         "Email único validado en POST y PUT (v4.5). Constraint @unique en BD."),
        # TC-CLI-014 — Email duplicado entre clientes (el que se reparó)
        (18, "TC-CLI-014",
         "HTTP 409 con codigo=EMAIL_DUPLICADO. BD rechaza con Prisma P2002 (v4.5)",
         "REPARO v4.5: Schema email @unique + API POST/PUT validan con 409. "
         "BD desempatada (jsadr23@gmail.com → solo Carolina Alvarez, "
         "Johan Alvarez → email=NULL). Auditoría post-fix: 0 duplicados."),
    ],
}

green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
green_font = Font(color="006100", bold=True)
yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

wb = load_workbook(WB_PATH)
total_updated = 0

for sheet_name, items in updates.items():
    ws = wb[sheet_name]
    for row_num, tc_id, nuevo_riesgo, nuevo_hallazgo in items:
        cell_id = ws.cell(row=row_num, column=2).value
        if cell_id != tc_id:
            print(f"⚠ {sheet_name} fila {row_num}: esperado {tc_id}, encontrado {cell_id}")
            continue

        # Columna Estado = M (13)
        cell_state = ws.cell(row=row_num, column=13)
        cell_state.value = "Aprobado"
        cell_state.fill = green_fill
        cell_state.font = green_font
        cell_state.alignment = Alignment(horizontal="center", vertical="center")

        # Columna Riesgo = K (11)
        cell_riesgo = ws.cell(row=row_num, column=11)
        cell_riesgo.value = nuevo_riesgo
        cell_riesgo.fill = yellow_fill
        cell_riesgo.alignment = Alignment(wrap_text=True, vertical="top")

        # Columna Hallazgo = L (12)
        cell_hallazgo = ws.cell(row=row_num, column=12)
        cell_hallazgo.value = nuevo_hallazgo
        cell_hallazgo.fill = yellow_fill
        cell_hallazgo.alignment = Alignment(wrap_text=True, vertical="top")

        print(f"✓ {sheet_name} fila {row_num} ({tc_id}): Estado=Aprobado, Riesgo+Hallazgo actualizados")
        total_updated += 1

# Adicional: refrescar el resto de TCs ya aprobados con fill verde
for sheet_name in ["4. M02-Clientes"]:
    ws = wb[sheet_name]
    for r in range(5, 20):
        tc_id = ws.cell(row=r, column=2).value
        if tc_id and str(tc_id).startswith('TC-CLI'):
            cell_state = ws.cell(row=r, column=13)
            if cell_state.value == "Aprobado":
                cell_state.fill = green_fill
                cell_state.font = green_font
                cell_state.alignment = Alignment(horizontal="center", vertical="center")

wb.save(WB_PATH)
print(f"\n=== Total actualizado: {total_updated} items con hallazgo ampliado ===")
print(f"Archivo guardado: {WB_PATH}")

# Copiar a download para entrega
import shutil
dest = "/home/z/my-project/download/plan-pruebas-qa-jsadr-actualizado.xlsx"
shutil.copy(WB_PATH, dest)
print(f"Copia para entrega: {dest}")
