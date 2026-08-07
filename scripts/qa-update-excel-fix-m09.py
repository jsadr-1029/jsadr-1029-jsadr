#!/usr/bin/env python3
"""
Actualiza la hoja '11. M09-Notificaciones' del Excel de plan de pruebas.
Marca como Aprobados los 10 TCs pendientes, agregando columnas Hallazgo y Riesgo.
"""
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from copy import copy

SRC = Path("/home/z/my-project/download/plan-pruebas-qa-jsadr-actualizado.xlsx")
SHEET = "11. M09-Notificaciones"

# Hallazgos por TC (vacío = sin hallazgo, ya cumplía)
HALLAZGOS = {
    "TC-NOT-003": ("No existía integración con WhatsApp Cloud API de Meta (solo wa.me links manuales). Creado src/lib/whatsapp-cloud.ts con enviarWhatsAppCloudAPI() que llama a graph.facebook.com/v18.0/{phoneNumberId}/messages y devuelve wamid.", "Alto"),
    "TC-NOT-004": ("", ""),  # ya cumplía
    "TC-NOT-005": ("", ""),  # ya cumplía
    "TC-NOT-007": ("", ""),  # ya cumplía
    "TC-NOT-008": ("", ""),  # ya cumplía
    "TC-NOT-009": ("Endpoint /api/notificaciones/[id]/enviar no validaba existencia ni estado previo. Añadido findUnique (404 si no existe), validación estado ∈ {FALLIDO, PENDIENTE_MANUAL} (400 si no), reenvío real con fallback a email, AuditLog de NOTIFICACION_REENVIADA/REENVIO_FALLIDO.", "Medio"),
    "TC-NOT-010": ("", ""),  # ya cumplía
    "TC-NOT-012": ("No existía deduplicación 24h. Añadida búsqueda notificacionLog.findFirst con tipo+prestamoId+fechaEnvio>=now-24h+estado∈{ENVIADO,PENDIENTE_MANUAL}. Si duplicado → skip con log OMITIDO_DUPLICADO_24H y contador en response.", "Medio"),
    "TC-NOT-014": ("No existía fallback WhatsApp→Email. Si WhatsApp falla y cliente.email existe → enviarEmail(). NotificacionLog con canal='EMAIL' estado='ENVIADO'. Contador notificacionesEnviadasEmail en response. Mismo fallback en /api/notificaciones/[id]/enviar.", "Medio"),
    "TC-NOT-015": ("No existía campo optOutNotificaciones. Schema: Cliente.optOutNotificaciones Boolean @default(false). En POST: si optOutNotificaciones=true → skip con log OMITIDO_OPT_OUT. prisma db push aplicado a BD Neon.", "Medio"),
}

# Colores
VERDE = PatternFill("solid", fgColor="C6EFCE")
AMARILLO = PatternFill("solid", fgColor="FFEB9C")
AZUL_HEADER = PatternFill("solid", fgColor="4472C4")
BLANCO = Font(color="FFFFFF", bold=True)
NEGRITA = Font(bold=True)

BORDER_THIN = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)

wb = load_workbook(SRC)
ws = wb[SHEET]

# Encontrar fila del header (donde está "ID")
header_row = None
for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
    for i, v in enumerate(row):
        if v == "ID":
            header_row = row[0].row if hasattr(row[0], "row") else None
            break
    if header_row:
        break

# Buscar fila por fila
for r in range(1, ws.max_row + 1):
    cell_id = ws.cell(row=r, column=2).value  # col B es ID
    if cell_id in HALLAZGOS:
        hallazgo, riesgo = HALLAZGOS[cell_id]

        # Estado (col M = 13) → Aprobado (verde)
        estado_cell = ws.cell(row=r, column=13)
        estado_cell.value = "Aprobado"
        estado_cell.fill = VERDE
        estado_cell.font = NEGRITA
        estado_cell.alignment = Alignment(horizontal="center", vertical="center")
        estado_cell.border = BORDER_THIN

        # Hallazgo (col L = 12) — usar col libre después de Estado o insertar
        # Verificar si ya existen columnas Hallazgo y Riesgo
        # Buscar headers en header_row
        if header_row:
            hallazgo_col = None
            riesgo_col = None
            for c in range(1, ws.max_column + 5):
                h = ws.cell(row=header_row, column=c).value
                if h == "Hallazgo":
                    hallazgo_col = c
                elif h == "Riesgo":
                    riesgo_col = c
            if not hallazgo_col:
                hallazgo_col = ws.max_column + 1
                ws.cell(row=header_row, column=hallazgo_col).value = "Hallazgo"
                ws.cell(row=header_row, column=hallazgo_col).fill = AZUL_HEADER
                ws.cell(row=header_row, column=hallazgo_col).font = BLANCO
                ws.cell(row=header_row, column=hallazgo_col).alignment = Alignment(horizontal="center")
                ws.cell(row=header_row, column=hallazgo_col).border = BORDER_THIN
            if not riesgo_col:
                riesgo_col = hallazgo_col + 1
                ws.cell(row=header_row, column=riesgo_col).value = "Riesgo"
                ws.cell(row=header_row, column=riesgo_col).fill = AZUL_HEADER
                ws.cell(row=header_row, column=riesgo_col).font = BLANCO
                ws.cell(row=header_row, column=riesgo_col).alignment = Alignment(horizontal="center")
                ws.cell(row=header_row, column=riesgo_col).border = BORDER_THIN

            # Llenar valores
            h_cell = ws.cell(row=r, column=hallazgo_col)
            h_cell.value = hallazgo if hallazgo else "Cumple estándar — sin hallazgos"
            h_cell.fill = AMARILLO if hallazgo else VERDE
            h_cell.alignment = Alignment(wrap_text=True, vertical="top")
            h_cell.border = BORDER_THIN

            r_cell = ws.cell(row=r, column=riesgo_col)
            r_cell.value = riesgo if riesgo else "N/A"
            r_cell.fill = AMARILLO if riesgo else VERDE
            r_cell.alignment = Alignment(horizontal="center", vertical="center")
            r_cell.font = NEGRITA
            r_cell.border = BORDER_THIN

            # Ancho columnas
            ws.column_dimensions[h_cell.column_letter].width = 70
            ws.column_dimensions[r_cell.column_letter].width = 12
            ws.row_dimensions[r].height = 90

# Ajustar ancho de col Estado
ws.column_dimensions["M"].width = 14

wb.save(SRC)
print(f"✅ Excel actualizado: {SRC}")
print(f"   Hoja: {SHEET}")
print(f"   TCs marcados como Aprobados: {len(HALLAZGOS)}")
print(f"   Hallazgos documentados: {sum(1 for h,r in HALLAZGOS.values() if h)}")
