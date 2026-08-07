#!/usr/bin/env python3
"""Lee la hoja M12-UI-UX del Excel de plan de pruebas."""
from openpyxl import load_workbook

wb = load_workbook("/home/z/my-project/download/plan-pruebas-qa-jsadr-actualizado.xlsx", data_only=False)
print("Hojas disponibles:")
for s in wb.sheetnames:
    print(f"  - {s}")

target = None
for s in wb.sheetnames:
    if "M12" in s or "UI-UX" in s or "Mobile" in s:
        target = s
        break
print(f"\nHoja objetivo: {target}")
if not target:
    raise SystemExit("No se encontro hoja M12-UI-UX")

ws = wb[target]
print(f"Dimensiones: {ws.max_row} filas x {ws.max_column} columnas\n")

print("=== Encabezados (fila 1) ===")
for col in range(1, ws.max_column + 1):
    val = ws.cell(row=1, column=col).value
    if val:
        print(f"  Col {col}: {val}")

print("\n=== Filas de datos ===")
for row in range(1, ws.max_row + 1):
    fila = []
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=row, column=col).value
        fila.append(str(v) if v is not None else "")
    if any(fila):
        print(f"Fila {row}: {' | '.join(fila)}")
