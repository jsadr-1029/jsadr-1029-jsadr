"""Lee el detalle completo (todas las columnas) de los items pendientes de un módulo."""
import sys
import openpyxl
from openpyxl import load_workbook

WB_PATH = "/home/z/my-project/upload/plan-pruebas-qa-jsadr.xlsx"
sheet_name = sys.argv[1] if len(sys.argv) > 1 else "3. M01-Autenticación"

wb = load_workbook(WB_PATH, data_only=False)
ws = wb[sheet_name]

rows = list(ws.iter_rows(values_only=True))
# Header en fila 4 (índice 3)
header = [str(c).strip() if c else '' for c in rows[3]]

print(f"=== HOJA: {sheet_name} ===")
print(f"Header: {header}\n")

# Columnas: 0 vacío, 1 ID, 2 Módulo, 3 Función, 4 Caso de Prueba, 5 Tipo, 6 Prioridad,
# 7 Precondiciones, 8 Pasos, 9 Datos de Entrada, 10 Resultado Esperado, 11 Criterios de Aceptación, 12 Estado

for i, r in enumerate(rows[4:], start=5):
    if not any(r):
        continue
    estado = r[12] if len(r) > 12 else None
    estado_str = str(estado).strip().lower() if estado else ''
    if estado_str != 'pendiente':
        continue
    print(f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━")
    print(f"FILA {i}")
    for ci, h in enumerate(header):
        if not h:
            continue
        v = r[ci] if ci < len(r) else None
        if v is None or v == '':
            continue
        print(f"  [{h}]:")
        # multi-line values
        s = str(v)
        for line in s.split('\n'):
            print(f"      {line}")
    print()
