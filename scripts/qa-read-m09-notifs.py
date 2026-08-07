#!/usr/bin/env python3
"""Lee la hoja M09-Notificaciones del Excel de plan de pruebas."""
import openpyxl
from pathlib import Path

EXCEL = Path("/home/z/my-project/upload/plan-pruebas-qa-jsadr.xlsx")
wb = openpyxl.load_workbook(EXCEL, data_only=False)

# Buscar hoja M09
sheet_name = None
for name in wb.sheetnames:
    print(f"  Hoja: {name!r}")
    if "M09" in name or "Notif" in name.lower():
        sheet_name = name

if not sheet_name:
    print("\nNo se encontró hoja M09-Notificaciones")
    raise SystemExit(1)

print(f"\n>>> Hoja seleccionada: {sheet_name!r}")
ws = wb[sheet_name]
print(f"Dimensiones: {ws.dimensions} | max_row={ws.max_row} max_col={ws.max_column}")

# Imprimir todas las filas con contenido
print("\n=== Contenido completo ===")
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
    cells = []
    for c in row:
        v = c.value
        if v is None:
            v = ""
        cells.append(str(v).replace("\n", " | ")[:80])
    # Imprimir fila completa aunque tenga celdas vacías
    print(f"R{row[0].row:>3}: {cells}")
