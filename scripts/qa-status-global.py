"""Lista el estado de todos los módulos del Excel para identificar el siguiente pendiente."""
from openpyxl import load_workbook

WB = "/home/z/my-project/download/plan-pruebas-qa-jsadr-actualizado.xlsx"
wb = load_workbook(WB, data_only=True)

print("=" * 80)
print("ESTADO GLOBAL DEL PLAN DE PRUEBAS QA")
print("=" * 80)

resumen = []
for sn in wb.sheetnames:
    if not sn.startswith(tuple(f"{i}. " for i in range(3, 16))):
        continue
    if "M0" not in sn and "M1" not in sn:
        continue
    ws = wb[sn]
    # Buscar columna Estado (13) y columna ID (2)
    total = 0
    aprobados = 0
    pendientes = 0
    otros = 0
    for r in range(4, ws.max_row + 1):
        tc_id = ws.cell(row=r, column=2).value
        if not tc_id or 'TC' not in str(tc_id).upper():
            continue
        total += 1
        estado = ws.cell(row=r, column=13).value
        estado_str = str(estado or '').strip().lower()
        if estado_str == 'aprobado':
            aprobados += 1
        elif estado_str in ('pendiente', '', 'no ejecutado', 'no ejecutada'):
            pendientes += 1
        else:
            otros += 1
    if total > 0:
        resumen.append((sn, total, aprobados, pendientes, otros))
        pct = (aprobados / total) * 100 if total > 0 else 0
        print(f"\n{sn}")
        print(f"  Total TCs: {total} | Aprobados: {aprobados} | Pendientes: {pendientes} | Otros: {otros}")
        bar_len = 30
        filled = int(bar_len * aprobados / total)
        bar = '█' * filled + '░' * (bar_len - filled)
        print(f"  [{bar}] {pct:.0f}%")

print("\n" + "=" * 80)
print("MÓDULOS CON TCs PENDIENTES (ordenados por # pendientes):")
print("=" * 80)
pendientes_list = [(sn, pend, total, aprob) for sn, total, aprob, pend, otr in resumen if pend > 0]
pendientes_list.sort(key=lambda x: -x[1])
for sn, pend, total, aprob in pendientes_list:
    print(f"  • {sn}: {pend} pendientes de {total} ({aprob} aprobados)")

print("\n" + "=" * 80)
print("SIGUIENTE MÓDULO A EJECUTAR:")
print("=" * 80)
if pendientes_list:
    next_mod = pendientes_list[0]
    print(f"  → {next_mod[0]} ({next_mod[1]} TCs pendientes)")
else:
    print("  ✅ TODOS LOS MÓDULOS COMPLETOS")
