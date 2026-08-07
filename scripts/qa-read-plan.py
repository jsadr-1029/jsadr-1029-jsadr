"""Lee el plan de pruebas QA y muestra los items pendientes por módulo."""
import openpyxl
from openpyxl import load_workbook
from collections import defaultdict

WB_PATH = "/home/z/my-project/upload/plan-pruebas-qa-jsadr.xlsx"

wb = load_workbook(WB_PATH, data_only=False)

# Mostrar hojas disponibles
print("=== HOJAS ===")
for s in wb.sheetnames:
    print(f"  - {s}")

print("\n=== ÍNDICE DE MÓDULOS ===")
ws = wb["2. Índice de Módulos"]
for row in ws.iter_rows(values_only=False):
    vals = [(c.coordinate, c.value) for c in row if c.value is not None]
    if vals:
        print(vals)

# Para cada hoja de módulo, detectar header y filas con estado "pendiente"
module_sheets = [s for s in wb.sheetnames if s.startswith(("3.", "4.", "5.", "6.", "7.", "8.", "9.", "10.", "11.", "12.", "13.", "14.", "15."))]

print(f"\n=== MÓDULOS: {len(module_sheets)} hojas ===")

summary = []
for sname in module_sheets:
    ws = wb[sname]
    # Leer todo en una lista de listas
    rows = []
    for r in ws.iter_rows(values_only=True):
        rows.append(list(r))
    
    print(f"\n\n>>> HOJA: {sname}")
    print(f"   Dimensiones: {ws.max_row} filas × {ws.max_column} columnas")
    
    # Imprimir primeras 4 filas para detectar header
    for i, r in enumerate(rows[:4]):
        print(f"   ROW {i+1}: {r}")
    
    # Detectar header (fila con "ID" o "Módulo" o "Estado")
    header_idx = None
    for i, r in enumerate(rows[:6]):
        if r and any(c and isinstance(c, str) and c.strip().lower() in ('estado', 'id', 'módulo', 'modulo', 'caso', 'descripción', 'descripcion', 'item') for c in r):
            header_idx = i
            break
    print(f"   HEADER en fila: {header_idx+1 if header_idx is not None else 'NO DETECTADO'}")
    
    if header_idx is not None:
        header = [str(c).strip().lower() if c else '' for c in rows[header_idx]]
        # Buscar columna de estado
        estado_col = None
        for ci, h in enumerate(header):
            if 'estado' in h:
                estado_col = ci
                break
        # Buscar columna ID o Item
        id_col = None
        for ci, h in enumerate(header):
            if h in ('id', 'item', 'caso', 'caso id', 'caso_id', 'id caso', '#'):
                id_col = ci
                break
        desc_col = None
        for ci, h in enumerate(header):
            if 'desc' in h or 'caso' in h or 'prueba' in h:
                desc_col = ci
                break
        print(f"   Col Estado={estado_col}  Col ID={id_col}  Col Desc={desc_col}")
        print(f"   Header completo: {header}")
        
        # Contar items por estado
        counts = defaultdict(int)
        pending_items = []
        for i, r in enumerate(rows[header_idx+1:], start=header_idx+2):
            if not any(r):
                continue
            estado = r[estado_col] if estado_col is not None and estado_col < len(r) else None
            estado_str = str(estado).strip().lower() if estado else '(vacío)'
            counts[estado_str] += 1
            if estado_str == 'pendiente':
                item_id = r[id_col] if id_col is not None and id_col < len(r) else f"fila{i}"
                desc = r[desc_col] if desc_col is not None and desc_col < len(r) else ''
                pending_items.append((item_id, desc, i))
        
        total = sum(counts.values())
        print(f"   Total items: {total}")
        for est, n in sorted(counts.items()):
            print(f"     {est}: {n} ({100*n/total:.1f}%)" if total else f"     {est}: 0")
        
        if pending_items:
            print(f"   PENDIENTES ({len(pending_items)}):")
            for it_id, desc, rownum in pending_items[:5]:
                print(f"     - fila {rownum} | id={it_id} | {str(desc)[:80] if desc else ''}")
            if len(pending_items) > 5:
                print(f"     ... y {len(pending_items)-5} más")
        
        summary.append({
            'sheet': sname,
            'total': total,
            'pending': len(pending_items),
            'counts': dict(counts),
            'pending_items': pending_items,
        })

# Guardar resumen JSON
import json
with open("/home/z/my-project/tool-results/qa-summary.json", "w") as f:
    json.dump(summary, f, indent=2, default=str, ensure_ascii=False)
print(f"\n\n=== Resumen guardado en tool-results/qa-summary.json ===")
print(f"\n=== TOTALES GLOBALES ===")
total_global = sum(m['total'] for m in summary)
pending_global = sum(m['pending'] for m in summary)
print(f"Total items: {total_global}")
print(f"Pendientes:  {pending_global}")
if total_global:
    print(f"Avance:      {100*(total_global-pending_global)/total_global:.1f}%")
