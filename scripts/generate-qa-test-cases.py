#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Genera el Excel de Casos de Prueba QA — Sistema Jsadr v5.0
===========================================================
Genera un workbook con:
  - Sheet 1: Índice (resumen por módulo)
  - Sheet 2: Estándares QA (definiciones de tipos, prioridades, estados)
  - Sheets 3-20: Un sheet por cada módulo del sistema

Cada caso de prueba tiene:
  ID | Función | Caso de Prueba | Tipo | Prioridad | Pre-condiciones |
  Pasos | Datos de Entrada | Resultado Esperado | Criterios de Aceptación |
  Estado | Notas

Estándares QA Tester aplicados:
  - Tipos: Positivo / Negativo / Borde / Seguridad / Performance / Regresión
  - Prioridades: Crítica / Alta / Media / Baja
  - Estados: Pendiente / En Progreso / Aprobado / Fallido / Bloqueado
"""

import sys
import os

# Cargar el design system del skill xlsx
XLSX_SKILL_DIR = "/home/z/my-project/skills/xlsx"
for sub in [XLSX_SKILL_DIR, os.path.join(XLSX_SKILL_DIR, "templates")]:
    if sub not in sys.path:
        sys.path.insert(0, sub)

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from base import (
    FONT_NAME, HEADER_BOLD,
    PRIMARY, PRIMARY_LIGHT, SECONDARY,
    ACCENT_POSITIVE, ACCENT_NEGATIVE, ACCENT_WARNING,
    NEUTRAL_900, NEUTRAL_600, NEUTRAL_200, NEUTRAL_100, NEUTRAL_50, NEUTRAL_0,
    HEADER_TEXT,
    font_title, font_header, font_subheader, font_body, font_caption,
    fill_header, fill_total, fill_data_row,
    border_header, border_total,
    align_title, align_header, align_number, align_text,
    setup_sheet, style_header_row, style_data_row, style_total_row,
    auto_fit_columns,
    use_palette_explicit,
)

# Usar paleta profesional (azul profundo) — apropiado para documentación QA
use_palette_explicit("professional")

# === Headers comunes para todos los sheets de módulos ===
TEST_CASE_HEADERS = [
    "ID",
    "Función",
    "Caso de Prueba",
    "Tipo",
    "Prioridad",
    "Pre-condiciones",
    "Pasos",
    "Datos de Entrada",
    "Resultado Esperado",
    "Criterios de Aceptación",
    "Estado",
    "Notas",
]

# === Anchos de columna (en unidades Excel) ===
COLUMN_WIDTHS_TEST = {
    "B": 8,    # ID
    "C": 24,   # Función
    "D": 36,   # Caso de Prueba
    "E": 12,   # Tipo
    "F": 11,   # Prioridad
    "G": 32,   # Pre-condiciones
    "H": 38,   # Pasos
    "I": 28,   # Datos de Entrada
    "J": 38,   # Resultado Esperado
    "K": 32,   # Criterios de Aceptación
    "L": 13,   # Estado
    "M": 26,   # Notas
}

# ============================================================
# DEFINICIÓN DE CASOS DE PRUEBA POR MÓDULO
# ============================================================
# Estructura: { "Módulo": [ (función, caso, tipo, prioridad, pre, pasos, datos, esperado, criterios, notas), ... ] }
# Estado se inicializa como "Pendiente" para todos.

MODULES = {}

# ============================================================
# 1. AUTENTICACIÓN Y SEGURIDAD
# ============================================================
MODULES["1. Autenticación y Seguridad"] = [
    # Login admin
    ("Login admin válido", "POST /api/auth/login con credenciales válidas de ADMIN",
     "Positivo", "Crítica",
     "Usuario ADMIN activo en BD con passwordHash válido",
     "1. POST /api/auth/login\n2. Body: { username, password }\n3. Verificar respuesta",
     '{"username":"Adm-Jsadr","password":"<pass-real>"}',
     "200 OK con { success:true, accessToken, refreshToken, user:{rol:'ADMIN'} }",
     "JWT access 15min + refresh 7d generados. AuditLog registra LOGIN_EXITOSO. mustChangePassword=false.",
     "Revisar que cookie HttpOnly se establezca"),
    # Login admin pass incorrecta
    ("Login admin password incorrecta", "POST /api/auth/login con password inválida",
     "Negativo", "Alta",
     "Usuario ADMIN existe y está activo",
     "1. POST /api/auth/login con password errada\n2. Reintentar 5 veces",
     '{"username":"Adm-Jsadr","password":"wrong-pass"}',
     "401 Unauthorized. Tras 5 intentos: bloqueo 30min (bloqueadoHasta).",
     "intentosFallidos se incrementa. AuditLog registra INTENTO_FALLIDO. Rate limit activo."),
    # Login usuario inactivo
    ("Login usuario inactivo bloqueado", "POST /api/auth/login con usuario inactivo",
     "Negativo", "Alta",
     "Usuario con activo=false en BD",
     "1. POST /api/auth/login",
     '{"username":"inactive-user","password":"any"}',
     "403 Forbidden con code:'USER_INACTIVE'",
     "No se genera JWT. AuditLog registra login fallido."),
    # Recuperar clave
    ("Recuperar clave — flujo completo", "POST /api/auth/recuperar-clave con identificador válido",
     "Positivo", "Crítica",
     "Usuario o Cliente con email registrado y SMTP configurado",
     "1. POST /api/auth/recuperar-clave\n2. Verificar correo recibido\n3. Login con contraseña temporal\n4. Sistema pide cambiar clave",
     '{"identificador":"admin@jsadr.co"}',
     "200 OK con mensaje genérico. Email con contraseña temporal enviado. mustChangePassword=true al login.",
     "Respuesta idéntica exista o no el usuario. Contraseña temporal 12 chars alfanuméricos."),
    # Rate limit recuperar clave
    ("Rate limit recuperar clave", "Solicitar recuperación 2 veces seguidas desde misma IP",
     "Seguridad", "Alta",
     "Rate limit configurado a 5 min",
     "1. POST /api/auth/recuperar-clave\n2. Repetir inmediatamente",
     '{"identificador":"admin@jsadr.co"}',
     "Segunda solicitud: 429 con code:'RATE_LIMIT' y minutosRestantes",
     "Map en memoria controla rate limit por IP."),
    # MFA
    ("MFA setup y verificación", "POST /api/auth/mfa setup + verify con código TOTP válido",
     "Positivo", "Alta",
     "Usuario autenticado, sin MFA previo",
     "1. POST /api/auth/mfa action=setup\n2. Escanear QR con app autenticadora\n3. POST /api/auth/mfa action=verify con código TOTP",
     '{"action":"setup"} luego {"action":"verify","codigo":"123456"}',
     "200 OK con totpSecret y QR. verify:true tras código válido.",
     "totpEnabled=true en BD. totpSecret cifrado."),
    # Refresh token
    ("Refresh token válido", "POST /api/auth/refresh con refreshToken activo",
     "Positivo", "Alta",
     "refreshToken emitido hace <7d, no revocado",
     "1. POST /api/auth/refresh\n2. Body: { refreshToken }",
     '{"refreshToken":"<jwt-refresh>"}',
     "200 OK con nuevo accessToken (15min).",
     "refreshToken NO se rota (sigue válido hasta expirar)."),
    # Refresh expirado
    ("Refresh token expirado", "POST /api/auth/refresh con refreshToken >7d",
     "Negativo", "Media",
     "refreshToken emitido hace >7d",
     "1. POST /api/auth/refresh con token expirado",
     '{"refreshToken":"<jwt-expired>"}',
     "401 Unauthorized con code:'TOKEN_EXPIRED'",
     "Cliente debe hacer login de nuevo."),
    # Switch user (impersonate)
    ("Switch user (impersonate)", "POST /api/auth/switch-user como ADMIN a GESTOR",
     "Seguridad", "Alta",
     "ADMIN autenticado, usuario destino existe",
     "1. POST /api/auth/switch-user\n2. Realizar acciones como GESTOR\n3. POST /api/auth/switch-back",
     '{"targetUserId":"<gestor-id>"}',
     "200 OK con token del GESTOR. AuditLog registra IMPERSONATE_START. switch-back restaura token ADMIN.",
     "Solo ADMIN puede impersonate. AuditLog auditable."),
    # RBAC
    ("RBAC — GESTOR intenta acceso admin", "POST /api/email con token GESTOR",
     "Seguridad", "Crítica",
     "Usuario GESTOR autenticado",
     "1. POST /api/email con token GESTOR",
     '{"accion":"probar"}',
     "403 Forbidden con code:'FORBIDDEN'",
     "requireRole(['ADMIN']) bloquea. AuditLog registra acceso denegado."),
]

# ============================================================
# 2. GESTIÓN DE CLIENTES
# ============================================================
MODULES["2. Gestión de Clientes"] = [
    # Crear cliente
    ("Crear cliente válido", "POST /api/clientes con datos completos",
     "Positivo", "Crítica",
     "ADMIN o GESTOR autenticado. Cédula no existe.",
     "1. POST /api/clientes\n2. Verificar en BD y dashboard",
     '{"nombre":"Juan Pérez","cedula":"12345678","email":"juan@x.com","telefono":"3001234567",...}',
     "201 Created con cliente.id. cliente.activo=true.",
     "pinHash inicial null. cliente.createdAt=now(). AuditLog creado."),
    # Cédula duplicada
    ("Crear cliente cédula duplicada", "POST /api/clientes con cédula ya existente",
     "Negativo", "Alta",
     "Cliente con cédula '12345678' ya existe",
     "1. POST /api/clientes con misma cédula",
     '{"nombre":"Otro","cedula":"12345678"}',
     "409 Conflict con code:'CEDULA_DUPLICADA'",
     "No se crea registro. Mensaje claro al usuario."),
    # Validación email
    ("Crear cliente email inválido", "POST /api/clientes con email mal formado",
     "Negativo", "Media",
     "ADMIN/GESTOR autenticado",
     "1. POST /api/clientes con email='no-es-email'",
     '{"nombre":"X","cedula":"99","email":"no-es-email"}',
     "400 Bad Request con code:'INVALID_EMAIL'",
     "validateEmail() retorna false."),
    # Listar con paginación
    ("Listar clientes paginado", "GET /api/clientes?page=1&pageSize=20",
     "Positivo", "Media",
     "Al menos 25 clientes en BD",
     "1. GET /api/clientes?page=1&pageSize=20\n2. GET /api/clientes?page=2&pageSize=20",
     '?page=1&pageSize=20',
     "200 OK con { data:[...], total, page, pageSize }",
     "total coincide con count en BD. data.length <= pageSize."),
    # Buscar cliente
    ("Buscar cliente por nombre", "GET /api/clientes?q=juan",
     "Positivo", "Media",
     "Cliente 'Juan Pérez' existe",
     "1. GET /api/clientes?q=juan",
     '?q=juan',
     "200 OK con array filtrado (case-insensitive)",
     "Contiene clientes cuyo nombre/cédula/email coinciden."),
    # Editar cliente
    ("Editar cliente existente", "PATCH /api/clientes/:id con nuevos datos",
     "Positivo", "Alta",
     "Cliente existe y activo",
     "1. PATCH /api/clientes/:id\n2. Verificar updatedAt",
     '{"telefono":"3009876543"}',
     "200 OK con cliente actualizado. updatedAt > createdAt.",
     "AuditLog registra CLIENTE_UPDATE."),
    # Eliminar (soft delete)
    ("Eliminar cliente (soft delete)", "DELETE /api/clientes/:id",
     "Positivo", "Alta",
     "Cliente sin préstamos ACTIVOS",
     "1. DELETE /api/clientes/:id\n2. Verificar activo=false",
     'DELETE /api/clientes/<id>',
     "200 OK con { success:true }. cliente.activo=false (soft delete).",
     "No se borra físicamente. Préstamos históricos preservados."),
    # Intentar eliminar con préstamos activos
    ("Eliminar cliente con préstamo activo", "DELETE /api/clientes/:id con préstamo ACTIVO",
     "Negativo", "Alta",
     "Cliente tiene préstamo en estado ACTIVO",
     "1. DELETE /api/clientes/:id",
     'DELETE /api/clientes/<id-con-prestamo>',
     "409 Conflict con code:'CLIENTE_TIENE_PRESTAMOS_ACTIVOS'",
     "Operación bloqueada. No se modifica cliente."),
    # Asignar categoría
    ("Asignar categoría a cliente", "PATCH /api/clientes/:id categoriaId",
     "Positivo", "Media",
     "Categoría existe y está activa",
     "1. PATCH /api/clientes/:id\n2. Verificar categoría aplicada",
     '{"categoriaId":"<cat-id>"}',
     "200 OK con cliente.categoriaId actualizado.",
     "Si categoría tiene tasa personalizada, se aplica a futuros préstamos."),
    # Referido
    ("Crear cliente con referido", "POST /api/clientes con referidoPorId",
     "Positivo", "Media",
     "Cliente referidor existe y activo",
     "1. POST /api/clientes con referidoPorId\n2. Verificar relación",
     '{"nombre":"Nuevo","cedula":"99","referidoPorId":"<ref-id>"}',
     "201 Created. cliente.referidoPorId = ref-id.",
     "Listado de referidos del referidor incluye al nuevo cliente."),
]

# ============================================================
# 3. PRÉSTAMOS
# ============================================================
MODULES["3. Préstamos"] = [
    # Crear préstamo
    ("Crear préstamo válido", "POST /api/prestamos con datos completos",
     "Positivo", "Crítica",
     "Cliente existe y activo. Categoría configurada.",
     "1. POST /api/prestamos\n2. Verificar cálculo automático\n3. Verificar estado=PENDIENTE_ACEPTACION",
     '{"clienteId":"...","montoPrincipal":1000000,"tasaInteresAnual":24,"plazoMeses":12,"frecuencia":"MENSUAL"}',
     "201 Created con código único (PR-XXXX). montoCuota y totalPagar calculados.",
     "Estado inicial PENDIENTE_ACEPTACION. fechaSolicitud=now(). Bitácora creada."),
    # Cálculo incorrecto
    ("Crear préstamo monto negativo", "POST /api/prestamos con monto=-1000",
     "Negativo", "Alta",
     "ADMIN/GESTOR autenticado",
     "1. POST /api/prestamos con montoPrincipal=-1000",
     '{"montoPrincipal":-1000,...}',
     "400 Bad Request con code:'INVALID_AMOUNT'",
     "Validación de monto > 0."),
    # Tasa excesiva
    ("Crear préstamo tasa > 100%", "POST /api/prestamos con tasaAnual=150",
     "Borde", "Media",
     "Sistema permite tasas hasta tope legal",
     "1. POST /api/prestamos con tasaInteresAnual=150",
     '{"tasaInteresAnual":150,...}',
     "400 Bad Request o warning. Tasa máxima permitida configurable.",
     "Si se permite, marcar como revisión legal pendiente."),
    # Aceptar TyC con OTP
    ("Aceptar TyC con OTP correo", "POST /api/prestamos/:id/aceptar-tyc-otp action=enviar_otp canal=EMAIL",
     "Positivo", "Crítica",
     "Préstamo en PENDIENTE_ACEPTACION. Cliente con email.",
     "1. POST enviar_otp canal=EMAIL\n2. Verificar correo recibido\n3. POST validar_otp\n4. POST confirmar_con_foto con cédula+selfie",
     '{"accion":"enviar_otp","canal":"EMAIL"}',
     "200 OK. OTP hasheado SHA-256 en firma.otpCodigo. Estado firma OTP_ENVIADO.",
     "OTP 6 dígitos. Expira en 5min. NotificacionLog creado. AccesoPortal creado."),
    # Reutilizar OTP vigente
    ("Solicitar OTP cuando hay vigente", "POST aceptar-tyc-otp con OTP aún activo",
     "Positivo", "Alta",
     "Firma con OTP_ENVIADO hace <5min",
     "1. POST enviar_otp nuevamente",
     '{"accion":"enviar_otp"}',
     "200 OK con reutilizado:true y segundosRestantes. NO se genera nuevo OTP.",
     "Sistema reutiliza OTP vigente para evitar spam de correos."),
    # OTP expirado
    ("Validar OTP expirado", "POST validar_otp con OTP >5min",
     "Negativo", "Alta",
     "OTP enviado hace >5min",
     "1. POST validar_otp",
     '{"accion":"validar_otp","otpIngresado":"123456"}',
     "400 Bad Request: 'El código ha expirado'",
     "firma.usado=true. Cliente debe solicitar nuevo OTP."),
    # 3 intentos fallidos bloqueo
    ("Bloqueo tras 3 intentos OTP fallidos", "POST validar_otp 3 veces con código errado",
     "Negativo", "Alta",
     "OTP activo. intentosOTP=0.",
     "1. POST validar_otp código errado (1)\n2. Repetir (2)\n3. Repetir (3)",
     '{"accion":"validar_otp","otpIngresado":"000000"} (3 veces)',
     "Tercer intento: 400 con estadoFirma=RECHAZADA. Préstamo no se puede activar por OTP.",
     "intentosOTP=maxIntentos=3. firma.bloqueado=true."),
    # Confirmar con foto
    ("Confirmar TyC con fotos válidas", "POST confirmar_con_foto con cédula+selfie JPEG",
     "Positivo", "Crítica",
     "OTP validado. EstadoFirma OTP_ENVIADO. otpValidado=true.",
     "1. POST confirmar_con_foto con fotoDocumentoBase64 y fotoSelfieBase64",
     '{"accion":"confirmar_con_foto","fotoDocumentoBase64":"data:image/jpeg...","fotoSelfieBase64":"data:image/jpeg..."}',
     "200 OK. Préstamo estado=ACTIVO. fechaDesembolso=now(). firma COMPLETADA.",
     "DocumentoGestor crea 2 entradas (FOTO_CEDULA, FOTO_SELFI). Hashes SHA-256 guardados. Pagaré PDF generado."),
    # Foto SVG rechazada
    ("Confirmar con foto SVG (XSS)", "POST confirmar_con_foto con SVG",
     "Seguridad", "Alta",
     "OTP validado",
     "1. POST confirmar_con_foto con data:image/svg+xml",
     '{"fotoDocumentoBase64":"data:image/svg+xml,..."}',
     "400 Bad Request: 'La foto del documento debe ser JPEG, PNG o WebP.'",
     "SVG bloqueado por riesgo XSS."),
    # Foto >5MB
    ("Confirmar con foto >5MB", "POST confirmar_con_foto con imagen >7MB base64",
     "Borde", "Media",
     "OTP validado",
     "1. POST con fotoDocumentoBase64 >7MB",
     '{"fotoDocumentoBase64":"<7MB>"}',
     "400 Bad Request: 'Las fotos exceden el tamaño máximo (5MB cada una).'",
     "MAX_SIZE = 7*1024*1024 chars (≈5MB binario)."),
    # Codeudor
    ("Crear préstamo con codeudor", "POST /api/prestamos con tieneCodeudor=true",
     "Positivo", "Alta",
     "Cliente principal y codeudor existen",
     "1. POST /api/prestamos con codeudorId",
     '{"tieneCodeudor":true,"codeudorId":"<cod-id>"}',
     "201 Created. prestamo.codeudorId poblado.",
     "Codeudor requiere firma separada. Notificaciones a ambos."),
    # Flexibilidad financiera
    ("Activar flexibilidad financiera", "PATCH /api/prestamos/:id flexibilidadActivada=true",
     "Positivo", "Media",
     "Préstamo ACTIVO con cuotas pendientes",
     "1. PATCH /api/prestamos/:id\n2. Verificar recalculo",
     '{"flexibilidadActivada":true,"flexibilidadCosto":50000}',
     "200 OK. flexibilidadActivada=true. fechaActivacion=now().",
     "Recálculo de cuotas pendientes aplica flexibilidad."),
    # Mora
    ("Cálculo mora compuesta diaria", "GET /api/prestamos/:id tras vencimiento",
     "Positivo", "Alta",
     "Préstamo con cuota vencida hace 10 días",
     "1. GET /api/prestamos/:id\n2. Verificar montoMora y diasMora",
     'GET /api/prestamos/<id-vencido>',
     "200 OK. diasMora=10. montoMora calculado con tasa mora diaria compuesta.",
     "moraCompuestaDiaria=true aplica interés sobre interés."),
    # Renegociar mora
    ("Renegociar mora", "PATCH /api/prestamos/:id moraRenegociada",
     "Positivo", "Media",
     "Préstamo con mora acumulada",
     "1. PATCH moraRenegociada con acción y observación",
     '{"moraRenegociada":50000,"moraRenegociadaAccion":"CONDONACION","moraRenegociadaObservacion":"Cliente cumplió parcial"}',
     "200 OK. moraRenegociada registrada con fecha y usuario.",
     "AuditLog registra MORA_RENEGOCIADA. moraRenegociadaPorId poblado."),
]

# ============================================================
# 4. PAGOS
# ============================================================
MODULES["4. Pagos"] = [
    # Registrar pago
    ("Registrar pago completo", "POST /api/pagos con monto total de cuota",
     "Positivo", "Crítica",
     "Préstamo ACTIVO con cuota pendiente",
     "1. POST /api/pagos\n2. Verificar saldo actualizado\n3. Verificar recibo generado",
     '{"prestamoId":"...","montoTotal":50000,"metodoPago":"PSE","cuentaRecaudoId":"..."}',
     "201 Created con código P-XXXX. Pago estado=CONFIRMADO. Préstamo.saldoTotal reducido.",
     "cuotasPagadas incrementado. reciboHash generado. Bitácora registra PAGO_CONFIRMADO."),
    # Pago parcial
    ("Registrar pago parcial", "POST /api/pagos con monto < cuota",
     "Positivo", "Alta",
     "Préstamo ACTIVO",
     "1. POST /api/pagos con montoTotal < montoCuota",
     '{"montoTotal":20000,"montoCuota":50000}',
     "201 Created. Pago marcado como parcial. saldoTotal reducido proporcional.",
     "Bitácora registra PAGO_PARCIAL. No marca cuota como pagada."),
    # Pago solo intereses
    ("Pago solo intereses", "POST /api/pagos esSoloIntereses=true",
     "Positivo", "Alta",
     "Préstamo ACTIVO, cliente paga solo intereses",
     "1. POST /api/pagos esSoloIntereses=true",
     '{"esSoloIntereses":true,"montoInteres":10000}',
     "201 Created. cuotaAplazadaDe registrada. fechaOriginalVencimiento preservada.",
     "Capital NO se reduce. Cuota se aplaza (no cuenta como pagada)."),
    # Anular pago
    ("Anular pago", "POST /api/pagos/:id/anular con motivo",
     "Positivo", "Alta",
     "Pago CONFIRMADO. Usuario ADMIN o GESTOR.",
     "1. POST /api/pagos/:id/anular\n2. Verificar reversión de saldo",
     '{"motivoAnulacion":"Pago devuelto por contracargo"}',
     "200 OK. Pago estado=ANULADO. Préstamo.saldoTotal restaurado.",
     "anuladoPorId registrado. fechaAnulacion=now(). AuditLog registra PAGO_ANULADO."),
    # Reversar pago
    ("Reversar pago", "POST /api/pagos/:id/reversar",
     "Seguridad", "Alta",
     "Pago CONFIRMADO. Solo ADMIN.",
     "1. POST /api/pagos/:id/reversar con motivo",
     '{"motivoReversion":"Error de registro"}',
     "200 OK. Pago estado=REVERSADO. Saldo restaurado.",
     "Solo ADMIN puede reversar. Bitácora registra REVERSION."),
    # Reversar pago ya reversado
    ("Reversar pago ya reversado", "POST /api/pagos/:id/reversar dos veces",
     "Negativo", "Media",
     "Pago ya está REVERSADO",
     "1. POST /api/pagos/:id/reversar (segunda vez)",
     '{"motivoReversion":"test"}',
     "409 Conflict con code:'PAGO_YA_REVERSADO'",
     "No se puede reversar dos veces."),
    # Validar comprobante
    ("Validar comprobante de pago", "POST /api/pagos/:id/validar-comprobante",
     "Positivo", "Media",
     "Pago con comprobanteUrl subido",
     "1. POST /api/pagos/:id/validar-comprobante",
     '{"comprobanteValidado":true}',
     "200 OK. comprobanteValidado=true. comprobanteValidadoPorId=usuarioActual.",
     "AuditLog registra COMPROBANTE_VALIDADO."),
    # Conciliación bancaria
    ("Conciliación bancaria automática", "POST /api/pagos/conciliar con archivo bancario",
     "Positivo", "Media",
     "Archivo de extracto bancario cargado",
     "1. POST /api/pagos/conciliar\n2. Verificar matches",
     '{"archivoBase64":"...","cuentaRecaudoId":"..."}',
     "200 OK con { matches, noMatches, totalProcesado }",
     "Pagos con referencia coincidente se marcan CONFIRMADO automáticamente."),
    # Link de pago
    ("Generar link de pago", "POST /api/pagos/:id/generar-link",
     "Positivo", "Media",
     "Préstamo ACTIVO con cuota pendiente",
     "1. POST /api/pagos/:id/generar-link\n2. Verificar link y expiración",
     '{}',
     "200 OK con linkPago y linkExpira (24h).",
     "Link incluye token único. Acceso al link registra en Bitácora."),
]

# ============================================================
# 5. CAJAS Y CONTABILIDAD
# ============================================================
MODULES["5. Cajas y Contabilidad"] = [
    # Movimiento caja
    ("Registrar movimiento de caja", "POST /api/cajas/:id/movimientos",
     "Positivo", "Alta",
     "Caja abierta por usuario actual",
     "1. POST /api/cajas/:id/movimientos\n2. Verificar saldo",
     '{"tipo":"INGRESO","monto":50000,"concepto":"Pago cliente X"}',
     "201 Created. MovimientoCaja con codigo único. Saldo de caja actualizado.",
     "tipo INGRESO suma, EGRESO resta. AuditLog registra MOVIMIENTO_CAJA."),
    # Cerrar caja
    ("Cerrar caja con cuadre", "POST /api/cajas/:id/cerrar",
     "Positivo", "Alta",
     "Caja abierta con movimientos",
     "1. POST /api/cajas/:id/cerrar\n2. Verificar cuadre",
     '{"saldoFinalEsperado":500000}',
     "200 OK. Caja estado=CERRADA. fechaCierre=now().",
     "Si saldoFinal != esperado: warning pero permite cerrar con observación."),
    # Conciliación bancaria
    ("Conciliación bancaria", "POST /api/contabilidad/conciliar",
     "Positivo", "Media",
     "Extracto bancario cargado",
     "1. POST /api/contabilidad/conciliar\n2. Revisar diferencias",
     '{"extracto":"...","cuentaId":"..."}',
     "200 OK con { conciliados, pendientes, diferencias }",
     "Movimientos bancarios matching con Pagos confirmados."),
    # Contabilidad unificada
    ("Vista contabilidad unificada", "GET /api/contabilidad/unificada",
     "Positivo", "Media",
     "Datos en Pagos, MovimientoCaja, CasoJuridico",
     "1. GET /api/contabilidad/unificada?desde=2026-01-01&hasta=2026-12-31",
     '?desde=2026-01-01&hasta=2026-12-31',
     "200 OK con asientos consolidados.",
     "Incluye ingresos (pagos), egresos (cajas), honorarios jurídicos."),
]

# ============================================================
# 6. PORTAL CLIENTE
# ============================================================
MODULES["6. Portal Cliente"] = [
    # Solicitar OTP chat
    ("Solicitar OTP chat — cliente válido", "POST /api/chat/otp accion=solicitar",
     "Positivo", "Crítica",
     "Cliente existe, activo, con email",
     "1. POST /api/chat/otp\n2. Verificar correo recibido con código",
     '{"accion":"solicitar","clienteId":"<id>"}',
     "200 OK con otpId, expiraEn (5min), destinatario enmascarado.",
     "OTP hasheado SHA-256 en OtpChat.codigoHash. OtpRegistro creado. Email enviado por SMTP/API Brevo."),
    # Sin email
    ("Solicitar OTP chat — sin email", "POST /api/chat/otp sin email en cliente",
     "Negativo", "Alta",
     "Cliente con email=null",
     "1. POST /api/chat/otp",
     '{"accion":"solicitar","clienteId":"<id-sin-email>"}',
     "400 Bad Request con code:'NO_EMAIL'",
     "Sistema exige email registrado. Mensaje pide contactar admin."),
    # Verificar OTP
    ("Verificar OTP chat correcto", "POST /api/chat/otp accion=verificar",
     "Positivo", "Crítica",
     "OTP activo no expirado, no usado",
     "1. POST /api/chat/otp accion=verificar con código correcto",
     '{"accion":"verificar","otpId":"<id>","codigo":"123456"}',
     "200 OK con sessionId y tokenExpira (2h).",
     "cliente.tokenSesion=sessionId. cliente.tokenExpira=ahora+2h. AccesoPortal registra LOGIN exitoso."),
    # 3 intentos fallidos
    ("Bloqueo OTP chat tras 3 intentos", "POST verificar 3 veces con código errado",
     "Negativo", "Alta",
     "OTP activo",
     "1. POST verificar código errado (1)\n2. Repetir (2)\n3. Repetir (3)",
     '{"accion":"verificar","codigo":"000000"} x3',
     "Tercer intento: 403 Forbidden con code:'BLOCKED' y minutosRestantes=15.",
     "OtpChat.bloqueado=true. fechaBloqueo=now(). Cliente debe esperar 15min."),
    # Login portal PIN
    ("Login portal con PIN", "POST /api/portal/auth action=login",
     "Positivo", "Alta",
     "Cliente con pinHash válido",
     "1. POST /api/portal/auth con cédula y PIN",
     '{"action":"login","cedula":"12345678","pin":"1234"}',
     "200 OK con token de sesión. ultimoAccesoPortal=now().",
     "pinIntentos reset a 0. AccesoPortal registra LOGIN_EXITOSO."),
    # PIN bloqueado
    ("Login portal con PIN bloqueado", "POST /api/portal/auth con pinBloqueadoHasta > now",
     "Negativo", "Alta",
     "Cliente.bloqueadoHasta > now()",
     "1. POST /api/portal/auth con cédula",
     '{"action":"login","cedula":"...","pin":"1234"}',
     "403 Forbidden con code:'PIN_BLOQUEADO' y minutosRestantes",
     "Sistema pide esperar. No se incrementan intentos."),
    # Chat conversaciones
    ("Listar conversaciones chat", "GET /api/chat/conversaciones con x-portal-token",
     "Positivo", "Media",
     "Cliente autenticado con tokenSesion válido",
     "1. GET /api/chat/conversaciones\n2. Header x-portal-token: <sessionId>",
     'Header: x-portal-token: <token>',
     "200 OK con array de conversaciones del cliente.",
     "Token validado contra cliente.tokenSesion. Si expiró: 401."),
    # Enviar mensaje
    ("Enviar mensaje al chat", "POST /api/chat/mensajes",
     "Positivo", "Media",
     "Conversación existe, cliente autenticado",
     "1. POST /api/chat/mensajes con texto",
     '{"conversacionId":"...","texto":"Hola"}',
     "201 Created con mensaje. Bot responde automáticamente.",
     "Mensaje con emisor='CLIENTE'. Bot procesa y responde con NLU."),
]

# ============================================================
# 7. PORTAL ABOGADO
# ============================================================
MODULES["7. Portal Abogado"] = [
    # Auth
    ("Login portal abogado", "POST /api/juridico/portal/auth",
     "Positivo", "Alta",
     "Abogado con usuario/clave válidos",
     "1. POST /api/juridico/portal/auth",
     '{"usuario":"abogado1","clave":"<pass>"}',
     "200 OK con token de abogado. abogadoAsignado=true.",
     "AuditLog registra ABOGADO_LOGIN."),
    # Listar casos
    ("Listar casos asignados", "GET /api/juridico/portal/casos",
     "Positivo", "Media",
     "Abogado autenticado con casos asignados",
     "1. GET /api/juridico/portal/casos",
     'Header: Authorization: Bearer <token>',
     "200 OK con array de CasoJuridico donde abogadoNombre coincide.",
     "Solo casos del abogado actual. No ve casos de otros."),
    # Chat abogado
    ("Chat abogado con cliente", "POST /api/juridico/portal/chat",
     "Positivo", "Media",
     "Abogado autenticado, caso activo",
     "1. POST /api/juridico/portal/chat con mensaje",
     '{"casoId":"...","texto":"Revisar documentos"}',
     "201 Created con mensaje. Cliente notificado.",
     "Mensaje con emisor='ABOGADO'."),
]

# ============================================================
# 8. JURÍDICO
# ============================================================
MODULES["8. Jurídico"] = [
    # Abrir caso
    ("Abrir caso jurídico", "POST /api/juridico",
     "Positivo", "Alta",
     "Préstamo con mora > 90 días",
     "1. POST /api/juridico con prestamoId y datos del caso",
     '{"prestamoId":"...","tipoProceso":"COBRO_JUDICIAL","valorReclamado":5000000}',
     "201 Created con estado='EN_PROCESO'. fechaApertura=now().",
     "Abogado asignado. Préstamo marcado como en proceso jurídico."),
    # Actualizar caso
    ("Actualizar caso — admisión demanda", "PATCH /api/juridico/:id",
     "Positivo", "Media",
     "Caso existe, estado EN_PROCESO",
     "1. PATCH /api/juridico/:id con fechaAdmision",
     '{"fechaAdmision":"2026-08-05","radicado":"2026-00123"}',
     "200 OK con caso actualizado.",
     "Bitácora de cronología actualizada."),
    # Cronología
    ("Agregar evento a cronología", "POST /api/juridico/:id/cronologia",
     "Positivo", "Media",
     "Caso existe",
     "1. POST /api/juridico/:id/cronologia con evento",
     '{"tipo":"AUDIENCIA","fecha":"2026-09-01","descripcion":"Audiencia preliminar"}',
     "201 Created. Cronología actualizada.",
     "Eventos ordenados por fecha. AuditLog registra CRONOLOGIA_ADD."),
    # Cerrar caso
    ("Cerrar caso con resultado", "PATCH /api/juridico/:id cerrar",
     "Positivo", "Media",
     "Caso EN_PROCESO",
     "1. PATCH cerrar con resultadoFinal",
     '{"estado":"CERRADO","resultadoFinal":"FAVORABLE","fechaCierre":"2026-12-31"}',
     "200 OK con caso cerrado.",
     "Honorarios pagados se registran. Préstamo actualizado."),
]

# ============================================================
# 9. NOTIFICACIONES
# ============================================================
MODULES["9. Notificaciones"] = [
    # Crear notificación
    ("Crear notificación", "POST /api/notificaciones",
     "Positivo", "Media",
     "ADMIN/GESTOR autenticado",
     "1. POST /api/notificaciones con tipo, mensaje, destino",
     '{"tipo":"RECORDATORIO_PAGO","mensaje":"Su cuota vence mañana","prestamoId":"..."}',
     "201 Created con estado='PENDIENTE'.",
     "NotificacionLog creado. Programado para envío."),
    # Marcar enviada
    ("Marcar notificación como enviada", "POST /api/notificaciones/:id/enviar",
     "Positivo", "Media",
     "Notificación PENDIENTE",
     "1. POST /api/notificaciones/:id/enviar",
     '{}',
     "200 OK. estado='ENVIADO'. fechaEnvio=now().",
     "AuditLog registra NOTIFICACION_ENVIADA."),
    # WhatsApp
    ("Enviar WhatsApp — plantilla OTP", "POST /api/whatsapp/send",
     "Positivo", "Alta",
     "Conexión WhatsApp API activa",
     "1. POST /api/whatsapp/send con plantilla OTP",
     '{"telefono":"3001234567","plantilla":"OTP","variables":{"codigo":"123456"}}',
     "200 OK con messageId de WhatsApp API.",
     "NotificacionLog creado con estado=ENVIADO. Si falla: estado=FALLIDO."),
    # Sin conexión
    ("Enviar WhatsApp sin conexión activa", "POST /api/whatsapp/send sin API",
     "Negativo", "Media",
     "ConexionAPI.WHATSAPP activa=false",
     "1. POST /api/whatsapp/send",
     '{"telefono":"...","plantilla":"..."}',
     "503 Service Unavailable con code:'NO_WHATSAPP_CONFIG'",
     "Sistema no intenta enviar. NotificacionLog estado=FALLIDO."),
    # Centro comunicaciones
    ("Centro de comunicaciones — broadcast", "POST /api/centro-comunicaciones/broadcast",
     "Positivo", "Media",
     "ADMIN autenticado, lista de destinatarios",
     "1. POST broadcast con mensaje y destinatarios[]",
     '{"mensaje":"...","destinatarios":["300...","301..."]}',
     "200 OK con { enviados, fallidos, total }",
     "Cada envío registra en NotificacionLog. Respeta rate limit del proveedor."),
]

# ============================================================
# 10. CONEXIONES API Y CORREOS
# ============================================================
MODULES["10. Conexiones API y Correos"] = [
    # Crear conexión SMTP
    ("Crear conexión SMTP", "POST /api/conexiones",
     "Positivo", "Crítica",
     "ADMIN autenticado",
     "1. POST /api/conexiones con tipo=EMAIL_SMTP y credenciales",
     '{"tipo":"EMAIL_SMTP","url":"smtp-relay.brevo.com:587","usuario":"user@smtp-brevo.com","password":"xsmtpsib-...","apiKey":"xkeysib-...","configuracionExtra":{"host":"smtp-relay.brevo.com","port":587,"secure":false,"fromName":"Sistema","fromEmail":"jsa@jsadr.com.co"}}',
     "201 Created. Password y apiKey cifrados AES-256-CBC. activa=true.",
     "ConexionAPI con tipo=EMAIL_SMTP. Cifrado con API_ENCRYPTION_KEY de .env."),
    # Probar conexión
    ("Probar conexión SMTP", "POST /api/email action=probar",
     "Positivo", "Alta",
     "Conexión EMAIL_SMTP activa configurada",
     "1. POST /api/email action=probar",
     '{"accion":"probar"}',
     "200 OK con { success:true, message:'Conexión SMTP verificada' }",
     "transporter.verify() exitoso. host:puerto en respuesta."),
    # Enviar correo prueba
    ("Enviar correo de prueba", "POST /api/email action=enviar-prueba",
     "Positivo", "Alta",
     "SMTP configurado o BREVO_API_KEY en env",
     "1. POST /api/email action=enviar-prueba con destinatario",
     '{"accion":"enviar-prueba","to":"admin@jsadr.co"}',
     "200 OK con { success:true, messageId }. EnvioCorreo estado=ENVIADO.",
     "Camino principal: Brevo HTTPS API. Fallback: SMTP. Dev: Ethereal."),
    # Desencripción falla
    ("Detectar credenciales no desencriptables", "POST /api/email action=probar con llave incorrecta",
     "Negativo", "Crítica",
     "API_ENCRYPTION_KEY de .env no coincide con la usada al cifrar",
     "1. POST /api/email action=probar\n2. Ver log del servidor",
     '{"accion":"probar"}',
     "200 pero smtpConfigurado:false. Log muestra: '[email][SMTP] password cifrado... Ejecuta: BREVO_API_KEY=... node scripts/save-brevo-creds.js'",
     "Sistema NO intenta enviar con credenciales corruptas. Mensaje claro guía al admin."),
    # Re-cifrar credenciales
    ("Re-cifrar credenciales Brevo", "Ejecutar scripts/save-brevo-creds.js",
     "Positivo", "Crítica",
     "BREVO_API_KEY y BREVO_SMTP_KEY obtenidas del panel Brevo",
     "1. node scripts/save-brevo-creds.js con env vars\n2. POST /api/email action=enviar-prueba",
     'BREVO_API_KEY=xkeysib-... BREVO_SMTP_KEY=xsmtpsib-... node scripts/save-brevo-creds.js',
     "Credenciales re-cifradas con API_ENCRYPTION_KEY actual. Prueba de envío exitosa.",
     "Correo de prueba llega a la bandeja. EnvioCorreo estado=ENVIADO."),
    # Listar conexiones
    ("Listar conexiones API", "GET /api/conexiones",
     "Positivo", "Media",
     "ADMIN autenticado, varias conexiones en BD",
     "1. GET /api/conexiones",
     '{}',
     "200 OK con array. Passwords nunca se exponen.",
     "Solo metadata: tipo, activa, url, usuario, createdAt. No password/apiKey."),
    # Probar conexión inactiva
    ("Probar conexión inactiva", "POST /api/conexiones/:id/probar con activa=false",
     "Negativo", "Baja",
     "Conexión existe pero activa=false",
     "1. POST /api/conexiones/:id/probar",
     '{}',
     "400 Bad Request con code:'CONEXION_INACTIVA'",
     "Sistema no prueba conexiones inactivas."),
    # BorrarEnvioCorreo — auditoría
    ("Auditar envío de correo", "GET /api/email/auditoria after sending",
     "Positivo", "Media",
     "Envío de correo realizado",
     "1. Realizar envío\n2. Verificar EnvioCorreo creado",
     '{}',
     "EnvioCorreo con destinatario, asunto, estado=ENVIADO, via='BREVO_HTTPS_API'.",
     "Si falla: estado=FALLIDO con mensajeError."),
]

# ============================================================
# 11. DOCUMENTOS Y GESTIÓN DOCUMENTAL
# ============================================================
MODULES["11. Documentos y Gestión Documental"] = [
    # Subir documento
    ("Subir documento", "POST /api/documentos",
     "Positivo", "Alta",
     "Préstamo existe, usuario autenticado",
     "1. POST /api/documentos con archivoBase64",
     '{"prestamoId":"...","tipo":"CONTRATO","titulo":"Contrato","archivoBase64":"data:application/pdf;base64,...","archivoNombre":"contrato.pdf"}',
     "201 Created con documentoGestor.id.",
     "archivoHash SHA-256 calculado. subidoPor registrado. Tamaño validado."),
    # Descargar
    ("Descargar documento", "GET /api/documentos/:id",
     "Positivo", "Media",
     "Documento existe, usuario con permisos",
     "1. GET /api/documentos/:id",
     'GET /api/documentos/<id>',
     "200 OK con archivoBase64 y metadata.",
     "AuditLog registra DOCUMENTO_DESCARGADO."),
    # Snapshot
    ("Crear snapshot de proyecto", "POST /api/snapshots",
     "Positivo", "Media",
     "ADMIN autenticado",
     "1. POST /api/snapshots con descripción",
     '{"descripcion":"Snapshot pre-despliegue v5.1"}',
     "201 Created con snapshot.id. Estado=CREADO.",
     "Snapshot incluye estado completo de Préstamos, Clientes, Pagos."),
    # Restaurar snapshot
    ("Restaurar snapshot", "POST /api/snapshots/:id/restaurar",
     "Positivo", "Alta",
     "Snapshot existe, ADMIN autenticado",
     "1. POST /api/snapshots/:id/restaurar",
     '{}',
     "200 OK. Datos restaurados al estado del snapshot.",
     "AuditLog registra SNAPSHOT_RESTAURADO. Versión actual guardada como backup."),
    # Pagaré PDF
    ("Generar pagaré PDF", "POST /api/prestamos/:id/pagare",
     "Positivo", "Alta",
     "Préstamo ACTIVO con firma completada",
     "1. POST /api/prestamos/:id/pagare",
     '{}',
     "200 OK con PDF en base64. Incluye fotos cédula+selfie como respaldo.",
     "Hashes SHA-256 de fotos embebidos. Firma electrónica válida."),
]

# ============================================================
# 12. REPORTES Y EXPORTACIÓN
# ============================================================
MODULES["12. Reportes y Exportación"] = [
    # Dashboard
    ("Dashboard principal", "GET /api/dashboard",
     "Positivo", "Crítica",
     "Usuario autenticado",
     "1. GET /api/dashboard",
     '{}',
     "200 OK con KPIs: totalPréstamos, saldoTotal, pagosHoy, morosidad.",
     "Datos en tiempo real. Cálculos optimizados (índices DB)."),
    # Exportar Excel
    ("Exportar préstamos a Excel", "POST /api/export",
     "Positivo", "Alta",
     "ADMIN/GESTOR autenticado, datos en BD",
     "1. POST /api/export con tipo=PRESTAMOS y filtros",
     '{"tipo":"PRESTAMOS","formato":"xlsx","filtros":{"estado":"ACTIVO"}}',
     "200 OK con archivo xlsx en base64.",
     "Incluye todas las columnas relevantes. Respete filtros."),
    # Reporte accesos portal
    ("Reporte accesos portal", "GET /api/reportes/accesos-portal",
     "Positivo", "Media",
     "ADMIN autenticado, accesos registrados",
     "1. GET /api/reportes/accesos-portal?desde=2026-07-01",
     '?desde=2026-07-01',
     "200 OK con array de AccesoPortal.",
     "Incluye IP, userAgent, accion, exito. Filtrado por fecha."),
    # Reporte morosidad
    ("Reporte morosidad", "GET /api/reportes?tipo=MOROSIDAD",
     "Positivo", "Media",
     "Préstamos con mora en BD",
     "1. GET /api/reportes?tipo=MOROSIDAD",
     '?tipo=MOROSIDAD',
     "200 OK con préstamos en mora, días de mora, monto.",
     "Ordenado por diasMora descendente. Total al final."),
]

# ============================================================
# 13. AUTOMATIZACIONES
# ============================================================
MODULES["13. Automatizaciones"] = [
    # Crear automatización
    ("Crear automatización", "POST /api/automatizaciones",
     "Positivo", "Media",
     "ADMIN autenticado",
     "1. POST /api/automatizaciones con trigger y acción",
     '{"nombre":"Recordar pago","trigger":"CRON_DIARIO","accion":"ENVIAR_NOTIFICACION","config":{"hora":"08:00"}}',
     "201 Created con activa=true.",
     "AuditLog registra AUTOMATIZACION_CREATE."),
    # Ejecutar
    ("Ejecutar automatización manual", "POST /api/automatizaciones/ejecutar",
     "Positivo", "Media",
     "Automatización existe",
     "1. POST /api/automatizaciones/ejecutar con id",
     '{"id":"<auto-id>"}',
     "200 OK con resultado de ejecución.",
     "Bitácora registra ejecución, éxito/fallo, tiempo de ejecución."),
    # Desactivar
    ("Desactivar automatización", "PATCH /api/automatizaciones/:id activa=false",
     "Positivo", "Baja",
     "Automatización existe y activa",
     "1. PATCH /api/automatizaciones/:id",
     '{"activa":false}',
     "200 OK con activa=false.",
     "No se ejecuta automáticamente. Queda en BD para reactivar."),
]

# ============================================================
# 14. CONFIGURACIÓN DEL SISTEMA
# ============================================================
MODULES["14. Configuración del Sistema"] = [
    # Config global
    ("Actualizar configuración global", "PATCH /api/configuracion",
     "Positivo", "Alta",
     "ADMIN autenticado",
     "1. PATCH /api/configuracion con clave/valor",
     '{"clave":"TASA_MORA_DIARIA_DEFAULT","valor":"0.000666"}',
     "200 OK. Configuracion actualizada o creada (upsert).",
     "AuditLog registra CONFIG_UPDATE."),
    # Versiones
    ("Activar versión del sistema", "POST /api/versiones/:id/activar",
     "Positivo", "Alta",
     "Versión existe, ADMIN autenticado",
     "1. POST /api/versiones/:id/activar",
     '{}',
     "200 OK con versión activa.",
     "Versión anterior marcada inactiva. AuditLog registra VERSION_ACTIVADA."),
    # Usuarios CRUD
    ("Crear usuario GESTOR", "POST /api/usuarios",
     "Positivo", "Alta",
     "ADMIN autenticado, username no existe",
     "1. POST /api/usuarios con rol=GESTOR",
     '{"username":"gestor1","password":"<pass>","nombre":"Juan G","rol":"GESTOR","email":"gestor@jsadr.co"}',
     "201 Created. passwordHash bcrypt rounds=12. activo=true.",
     "mustChangePassword=true (primer login pide cambio)."),
    # Cambiar rol
    ("Cambiar rol usuario", "PATCH /api/usuarios/:id rol",
     "Seguridad", "Alta",
     "Usuario existe, ADMIN autenticado",
     "1. PATCH /api/usuarios/:id con nuevo rol",
     '{"rol":"CONSULTOR"}',
     "200 OK con rol actualizado.",
     "AuditLog registra ROL_CHANGE. Tokens JWT previos invalidados."),
    # Eliminar usuario
    ("Eliminar usuario (soft delete)", "DELETE /api/usuarios/:id",
     "Positivo", "Media",
     "Usuario existe, ADMIN autenticado",
     "1. DELETE /api/usuarios/:id",
     '{}',
     "200 OK. activo=false.",
     "No se borra físicamente. AuditLog registra USER_DEACTIVATED."),
    # MFA setup view
    ("Vista MFA setup", "GET /api/auth/mfa (vista setup)",
     "Positivo", "Media",
     "Usuario autenticado sin MFA",
     "1. GET /api/auth/mfa action=setup",
     '{"action":"setup"}',
     "200 OK con totpSecret nuevo y QR code base64.",
     "totpSecret no se guarda hasta verify. QR incluye otpauth URL."),
]

# ============================================================
# 15. AUDITORÍA Y BITÁCORA
# ============================================================
MODULES["15. Auditoría y Bitácora"] = [
    # AuditLog inmutable
    ("AuditLog inmutable — intentar update", "PATCH /api/audit-log/:id",
     "Seguridad", "Crítica",
     "AuditLog existe, ADMIN autenticado",
     "1. PATCH /api/audit-log/:id",
     '{"detalle":"modificado"}',
     "Error: 'AuditLog es inmutable: no se permite delete/update'",
     "Prisma extension bloquea mutación. No se puede modificar ni borrar."),
    # AuditLog inmutable delete
    ("AuditLog inmutable — intentar delete", "DELETE /api/audit-log/:id",
     "Seguridad", "Crítica",
     "AuditLog existe",
     "1. DELETE /api/audit-log/:id",
     '{}',
     "Error: 'AuditLog es inmutable: no se permite delete/update'",
     "Inmutable a nivel de Prisma client. Protegido contra borrado malicioso."),
    # Listar audit
    ("Listar audit log", "GET /api/audit-log",
     "Positivo", "Media",
     "ADMIN autenticado, eventos en BD",
     "1. GET /api/audit-log?page=1&pageSize=50",
     '?page=1&pageSize=50',
     "200 OK con eventos paginados.",
     "Ordenado por fecha desc. Filtros por modulo, usuario, accion."),
    # Bitácora préstamo
    ("Bitácora préstamo", "GET /api/prestamos/:id/bitacora",
     "Positivo", "Media",
     "Préstamo existe",
     "1. GET /api/prestamos/:id/bitacora",
     '{}',
     "200 OK con array de BitacoraPrestamo.",
     "Eventos: CREACION, PAGO, FIRMA, MORA_RENEGOCIADA, etc."),
    # Accesos portal
    ("Accesos portal — registrar", "POST internally from /api/chat/otp",
     "Positivo", "Media",
     "Cliente solicita OTP",
     "1. POST /api/chat/otp accion=solicitar\n2. Verificar AccesoPortal creado",
     '{}',
     "AccesoPortal creado con accion='CONSULTA', exito=true.",
     "ipOrigen y userAgent registrados."),
]

# ============================================================
# 16. SOLICITUDES DE CLIENTES
# ============================================================
MODULES["16. Solicitudes de Clientes"] = [
    # Crear solicitud web
    ("Solicitud web — pública", "POST /api/solicitudes-web",
     "Positivo", "Alta",
     "Endpoint público (sin auth)",
     "1. POST /api/solicitudes-web con datos",
     '{"nombre":"Juan","cedula":"123","email":"juan@x.com","telefono":"300...","montoSolicitado":1000000,"plazoMeses":12}',
     "201 Created con estado='PENDIENTE_REVISION'.",
     "Sistema valida email, cédula. Rate limit por IP. Notifica a gestores."),
    # Sin datos
    ("Solicitud web sin campos obligatorios", "POST /api/solicitudes-web",
     "Negativo", "Media",
     "Endpoint público",
     "1. POST /api/solicitudes-web sin email",
     '{"nombre":"X"}',
     "400 Bad Request con code:'MISSING_FIELDS'",
     "Campos obligatorios: nombre, cedula, email, telefono, monto."),
    # Aprobar solicitud
    ("Aprobar solicitud web", "POST /api/solicitudes-web/:id/aprobar",
     "Positivo", "Alta",
     "Solicitud existe, ADMIN/GESTOR autenticado",
     "1. POST /api/solicitudes-web/:id/aprobar\n2. Crear cliente + préstamo",
     '{"tasaInteresAnual":24,"plazoMeses":12}',
     "200 OK. Solicitud estado='APROBADA'. Cliente y Préstamo creados.",
     "Préstamo en estado PENDIENTE_ACEPTACION. Cliente recibe email."),
    # Rechazar
    ("Rechazar solicitud web", "POST /api/solicitudes-web/:id/rechazar",
     "Positivo", "Media",
     "Solicitud existe, ADMIN/GESTOR autenticado",
     "1. POST /api/solicitudes-web/:id/rechazar con motivo",
     '{"motivo":"No cumple políticas"}',
     "200 OK. estado='RECHAZADA'.",
     "AuditLog registra SOLICITUD_RECHAZADA."),
    # Consultar por cédula
    ("Consultar solicitud por cédula", "GET /api/solicitudes-web/cliente/:cedula",
     "Positivo", "Media",
     "Solicitud existe para la cédula",
     "1. GET /api/solicitudes-web/cliente/12345678",
     'GET /api/solicitudes-web/cliente/12345678',
     "200 OK con solicitudes del cliente.",
     "Solo ADMIN/GESTOR pueden consultar."),
]

# ============================================================
# 17. BOTS Y ASISTENTES IA
# ============================================================
MODULES["17. Bots y Asistentes IA"] = [
    # Iniciar conversación
    ("Iniciar conversación con bot", "POST /api/chat/iniciar",
     "Positivo", "Media",
     "Cliente autenticado en portal",
     "1. POST /api/chat/iniciar\n2. Bot saluda",
     '{"clienteId":"...","tipo":"GENERAL"}',
     "201 Created con conversacion.id. Bot envía mensaje inicial.",
     "NLU procesa entrada inicial. Conversacion creada con estado ACTIVA."),
    # Mensaje del bot
    ("Bot responde a pregunta frecuente", "POST /api/chat/mensajes",
     "Positivo", "Media",
     "Conversación activa",
     "1. POST /api/chat/mensajes con pregunta\n2. Bot procesa y responde",
     '{"texto":"¿Cuándo vence mi cuota?"}',
     "201 Created con respuesta del bot + metadata.",
     "NLU identifica intent SALDO_CUOTA. Responde con info del préstamo."),
    # Entrenar bot
    ("Entrenar bot con dataset", "POST /api/bots/entrenar",
     "Positivo", "Baja",
     "ADMIN autenticado, dataset cargado",
     "1. POST /api/bots/entrenar con dataset",
     '{"botId":"...","dataset":[{"intent":"SALUDO","ejemplos":["hola","buenas"]}]}',
     "200 OK con métricas de entrenamiento.",
     "Modelo NLU actualizado. Versionado."),
]

# ============================================================
# 18. CAMPAÑAS DE MARKETING
# ============================================================
MODULES["18. Campañas de Marketing"] = [
    # Crear campaña
    ("Crear campaña", "POST /api/campanas",
     "Positivo", "Media",
     "ADMIN autenticado",
     "1. POST /api/campanas con nombre, mensaje, destinatarios",
     '{"nombre":"Campaña Agosto","mensaje":"Oferta especial","destinatarios":["300...","301..."]}',
     "201 Created con estado='PENDIENTE'.",
     "Campaña con total destinatarios. Programada o inmediata."),
    # Enviar campaña
    ("Enviar campaña", "POST /api/campanas/:id/enviar",
     "Positivo", "Media",
     "Campaña PENDIENTE, ADMIN autenticado",
     "1. POST /api/campanas/:id/enviar",
     '{}',
     "200 OK con { enviados, fallidos }. estado='ENVIADA'.",
     "Cada envío registra en NotificacionLog. Respeta rate limit."),
    # Cancelar
    ("Cancelar campaña", "POST /api/campanas/:id/cancelar",
     "Positivo", "Baja",
     "Campaña PENDIENTE o ENVIANDO",
     "1. POST /api/campanas/:id/cancelar",
     '{}',
     "200 OK. estado='CANCELADA'.",
     "No se envían más mensajes. Ya enviados no se revierten."),
]


# ============================================================
# FUNCIÓN PARA ESCRIBIR UN SHEET DE MÓDULO
# ============================================================
def escribir_sheet_modulo(ws, modulo_nombre, casos, inicio_id=1):
    """Escribe un sheet de módulo con los casos de prueba."""
    titulo = f"Casos de Prueba QA — {modulo_nombre}"
    last_col = 1 + len(TEST_CASE_HEADERS)  # B..M = 12 cols
    setup_sheet(ws, title=titulo, last_col=last_col)

    # Escribir headers en fila 4
    header_row = 4
    for col_idx, header in enumerate(TEST_CASE_HEADERS, start=2):
        ws.cell(row=header_row, column=col_idx, value=header)
    style_header_row(ws, row_num=header_row, col_start=2, col_end=last_col)

    # Aplicar anchos de columna
    for col_letter, width in COLUMN_WIDTHS_TEST.items():
        ws.column_dimensions[col_letter].width = width

    # Escribir casos
    data_start_row = 5
    for i, caso in enumerate(casos):
        row_num = data_start_row + i
        # Casos pueden tener 9 o 10 elementos (notas es opcional)
        if len(caso) == 10:
            funcion, caso_prueba, tipo, prioridad, pre, pasos, datos, esperado, criterios, notas = caso
        elif len(caso) == 9:
            funcion, caso_prueba, tipo, prioridad, pre, pasos, datos, esperado, criterios = caso
            notas = ""
        else:
            raise ValueError(f"Caso con {len(caso)} elementos (esperado 9 o 10): {caso[:2]}")
        estado = "Pendiente"
        caso_id = f"TC-{inicio_id + i:04d}"

        ws.cell(row=row_num, column=2, value=caso_id)
        ws.cell(row=row_num, column=3, value=funcion)
        ws.cell(row=row_num, column=4, value=caso_prueba)
        ws.cell(row=row_num, column=5, value=tipo)
        ws.cell(row=row_num, column=6, value=prioridad)
        ws.cell(row=row_num, column=7, value=pre)
        ws.cell(row=row_num, column=8, value=pasos)
        ws.cell(row=row_num, column=9, value=datos)
        ws.cell(row=row_num, column=10, value=esperado)
        ws.cell(row=row_num, column=11, value=criterios)
        ws.cell(row=row_num, column=12, value=estado)
        ws.cell(row=row_num, column=13, value=notas)

        style_data_row(ws, row_num=row_num, col_start=2, col_end=last_col, row_index=i)

        # Alinear texto a la izquierda con wrap para celdas largas
        for col in [3, 4, 7, 8, 9, 10, 11, 13]:
            ws.cell(row=row_num, column=col).alignment = Alignment(
                horizontal='left', vertical='top', wrap_text=True
            )
        # Centro para Tipo, Prioridad, Estado
        for col in [2, 5, 6, 12]:
            ws.cell(row=row_num, column=col).alignment = Alignment(
                horizontal='center', vertical='top', wrap_text=True
            )

        # Altura adaptativa según contenido
        max_len = max(len(str(pasos or "")), len(str(esperado or "")), len(str(pre or "")))
        if max_len > 200:
            ws.row_dimensions[row_num].height = 90
        elif max_len > 100:
            ws.row_dimensions[row_num].height = 60
        else:
            ws.row_dimensions[row_num].height = 36

    # Congelar paneles (header siempre visible)
    ws.freeze_panes = 'C5'

    # Filtros automáticos en headers
    last_col_letter = get_column_letter(last_col)
    ws.auto_filter.ref = f"B{header_row}:{last_col_letter}{data_start_row + len(casos) - 1}"

    # Conditional formatting para Prioridad (col F = 6)
    prioridad_col = "F"
    data_range_prioridad = f"{prioridad_col}{data_start_row}:{prioridad_col}{data_start_row + len(casos) - 1}"
    # Crítica = rojo
    ws.conditional_formatting.add(
        data_range_prioridad,
        CellIsRule(operator='equal', formula=['"Crítica"'],
                   fill=PatternFill('solid', fgColor='FDEDEC'),
                   font=Font(name=FONT_NAME, color=ACCENT_NEGATIVE, bold=True))
    )
    ws.conditional_formatting.add(
        data_range_prioridad,
        CellIsRule(operator='equal', formula=['"Alta"'],
                   fill=PatternFill('solid', fgColor='FEF9E7'),
                   font=Font(name=FONT_NAME, color=ACCENT_WARNING, bold=True))
    )
    ws.conditional_formatting.add(
        data_range_prioridad,
        CellIsRule(operator='equal', formula=['"Media"'],
                   fill=PatternFill('solid', fgColor='E8F4FD'),
                   font=Font(name=FONT_NAME, color=PRIMARY))
    )

    # Conditional formatting para Estado (col L = 12)
    estado_col = "L"
    data_range_estado = f"{estado_col}{data_start_row}:{estado_col}{data_start_row + len(casos) - 1}"
    ws.conditional_formatting.add(
        data_range_estado,
        CellIsRule(operator='equal', formula=['"Aprobado"'],
                   fill=PatternFill('solid', fgColor='E8F5E9'),
                   font=Font(name=FONT_NAME, color=ACCENT_POSITIVE, bold=True))
    )
    ws.conditional_formatting.add(
        data_range_estado,
        CellIsRule(operator='equal', formula=['"Fallido"'],
                   fill=PatternFill('solid', fgColor='FDEDEC'),
                   font=Font(name=FONT_NAME, color=ACCENT_NEGATIVE, bold=True))
    )
    ws.conditional_formatting.add(
        data_range_estado,
        CellIsRule(operator='equal', formula=['"En Progreso"'],
                   fill=PatternFill('solid', fgColor='FEF9E7'),
                   font=Font(name=FONT_NAME, color=ACCENT_WARNING, bold=True))
    )

    # Configurar impresión
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = '4:4'

    return inicio_id + len(casos)


# ============================================================
# SHEET ÍNDICE
# ============================================================
def escribir_sheet_indice(ws, modulos_info):
    """Sheet índice con resumen de módulos."""
    titulo = "Índice — Casos de Prueba QA · Sistema Jsadr v5.0"
    last_col = 6  # B..F
    setup_sheet(ws, title=titulo, last_col=last_col)

    # Subtítulo
    ws.cell(row=3, column=2, value="Listado maestro de módulos y conteo de casos de prueba").font = font_caption()

    # Headers
    headers_indice = ["#", "Módulo", "Casos", "Críticas", "Altas", "Hoja"]
    header_row = 4
    for col_idx, h in enumerate(headers_indice, start=2):
        ws.cell(row=header_row, column=col_idx, value=h)
    style_header_row(ws, row_num=header_row, col_start=2, col_end=last_col)

    # Datos
    data_start_row = 5
    total_casos = 0
    total_criticas = 0
    total_altas = 0
    for i, info in enumerate(modulos_info):
        row_num = data_start_row + i
        ws.cell(row=row_num, column=2, value=i + 1)
        ws.cell(row=row_num, column=3, value=info["nombre"])
        ws.cell(row=row_num, column=4, value=info["casos"])
        ws.cell(row=row_num, column=5, value=info["criticas"])
        ws.cell(row=row_num, column=6, value=info["altas"])
        # Sheet name con hyperlink
        sheet_cell = ws.cell(row=row_num, column=7, value=info["sheet_name"])
        sheet_cell.hyperlink = f"#'{info['sheet_name']}'!A1"
        sheet_cell.font = Font(name=FONT_NAME, color=PRIMARY, underline="single")
        style_data_row(ws, row_num=row_num, col_start=2, col_end=last_col + 1, row_index=i)
        ws.cell(row=row_num, column=2).alignment = Alignment(horizontal='center')
        ws.cell(row=row_num, column=4).alignment = Alignment(horizontal='right')
        ws.cell(row=row_num, column=5).alignment = Alignment(horizontal='right')
        ws.cell(row=row_num, column=6).alignment = Alignment(horizontal='right')
        ws.cell(row=row_num, column=3).alignment = Alignment(horizontal='left', vertical='center')
        ws.cell(row=row_num, column=7).alignment = Alignment(horizontal='center', vertical='center')
        total_casos += info["casos"]
        total_criticas += info["criticas"]
        total_altas += info["altas"]

    # Fila de totales
    total_row = data_start_row + len(modulos_info)
    ws.cell(row=total_row, column=2, value="")
    ws.cell(row=total_row, column=3, value="TOTAL")
    ws.cell(row=total_row, column=4, value=total_casos)
    ws.cell(row=total_row, column=5, value=total_criticas)
    ws.cell(row=total_row, column=6, value=total_altas)
    ws.cell(row=total_row, column=7, value="")
    style_total_row(ws, row_num=total_row, col_start=2, col_end=last_col + 1)
    ws.cell(row=total_row, column=3).alignment = Alignment(horizontal='left')
    ws.cell(row=total_row, column=4).alignment = Alignment(horizontal='right')
    ws.cell(row=total_row, column=5).alignment = Alignment(horizontal='right')
    ws.cell(row=total_row, column=6).alignment = Alignment(horizontal='right')

    # Anchos de columna
    ws.column_dimensions['B'].width = 6
    ws.column_dimensions['C'].width = 42
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 22

    # Notas al final
    notes_row = total_row + 3
    ws.cell(row=notes_row, column=2, value="Notas:").font = font_subheader()
    notas = [
        "• Cada hoja contiene los casos de prueba del módulo correspondiente.",
        "• Los IDs son únicos (TC-0001 a TC-XXXX) para trazabilidad.",
        "• Usar filtros automáticos en la fila de headers para筛选.",
        "• Estados: Pendiente (default), En Progreso, Aprobado, Fallido, Bloqueado.",
        "• Prioridades: Crítica (bloquea release), Alta, Media, Baja.",
        "• Tipos: Positivo, Negativo, Borde, Seguridad, Performance, Regresión.",
        "• Conditional formatting aplica colores a Prioridad y Estado.",
    ]
    for j, nota in enumerate(notas):
        cell = ws.cell(row=notes_row + 1 + j, column=2, value=nota)
        cell.font = font_caption()
        cell.alignment = Alignment(horizontal='left', vertical='center')

    ws.freeze_panes = 'B5'


# ============================================================
# SHEET ESTÁNDARES QA
# ============================================================
def escribir_sheet_estandares(ws):
    """Sheet con definiciones de estándares QA."""
    titulo = "Estándares QA Tester — Definiciones y Convenciones"
    last_col = 4  # B..D
    setup_sheet(ws, title=titulo, last_col=last_col)

    header_row = 4
    headers = ["Categoría", "Valor", "Definición"]
    for col_idx, h in enumerate(headers, start=2):
        ws.cell(row=header_row, column=col_idx, value=h)
    style_header_row(ws, row_num=header_row, col_start=2, col_end=last_col)

    definiciones = [
        # Tipos de prueba
        ("Tipo", "Positivo", "Verifica que el sistema funciona como se espera con datos válidos. Camino feliz."),
        ("Tipo", "Negativo", "Verifica que el sistema rechaza datos inválidos o acciones no permitidas. Maneja errores esperados."),
        ("Tipo", "Borde", "Prueba en los límites de los rangos permitidos (valores mínimos, máximos, justos en la frontera)."),
        ("Tipo", "Seguridad", "Valida controles de seguridad: RBAC, autenticación, autorización, rate limit, validación de input, prevención XSS/SQLi."),
        ("Tipo", "Performance", "Mide tiempos de respuesta, throughput, uso de recursos bajo carga."),
        ("Tipo", "Regresión", "Verifica que funcionalidades previas siguen funcionando tras cambios."),
        ("Tipo", "Integración", "Valida que múltiples módulos funcionen juntos end-to-end."),
        ("Tipo", "Usabilidad", "Valida la experiencia de usuario, claridad de mensajes, flujos intuitivos."),
        # Prioridades
        ("Prioridad", "Crítica", "Bloquea release. Funcionalidad core no funciona. Ej: login, crear préstamo, pagos."),
        ("Prioridad", "Alta", "Debe resolverse antes de release. Funcionalidad importante con workaround."),
        ("Prioridad", "Media", "Puede resolverse en próximo sprint. Funcionalidad secundaria afectada."),
        ("Prioridad", "Baja", "Cosmético, mejoras menores. No afecta funcionalidad."),
        # Estados
        ("Estado", "Pendiente", "Caso creado pero no ejecutado aún. Default inicial."),
        ("Estado", "En Progreso", "Tester está ejecutando el caso actualmente."),
        ("Estado", "Aprobado", "Caso ejecutado, resultado esperado verificado. Sistema pasa la prueba."),
        ("Estado", "Fallido", "Caso ejecutado pero el resultado NO coincide con el esperado. Bug reportado."),
        ("Estado", "Bloqueado", "No se puede ejecutar por dependencia externa (ej: credenciales, entorno)."),
        # Tipos de datos
        ("Dato", "Pre-condiciones", "Estado inicial necesario en BD, configuración, autenticación, etc."),
        ("Dato", "Pasos", "Secuencia numerada de acciones a ejecutar para reproducir el caso."),
        ("Dato", "Datos de Entrada", "Body JSON, query params, headers, archivos necesarios."),
        ("Dato", "Resultado Esperado", "Respuesta HTTP, código de estado, cambios en BD, efectos secundarios esperados."),
        ("Dato", "Criterios de Aceptación", "Condiciones objetivas que deben cumplirse para aprobar el caso."),
        ("Dato", "Notas", "Información adicional, referencias a issues, decisiones de diseño."),
        # Convenciones de ID
        ("Convención", "ID TC-XXXX", "Test Case ID único secuencial. 4 dígitos. Mantiene trazabilidad entre hojas."),
        ("Convención", "Sheet por módulo", "Cada módulo tiene su propia hoja. Facilita filtrado y asignación a testers especializados."),
        ("Convención", "Freeze panes", "Headers y columna ID siempre visibles al hacer scroll."),
        ("Convención", "Auto-filter", "Filtros automáticos en fila de headers para筛选 por tipo, prioridad, estado."),
        # Estrategia QA
        ("Estrategia", "Cobertura", "Mínimo 1 caso positivo + 1 negativo por función. Casos de borde en validaciones numéricas."),
        ("Estrategia", "Seguridad primero", "Cada endpoint con RBAC debe tener caso negativo por rol."),
        ("Estrategia", "Trazabilidad", "Cada bug reportado debe referenciar el ID del caso TC-XXXX que falló."),
        ("Estrategia", "Re-ejecución", "Tras fix, re-ejecutar caso TC-XXXX marcado como Fallido. Si pasa → Aprobado."),
        ("Estrategia", "Regresión", "Tras cada release, re-ejecutar todos los casos Críticos y Altos."),
    ]

    data_start_row = 5
    for i, (cat, val, defn) in enumerate(definiciones):
        row_num = data_start_row + i
        ws.cell(row=row_num, column=2, value=cat)
        ws.cell(row=row_num, column=3, value=val)
        ws.cell(row=row_num, column=4, value=defn)
        style_data_row(ws, row_num=row_num, col_start=2, col_end=last_col, row_index=i)
        ws.cell(row=row_num, column=2).alignment = Alignment(horizontal='center', vertical='top')
        ws.cell(row=row_num, column=3).alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        ws.cell(row=row_num, column=4).alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        ws.row_dimensions[row_num].height = 36

    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 80

    ws.freeze_panes = 'B5'
    ws.auto_filter.ref = f"B{header_row}:D{data_start_row + len(definiciones) - 1}"


# ============================================================
# MAIN
# ============================================================
def main():
    wb = Workbook()

    # Sheet 1: Índice
    ws_index = wb.active
    ws_index.title = "Índice"

    # Sheet 2: Estándares QA
    ws_estandares = wb.create_sheet("Estándares QA")
    escribir_sheet_estandares(ws_estandares)

    # Sheets 3+: Uno por módulo
    modulos_info = []
    proximo_id = 1
    for modulo_nombre, casos in MODULES.items():
        # Sheet name: máximo 31 chars, sin caracteres especiales
        sheet_name = modulo_nombre.replace(".", " ", 1)[:31]
        # Eliminar caracteres prohibidos por Excel: []:*?/\
        for c in "[]:*?/\\\\":
            sheet_name = sheet_name.replace(c, "")
        sheet_name = sheet_name.strip()[:31]

        ws_mod = wb.create_sheet(sheet_name)
        proximo_id_despues = escribir_sheet_modulo(ws_mod, modulo_nombre, casos, inicio_id=proximo_id)

        # Conteos
        criticas = sum(1 for c in casos if c[3] == "Crítica")
        altas = sum(1 for c in casos if c[3] == "Alta")
        modulos_info.append({
            "nombre": modulo_nombre,
            "sheet_name": sheet_name,
            "casos": len(casos),
            "criticas": criticas,
            "altas": altas,
        })
        proximo_id = proximo_id_despues

    # Escribir índice al final (necesita info de todos los módulos)
    escribir_sheet_indice(ws_index, modulos_info)

    # Mover Índice al principio (posición 0)
    wb.move_sheet(ws_index, offset=-(len(wb.sheetnames) - 1))

    # Metadata
    wb.properties.creator = "Z.ai · Sistema Jsadr v5.0"
    wb.properties.title = "Casos de Prueba QA — Sistema Jsadr"
    wb.properties.subject = "Listado maestro de casos de prueba QA por módulo"

    # Guardar
    output_path = "/home/z/my-project/download/QA-Casos-de-Prueba-Sistema-Jsadr.xlsx"
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    print(f"\n✓ Excel generado: {output_path}")
    print(f"  Total módulos: {len(MODULES)}")
    print(f"  Total casos: {sum(len(casos) for casos in MODULES.values())}")
    print(f"  Sheets: Índice + Estándares QA + {len(MODULES)} módulos = {len(wb.sheetnames)} hojas")


if __name__ == "__main__":
    main()
