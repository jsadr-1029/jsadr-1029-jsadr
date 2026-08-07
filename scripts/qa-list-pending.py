"""Lista todos los casos de prueba del Excel con su estado."""
from openpyxl import load_workbook

WB_PATH = "/home/z/my-project/upload/plan-pruebas-qa-jsadr.xlsx"
wb = load_workbook(WB_PATH, data_only=True)

print(f"Sheets: {wb.sheetnames}\n")

# Recorrer todos los sheets y listar test cases con su estado
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n{'='*80}")
    print(f"SHEET: {sheet_name}  (rows: {ws.max_row}, cols: {ws.max_column})")
    print(f"{'='*80}")
    
    # Detectar header row (buscar fila que contenga "ID" o "Caso")
    header_row = None
    for r in range(1, min(10, ws.max_row + 1)):
        row_vals = [str(ws.cell(row=r, column=c).value or '').lower() for c in range(1, ws.max_column + 1)]
        if any('id' in v or 'caso' in v or 'estado' in v for v in row_vals):
            header_row = r
            break
    
    if not header_row:
        print("  (no se detectó header)")
        continue
    
    headers = [str(ws.cell(row=header_row, column=c).value or '').strip() for c in range(1, ws.max_column + 1)]
    print(f"Headers (fila {header_row}): {headers}")
    
    # Encontrar índices de columnas clave
    col_id = None
    col_func = None
    col_estado = None
    col_caso = None
    for i, h in enumerate(headers):
        hl = h.lower()
        if hl == 'id' or hl.startswith('id'):
            col_id = i + 1
        if hl.startswith('estado'):
            col_estado = i + 1
        if 'caso' in hl:
            col_caso = i + 1
        if 'funci' in hl:
            col_func = i + 1
    
    if not col_id or not col_estado:
        print(f"  No se encontraron columnas ID/Estado (col_id={col_id}, col_estado={col_estado})")
        continue
    
    # Listar cada fila con su ID y estado
    pendientes = []
    aprobados = []
    fallidos = []
    for r in range(header_row + 1, ws.max_row + 1):
        tc_id = ws.cell(row=r, column=col_id).value
        estado = ws.cell(row=r, column=col_estado).value
        caso_desc = ws.cell(row=r, column=col_caso).value if col_caso else ''
        func_desc = ws.cell(row=r, column=col_func).value if col_func else ''
        
        if not tc_id:
            continue
        
        estado_str = str(estado or '').strip()
        if not estado_str or estado_str.lower() == 'none':
            estado_str = '(vacío)'
        
        line = f"  row {r:3d} | {str(tc_id):18s} | {estado_str:15s} | {str(func_desc or '')[:35]:35s} | {str(caso_desc or '')[:50]}"
        print(line)
        
        if 'pendiente' in estado_str.lower() or 'vacío' in estado_str.lower() or 'progreso' in estado_str.lower():
            pendientes.append((r, tc_id, func_desc, caso_desc))
        elif 'aprobado' in estado_str.lower():
            aprobados.append((r, tc_id))
        elif 'fall' in estado_str.lower() or 'fail' in estado_str.lower():
            fallidos.append((r, tc_id))

print(f"\n\n{'#'*80}")
print(f"RESUMEN:")
print(f"{'#'*80}")
