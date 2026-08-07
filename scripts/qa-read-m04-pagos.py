"""Detalle completo de TCs de M04-Pagos."""
from openpyxl import load_workbook
import sys

# Usar el Excel actualizado que está en download/ si existe, si no, el original en upload/
paths = [
    "/home/z/my-project/download/plan-pruebas-qa-jsadr-actualizado.xlsx",
    "/home/z/my-project/upload/plan-pruebas-qa-jsadr.xlsx",
]
wb = None
for p in paths:
    try:
        wb = load_workbook(p, data_only=True)
        print(f"Usando Excel: {p}")
        break
    except Exception as e:
        continue

if wb is None:
    print("ERROR: No se encontró ningún Excel de plan de pruebas")
    sys.exit(1)

print(f"Hojas disponibles: {wb.sheetnames}\n")

# Buscar la hoja de M04-Pagos
sheet_name = None
for sn in wb.sheetnames:
    if 'M04' in sn or ('Pago' in sn and 'M0' in sn):
        sheet_name = sn
        break

if sheet_name is None:
    # intentar buscar por nombre parcial
    for sn in wb.sheetnames:
        if 'Pago' in sn:
            sheet_name = sn
            break

if sheet_name is None:
    print("ERROR: No se encontró la hoja M04-Pagos")
    print(f"Hojas: {wb.sheetnames}")
    sys.exit(1)

ws = wb[sheet_name]
print(f"=== {ws.title}  max_row={ws.max_row} max_col={ws.max_column} ===\n")

# Cabecera - buscar en filas 1-5
for header_row in range(1, 6):
    print(f"\n--- Fila {header_row} (cabecera) ---")
    for c in range(1, ws.max_column + 1):
        h = ws.cell(row=header_row, column=c).value
        if h:
            print(f"  col {c:2d}: {str(h)[:100]}")

print("\n" + "=" * 80)
print("TCs ENCONTRADOS:")
print("=" * 80 + "\n")

# Buscar filas con TC-PAG
found = 0
for r in range(1, ws.max_row + 1):
    tc_id = ws.cell(row=r, column=2).value
    if not tc_id:
        # intentar columna 1
        tc_id = ws.cell(row=r, column=1).value
    if not tc_id or 'TC' not in str(tc_id).upper():
        continue
    tc_id_str = str(tc_id)
    if 'PAG' not in tc_id_str.upper() and 'M04' not in tc_id_str.upper():
        continue

    found += 1
    # Mostrar todas las columnas con datos
    print(f"━━━ row {r} | {tc_id_str} ━━━")
    for c in range(1, ws.max_column + 1):
        h = ws.cell(row=3, column=c).value or ws.cell(row=2, column=c).value or ws.cell(row=1, column=c).value
        v = ws.cell(row=r, column=c).value
        if v is not None and str(v).strip():
            print(f"  [{c}] {str(h or 'col' + str(c))[:40]:40s} = {str(v)[:300]}")
    print()

print(f"\nTotal TCs encontrados: {found}")
