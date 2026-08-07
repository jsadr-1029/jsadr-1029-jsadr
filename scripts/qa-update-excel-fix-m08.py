#!/usr/bin/env python3
"""Actualiza el Excel de plan de pruebas QA con los TCs aprobados de M08-Portal Jurídico."""
import openpyxl
from pathlib import Path
from openpyxl.styles import PatternFill, Font, Alignment

EXCEL_OUT = Path("/home/z/my-project/download/plan-pruebas-qa-jsadr-actualizado.xlsx")

TCS = {
    'TC-JUR-005': {
        'hallazgo': 'Cumple: SESSION_EXPIRY_HOURS=8 en auth, verificarTokenPortal valida tokenExpira < now en cada request (casos, chat).',
        'riesgo': 'N/A — sin riesgo (cumplimiento verificado).',
    },
    'TC-JUR-006': {
        'hallazgo': 'Cumple: existe DELETE /api/juridico/portal/auth que limpia tokenSesion=null y tokenExpira=null en BD, retorna HTTP 200.',
        'riesgo': 'N/A — sin riesgo (cumplimiento verificado).',
    },
    'TC-JUR-007': {
        'hallazgo': 'Cumple: GET /api/juridico/portal/casos filtra por abogadoEmail/abogadoNombre; GESTOR ve todos los no cerrados.',
        'riesgo': 'N/A — sin riesgo (cumplimiento verificado).',
    },
    'TC-JUR-008': {
        'hallazgo': 'No existía endpoint /api/juridico/portal/casos/[id] con auth. El /api/juridico/[id] existente NO validaba token del portal. Creado endpoint con verificarTokenPortal + includes (prestamo, cronologias, documentos, alertas).',
        'riesgo': 'Alto — abogados podían consultar cualquier caso sin autenticación; carencia funcional del portal.',
    },
    'TC-JUR-009': {
        'hallazgo': 'No existía endpoint POST en el portal para subir documentos. Creado /api/juridico/portal/casos/[id]/documentos con auth, transacción atómica (documentoLegal + cronologiaCaso), audit log y HTTP 201.',
        'riesgo': 'Medio — abogados no podían subir documentos desde el portal; el endpoint legacy /api/juridico/[id]/documentos no tenía auth.',
    },
    'TC-JUR-010': {
        'hallazgo': 'CRÍTICO: no existía validación de asignación del caso al abogado. Creada validación en /api/juridico/portal/casos/[id]: ABOGADO valida abogadoEmail/abogadoNombre → HTTP 403 CASO_AJENO + audit log de intento; GESTOR bypass.',
        'riesgo': 'CRÍTICO — cualquier abogado podía ver casos jurídicos de otros abogados; fuga de información legal/financiera.',
    },
    'TC-JUR-011': {
        'hallazgo': 'No existía endpoint para agregar notas internas al caso. Creado /api/juridico/portal/casos/[id]/notas-internas: auth + asignación + nota como CronologiaCaso (actor=usuario.nombre) + audit log NOTA_INTERNA_CREADA + HTTP 201.',
        'riesgo': 'Medio — no había forma de registrar notas internas del caso con trazabilidad de autor.',
    },
    'TC-JUR-012': {
        'hallazgo': 'No existía endpoint dedicado para ver bitácora. Creado /api/juridico/portal/casos/[id]/bitacora: auth + asignación + cronologias ordenadas por fecha (?order=asc|desc) + audit log + HTTP 200.',
        'riesgo': 'Medio — abogados debían consultar el detalle completo del caso para ver la cronología; carencia funcional.',
    },
    'TC-JUR-013': {
        'hallazgo': 'Cumple: rateLimit(`juridico-portal:${ip}`, 20) en POST /auth, retorna HTTP 429 "Demasiadas solicitudes" a partir de la 21ava solicitud/min.',
        'riesgo': 'N/A — sin riesgo (cumplimiento verificado).',
    },
    'TC-JUR-015': {
        'hallazgo': 'Solo el login registraba AuditLog. Añadido audit log a: logout (LOGOUT_PORTAL_JURIDICO), listar casos (CONSULTA_CASOS_JURIDICOS), ver caso (CONSULTA_CASO_JURIDICO + ACCESO_CASO_AJENO_BLOQUEADO), subir documento (CREATE), nota interna (NOTA_INTERNA_CREADA), ver bitácora (CONSULTA_BITACORA_CASO).',
        'riesgo': 'Alto — acciones críticas del portal jurídico no quedaban registradas; imposibilidad de auditoría forense.',
    },
}

# Cargar Excel
wb = openpyxl.load_workbook(EXCEL_OUT)
ws = wb['10. M08-Portal Jurídico']

# Estilos
fill_amarillo = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
fill_verde = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
font_amarilla = Font(color='9C5700', bold=True)
font_verde = Font(color='006100', bold=True)

# Buscar columnas
headers = {}
for col in range(1, ws.max_column + 1):
    v = ws.cell(row=4, column=col).value
    if v:
        headers[str(v).strip().lower()] = col

col_hallazgo = headers.get('hallazgo')
col_riesgo = headers.get('riesgo')
col_estado = headers.get('estado')
col_id = headers.get('id')

if col_hallazgo is None:
    col_hallazgo = ws.max_column + 1
    ws.cell(row=4, column=col_hallazgo, value='Hallazgo').font = Font(bold=True)
    col_riesgo = col_hallazgo + 1
    ws.cell(row=4, column=col_riesgo, value='Riesgo').font = Font(bold=True)

print(f"Columnas: ID={col_id}, Hallazgo={col_hallazgo}, Riesgo={col_riesgo}, Estado={col_estado}")

updated = 0
for row in range(5, ws.max_row + 1):
    tc_id = ws.cell(row=row, column=col_id).value if col_id else None
    if not tc_id:
        continue
    tc_id = str(tc_id).strip()
    if tc_id in TCS:
        info = TCS[tc_id]
        if col_estado:
            cell = ws.cell(row=row, column=col_estado, value='Aprobado')
            cell.fill = fill_verde
            cell.font = font_verde
            cell.alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=col_hallazgo, value=info['hallazgo']).fill = fill_amarillo
        ws.cell(row=row, column=col_hallazgo).font = font_amarilla
        ws.cell(row=row, column=col_hallazgo).alignment = Alignment(wrap_text=True, vertical='top')
        ws.cell(row=row, column=col_riesgo, value=info['riesgo']).fill = fill_amarillo
        ws.cell(row=row, column=col_riesgo).font = font_amarilla
        ws.cell(row=row, column=col_riesgo).alignment = Alignment(wrap_text=True, vertical='top')
        updated += 1
        print(f"  Actualizado {tc_id}")

print(f"\nTotal TCs actualizados: {updated}")
wb.save(EXCEL_OUT)
print(f"Excel guardado: {EXCEL_OUT}")
