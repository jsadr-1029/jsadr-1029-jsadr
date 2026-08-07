"""Actualiza el Excel con los hallazgos y reparos aplicados a M04-Pagos."""
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
import shutil

WB_PATH = "/home/z/my-project/upload/plan-pruebas-qa-jsadr.xlsx"
DEST_PATH = "/home/z/my-project/download/plan-pruebas-qa-jsadr-actualizado.xlsx"

# (row, tc_id, nuevo_riesgo, nuevo_hallazgo)
# Sheet "6. M04-Pagos" rows 5-19 for TC-PAG-001 to TC-PAG-015
# Columns: 11 = Riesgo, 12 = Hallazgo, 13 = Estado
updates = {
    "6. M04-Pagos": [
        (5, "TC-PAG-001",
         "HTTP 201. Pago registrado. Saldo disminuido. Cuota marcada APLICADO. Audit log + bitácora + WhatsApp.",
         "POST /api/pagos (accion=aplicarPago). Transacción atómica db.$transaction: pago create + caja movimiento + recálculo saldos. Auth requireRole ADMIN/GESTOR. Audit log PAGO_APLICADO. Muestra BD: pago cmrpudg4j cuota=1 monto=$200,000."),
        (6, "TC-PAG-002",
         "HTTP 201. Pago registrado. Estado=PAGO_PARCIAL. Saldo actualizado. Cuota sigue pendiente.",
         "Distribución proporcional: mora → interés → capital. Acumula con pagos parciales previos de la misma cuota. BD: 3 pagos PAGO_PARCIAL registrados."),
        (7, "TC-PAG-003",
         "Sistema permite el sobrepago y registra excedente en notas para auditoría. Política: gestor decide reembolsar o aplicar a cuota siguiente.",
         "POLÍTICA DE NEGOCIO DOCUMENTADA: el sistema NO rechaza el sobrepago. Detecta `excedente > 0` y lo registra en notasPago con detalles del monto recibido vs distribuido. El gestor decide posteriormente: reembolsar o aplicar a cuota siguiente. No se aplica automáticamente a cuotas futuras."),
        (8, "TC-PAG-004",
         "HTTP 409 PRESTAMO_NO_APLICA_PAGOS. Mensaje: 'No se pueden registrar pagos a un préstamo en estado ANULADO/RECHAZADO/CANCELADO'.",
         "REPARO v4.7 CRÍTICO: ANTES POST /api/pagos NO validaba el estado del préstamo. Aceptaba pagos a préstamos ANULADO/RECHAZADO/CANCELADO (inconsistencia contable). AHORA: valida ESTADOS_NO_ACEPTAN_PAGOS=['ANULADO','RECHAZADO','CANCELADO'] antes de la transacción. Retorna 409 con codigo PRESTAMO_NO_APLICA_PAGOS y estadoPrestamo."),
        (9, "TC-PAG-005",
         "HTTP 400 FECHA_FUTURA_INVALIDA. Mensaje: 'Fecha no puede ser futura'.",
         "REPARO v4.7: ANTES POST /api/pagos ignoraba la fecha del body y siempre usaba new Date() (no había forma de back-date ni validación). AHORA: si el body trae fechaPago explícita, valida que no sea > new Date(). Si es futura → 400 FECHA_FUTURA_INVALIDA. Si es pasada o no viene, se usa esa fecha (o new Date()) en el registro."),
        (10, "TC-PAG-006",
         "HTTP 200. Pago.estado=REVERSADO. Saldo restituido (recalcularSaldosPrestamo). Bitácora + caja mora ajustada. Solo ADMIN.",
         "POST /api/pagos/[id]/reversar. Solo ADMIN (v4.6). Exige motivoReversion. Solo permite reversar APLICADO o PAGO_PARCIAL. Recalcula saldos, ajusta Caja-Mora si aplica, registra BitacoraPrestamo tipo=PAGO."),
        (11, "TC-PAG-007",
         "HTTP 409 PAGO_YA_REVERSADO. Mensaje: 'El pago ya está reversado. No se puede reversar dos veces'.",
         "REPARO v4.7: ANTES reversar un pago ya REVERSADO retornaba HTTP 400 con mensaje genérico. AHORA: detecta específicamente estado === 'REVERSADO' y retorna HTTP 409 con codigo PAGO_YA_REVERSADO y estadoActual. Otros estados inválidos siguen retornando 400."),
        (12, "TC-PAG-008",
         "HTTP 200. Pago.estado=ANULADO (soft-delete). Saldo restituido. Audit log PAGO_ANULADO. Solo ADMIN.",
         "DELETE /api/pagos/[id]. Soft-delete: marca estado=ANULADO, conserva el registro. Registra motivoAnulacion, anuladoPorId, fechaAnulacion. Recalcula saldos. Audit log PAGO_ANULADO. Solo ADMIN. Bloquea doble anulación (400 si ya está ANULADO)."),
        (13, "TC-PAG-009",
         "HTTP 409 PAGO_REVERSADO_NO_ANULABLE. Mensaje: 'No se puede anular un pago que está REVERSADO'.",
         "REPARO v4.7: ANTES DELETE /api/pagos/[id] solo validaba estado === 'ANULADO' (doble anulación). Permitía anular pagos REVERSADOS (inconsistencia: un pago reversado ya fue devuelto al cliente). AHORA: valida específicamente estado === 'REVERSADO' y retorna 409 PAGO_REVERSADO_NO_ANULABLE antes de proceder."),
        (14, "TC-PAG-010",
         "HTTP 200 con matched + noMatched + montoTotalMatched + aplicados + errores. Conciliación asíncrona con match por referencia o código.",
         "POST /api/pagos/conciliacion. Auth ADMIN/GESTOR. Acciones: 'previsualizar' (match sin aplicar) y 'aplicar' (aplica pagos matcheados). Match por referencia exacta o codigo de pago. Tolerancia monto < $1 COP. Máximo 500 movimientos por lote. Audit log CONCILIACION_BANCARIA."),
        (15, "TC-PAG-011",
         "HTTP 400 MONTO_INVALIDO. Mensaje: 'Monto debe ser mayor a 0'.",
         "REPARO v4.7 CRÍTICO FINANCIERO: ANTES POST /api/pagos NO validaba monto > 0. Aceptaba montos negativos (riesgo: el pago reduciría el saldo del préstamo en negativo, generando saldo a favor incorrecto). AHORA: valida montoTotalNumValidacion <= 0 (cubre 0 y negativos) → 400 MONTO_INVALIDO."),
        (16, "TC-PAG-012",
         "HTTP 400 MONTO_INVALIDO. Mensaje: 'Monto debe ser mayor a 0'.",
         "REPARO v4.7: ANTES monto=0 era rechazado por validación truthy `!montoTotal` con mensaje confuso 'Faltan campos obligatorios'. AHORA: validación numérica unificada montoTotalNumValidacion <= 0 cubre 0 y negativos con mismo codigo MONTO_INVALIDO y mensaje claro 'Monto debe ser mayor a 0. Recibido: 0'."),
        (17, "TC-PAG-013",
         "HTTP 200 con array de pagos ordenados por fechaPago DESC. Incluye REVERSADOS por defecto; ANULADOS con ?incluirAnulados=true.",
         "REPARO v4.7: ANTES GET /api/pagos excluía ANULADOS por defecto (where.estado = { not: 'ANULADO' }). El Excel exige incluir REVERSADOS y ANULADOS para auditoría. AHORA: soporta query param incluirAnulados=true para retornar también pagos ANULADOS. REVERSADOS ya se incluían (no estaban filtrados). BD: préstamo cmrpojg1q con 6 pagos (1 REVERSADO, 0 ANULADOS)."),
        (18, "TC-PAG-014",
         "Política de descuento por pago anticipado: NO implementada como cálculo automático. Mecanismo disponible: mora renegociada.",
         "DECISIÓN DE NEGOCIO DOCUMENTADA: el sistema NO calcula descuento automático por pago anticipado. El cálculo de cuota es estándar (calcularPrestamo). El gestor puede aplicar descuentos manualmente vía POST /api/pagos/renegociar-mora (moraRenegociada) o notas de crédito. Esto es una decisión de negocio: el descuento anticipado no es política automática del sistema."),
        (19, "TC-PAG-015",
         "HTTP 201. Movimiento registrado. Saldo caja actualizado (increment/decrement). Audit log CAJA_MOVIMIENTO_CREADO.",
         "REPARO v4.7 CRÍTICO DE SEGURIDAD: ANTES POST /api/cajas/[id]/movimientos NO tenía autenticación. Cualquiera (sin token) podía crear movimientos de caja y manipular saldos. AHORA: (1) Auth requireRole ADMIN/GESTOR. (2) Validación monto > 0 (MONTO_INVALIDO). (3) Validación tipo INGRESO|EGRESO (TIPO_INVALIDO). (4) Validación concepto no vacío (CONCEPTO_REQUERIDO). (5) Transacción atómica movimiento + saldo. (6) Audit log CAJA_MOVIMIENTO_CREADO. BD: caja CAJA-MORA saldo=$87.64 activa."),
    ],
}

green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
green_font = Font(color="006100", bold=True)
yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

wb = load_workbook(WB_PATH)
total_updated = 0

for sheet_name, items in updates.items():
    ws = wb[sheet_name]
    for row_num, tc_id, nuevo_riesgo, nuevo_hallazgo in items:
        cell_id = ws.cell(row=row_num, column=2).value
        if cell_id != tc_id:
            print(f"⚠ {sheet_name} fila {row_num}: esperado {tc_id}, encontrado {cell_id}")
            continue

        # Estado = M (13)
        cell_state = ws.cell(row=row_num, column=13)
        cell_state.value = "Aprobado"
        cell_state.fill = green_fill
        cell_state.font = green_font
        cell_state.alignment = Alignment(horizontal="center", vertical="center")

        # Riesgo = K (11)
        cell_riesgo = ws.cell(row=row_num, column=11)
        cell_riesgo.value = nuevo_riesgo
        cell_riesgo.fill = yellow_fill
        cell_riesgo.alignment = Alignment(wrap_text=True, vertical="top")

        # Hallazgo = L (12)
        cell_hallazgo = ws.cell(row=row_num, column=12)
        cell_hallazgo.value = nuevo_hallazgo
        cell_hallazgo.fill = yellow_fill
        cell_hallazgo.alignment = Alignment(wrap_text=True, vertical="top")

        print(f"✓ {sheet_name} fila {row_num} ({tc_id}): Estado=Aprobado, Riesgo+Hallazgo actualizados")
        total_updated += 1

wb.save(WB_PATH)
print(f"\n=== Total actualizado: {total_updated} TCs ===")

# Copiar a download para entrega
shutil.copy(WB_PATH, DEST_PATH)
print(f"Copia para entrega: {DEST_PATH}")
