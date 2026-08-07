#!/usr/bin/env python3
"""Lee la hoja M10-Reportes del Excel de plan de pruebas."""
import openpyxl
from pathlib import Path

EXCEL = Path("/home/z/my-project/upload/plan-pruebas-qa-jsadr.xlsx")
wb = openpyxl.load_workbook(EXCEL, data_only=False)

sheet_name = "12. M10-Reportes"
ws = wb[sheet_name]
print(f"Hoja: {sheet_name!r} | Dim: {ws.dimensions} | max_row={ws.max_row}")

print("\n=== Contenido completo ===")
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
    cells = []
    for c in row:
        v = c.value if c.value is not None else ""
        cells.append(str(v).replace("\n", " | ")[:80])
    if any(cells):
        print(f"R{row[0].row:>3}: {cells}")
