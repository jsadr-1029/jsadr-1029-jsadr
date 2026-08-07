"""Lee los TCs pendientes de M05-Correo Electrónico."""
from openpyxl import load_workbook

WB = "/home/z/my-project/download/plan-pruebas-qa-jsadr-actualizado.xlsx"
wb = load_workbook(WB, data_only=True)
ws = wb["7. M05-Correo Electrónico"]

print(f"=== {ws.title}  max_row={ws.max_row} ===\n")

# Cabecera
for c in range(1, ws.max_column + 1):
    h = ws.cell(row=4, column=c).value
    if h:
        print(f"  col {c:2d}: {h}")
print()

# Listar todos los TCs y resaltar pendientes
print("=" * 80)
print("TODOS LOS TCs DE M05:")
print("=" * 80)
pendientes = []
for r in range(5, ws.max_row + 1):
    tc_id = ws.cell(row=r, column=2).value
    if not tc_id or 'TC' not in str(tc_id).upper():
        continue
    func = ws.cell(row=r, column=4).value
    caso = ws.cell(row=r, column=5).value
    tipo = ws.cell(row=r, column=6).value
    prioridad = ws.cell(row=r, column=7).value
    precond = ws.cell(row=r, column=8).value
    pasos = ws.cell(row=r, column=9).value
    datos = ws.cell(row=r, column=10).value
    esperado = ws.cell(row=r, column=11).value
    criterio = ws.cell(row=r, column=12).value
    estado = ws.cell(row=r, column=13).value
    estado_str = str(estado or '').strip()
    is_pending = estado_str.lower() in ('pendiente', '', 'no ejecutado')

    marker = '🔴' if is_pending else '🟢'
    print(f"\n{marker} row {r} | {tc_id} | estado='{estado_str}'")
    print(f"  Función:    {func}")
    print(f"  Caso:       {caso}")
    print(f"  Tipo:       {tipo} | Prioridad: {prioridad}")
    print(f"  Precond:    {str(precond or '')[:120]}")
    print(f"  Pasos:      {str(pasos or '')[:200]}")
    print(f"  Datos:      {str(datos or '')[:150]}")
    print(f"  Esperado:   {str(esperado or '')[:200]}")
    print(f"  Criterio:   {str(criterio or '')[:150]}")

    if is_pending:
        pendientes.append((r, tc_id, func, caso, pasos, datos, esperado, criterio))

print("\n" + "=" * 80)
print(f"PENDIENTES: {len(pendientes)} TCs")
print("=" * 80)
for r, tc_id, func, caso, *_ in pendientes:
    print(f"  row {r}: {tc_id} — {func} / {caso}")
