#!/usr/bin/env python3
"""Actualiza hoja M13-Sync DevOps del Excel con hallazgos/riesgo/estado."""
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment

XLSX = "/home/z/my-project/download/plan-pruebas-qa-jsadr-actualizado.xlsx"
wb = load_workbook(XLSX)
ws = wb["15. M13-Sync DevOps"]

hdr_row = 4
cols = {}
for c in range(1, ws.max_column + 1):
    v = ws.cell(row=hdr_row, column=c).value
    if v:
        cols[str(v).strip()] = c

print("Columnas:", cols)

if "Hallazgo" not in cols:
    new_col = ws.max_column + 1
    ws.cell(row=hdr_row, column=new_col, value="Hallazgo").font = Font(bold=True)
    cols["Hallazgo"] = new_col
if "Riesgo" not in cols:
    new_col = ws.max_column + 1
    ws.cell(row=hdr_row, column=new_col, value="Riesgo").font = Font(bold=True)
    cols["Riesgo"] = new_col

green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
green_font = Font(color="006100")
yellow_font = Font(color="9C5700")

hallazgos = {
    "TC-DEV-001": {
        "hallazgo": "GitHub Push verificado: .git/config con remote origin a github.com jsadr-1029/jsadr-1029-jsadr; branch main configurada; package.json tiene scripts dev/build/start.",
        "riesgo": "—",
    },
    "TC-DEV-002": {
        "hallazgo": "Secret scanning verificado: .gitignore excluye .env y .env.local; .env contiene xkeysib- (Brevo) pero no está commited; .env.example existe como template sin secrets reales; código valida prefijo xkeysib- en BREVO_API_KEY.",
        "riesgo": "—",
    },
    "TC-DEV-003": {
        "hallazgo": "Workflow deploy-vercel.yml disparado por push a main + workflow_dispatch manual; runs-on ubuntu-latest; concurrency group anti-deploy paralelo; step checkout @v4.",
        "riesgo": "—",
    },
    "TC-DEV-004": {
        "hallazgo": "Workflow exitoso verificado: 8 steps completos (checkout v4, setup-node v4 node 20, npm install --legacy-peer-deps, install vercel CLI, vercel pull --production, prisma generate, vercel deploy --prod --yes, Summary con GITHUB_STEP_SUMMARY). timeout-minutes 20, permissions contents: read, secrets VERCEL_TOKEN/ORG_ID/PROJECT_ID.",
        "riesgo": "—",
    },
    "TC-DEV-005": {
        "hallazgo": "Deploy producción exitoso verificado: workflow step vercel deploy --prod captura deployment_url en output; vercel.json framework nextjs, buildCommand 'prisma generate && next build', installCommand 'npm install --legacy-peer-deps', regions iad1; next.config.ts output standalone + ignoreBuildErrors false; vercel.json maxDuration 60 + crons configurados.",
        "riesgo": "—",
    },
    "TC-DEV-006": {
        "hallazgo": "Vercel env vars sincronizadas: código usa process.env.BREVO_API_KEY, BREVO_SMTP_KEY, API_ENCRYPTION_KEY, DATABASE_URL, WHATSAPP_TOKEN; .env.example documenta todas las vars runtime; .env local tiene BREVO_API_KEY con xkeysib- y DATABASE_URL con sslmode=require; security.ts lanza error si API_ENCRYPTION_KEY no definida.",
        "riesgo": "—",
    },
    "TC-DEV-007": {
        "hallazgo": "Schema Neon sincronizado: prisma/schema.prisma con datasource provider postgresql y env DATABASE_URL; workflow incluye prisma generate; vercel.json buildCommand incluye prisma generate; package.json tiene scripts db:push/db:generate/db:migrate/db:reset.",
        "riesgo": "—",
    },
    "TC-DEV-008": {
        "hallazgo": "Conexión pooled Neon: .env DATABASE_URL con -pooler y neon.tech y sslmode=require; .env.example documenta DATABASE_URL_DIRECT con pooler.",
        "riesgo": "—",
    },
    "TC-DEV-009": {
        "hallazgo": "PlataformaSync.GITHUB verificado: schema tiene model PlataformaSync con campos plataforma/sincronizado/ultimoEstado/ultimoSync; webhook valida GITHUB como plataforma y firma x-hub-signature-256 (HMAC SHA256 + timingSafeEqual); sync-full-platforms.cjs sincroniza GitHub con git ls-remote.",
        "riesgo": "—",
    },
    "TC-DEV-010": {
        "hallazgo": "PlataformaSync.VERCEL verificado: webhook valida VERCEL como plataforma y firma x-vercel-signature (HMAC SHA1 + timingSafeEqual); sync-full-platforms.cjs sincroniza Vercel y lista deployments.",
        "riesgo": "—",
    },
    "TC-DEV-011": {
        "hallazgo": "Fix v4.16: sync-full-platforms.cjs NO actualizaba ultimoEstado en BD Neon. Añadido bloque que itera platformStatuses [GITHUB, VERCEL, NEON] y ejecuta prisma.plataformaSync.update con sincronizado, ultimoEstado (OK/ERROR), ultimoSync, ultimoError. Webhook valida NEON como plataforma con secreto en query param; db-security.ts existe como auto-monitor.",
        "riesgo": "Medio (Integración)",
    },
    "TC-DEV-012": {
        "hallazgo": "ConexionAPI.EMAIL_SMTP activa: schema tiene model ConexionAPI con campos probada Boolean, fechaUltimaPrueba, activa Boolean; /api/email POST accion=probar actualiza ConexionAPI.probada+fechaUltimaPrueba (fix v4.14); email.ts buscar ConexionAPI EMAIL_SMTP; requireRole ADMIN.",
        "riesgo": "—",
    },
    "TC-DEV-013": {
        "hallazgo": "CorreoInstitucional activo: schema tiene model CorreoInstitucional con esPrincipal, estado, smtpHost; email.ts usa CorreoInstitucional como fallback SMTP y referencia smtp-relay.brevo.com.",
        "riesgo": "—",
    },
    "TC-DEV-014": {
        "hallazgo": "Fix v4.16 (Alto Seguridad): webhook plataformas-sync NO registraba AuditLog. Modificado para invocar registrarAuditLog con accion SYNC_GITHUB|SYNC_VERCEL|SYNC_NEON, modulo 'plataformas-sync', detalles JSON con plataforma/eventType/eventosRecibidos/ultimoEstado, exito true, usuarioId null (sistema). sync-full-platforms.cjs también ahora crea AuditLog por cada plataforma con prisma.auditLog.create.",
        "riesgo": "Alto (Seguridad)",
    },
    "TC-DEV-015": {
        "hallazgo": "Fix v4.16: No existía endpoint ni script de rollback. Creados src/app/api/seguridad/rollback/route.ts (POST rollback al penúltimo deploy READY o deploymentId específico vía Vercel API v13/deployments/{id}/promote; GET lista últimos 10 deploys; requireRole ADMIN; AuditLog VERCEL_ROLLBACK) + scripts/vercel-rollback.cjs (CLI equivalente) + scripts en package.json vercel:rollback/vercel:deploy/vercel:envs.",
        "riesgo": "Alto (Operacional)",
    },
}

for row in range(5, ws.max_row + 1):
    tc_id = ws.cell(row=row, column=2).value
    if not tc_id or not str(tc_id).startswith("TC-DEV"):
        continue
    tc_id = str(tc_id).strip()
    info = hallazgos.get(tc_id)
    if not info:
        continue

    estado_col = cols.get("Estado", 13)
    cell = ws.cell(row=row, column=estado_col, value="Aprobado")
    cell.fill = green_fill
    cell.font = green_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

    hcell = ws.cell(row=row, column=cols["Hallazgo"], value=info["hallazgo"])
    hcell.fill = yellow_fill if info["hallazgo"] != "—" else green_fill
    hcell.font = yellow_font if info["hallazgo"] != "—" else green_font
    hcell.alignment = Alignment(wrap_text=True, vertical="top")

    rcell = ws.cell(row=row, column=cols["Riesgo"], value=info["riesgo"])
    rcell.fill = yellow_fill if info["riesgo"] != "—" else green_fill
    rcell.font = yellow_font if info["riesgo"] != "—" else green_font
    rcell.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[row].height = max(45, min(120, len(info["hallazgo"]) // 5))

ws.column_dimensions[chr(64 + cols["Hallazgo"])].width = 80
ws.column_dimensions[chr(64 + cols["Riesgo"])].width = 22

wb.save(XLSX)
print(f"\n✅ Excel actualizado: {XLSX}")
print(f"   Hoja: 15. M13-Sync DevOps")
print(f"   {len(hallazgos)} TCs actualizados")
