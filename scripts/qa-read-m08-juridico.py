#!/usr/bin/env python3
"""Lee los TCs pendientes del módulo M08-Portal Jurídico desde el Excel."""
import openpyxl
from pathlib import Path

EXCEL = Path("/home/z/my-project/download/plan-pruebas-qa-jsadr-actualizado.xlsx")

wb = openpyxl.load_workbook(EXCEL, data_only=False)
hoja_nombre = None
for name in wb.sheetnames:
    if "M08" in name or "Portal Jur" in name:
        hoja_nombre = name
        break
print(f"Hoja encontrada: {hoja_nombre}")
print()
ws = wb[hoja_nombre]
print(f"Dimensiones: {ws.dimensions}, max_row={ws.max_row}, max_col={ws.max_column}")
print()

print("=== HEADERS (fila 4) ===")
for col in range(1, ws.max_column + 1):
    v = ws.cell(row=4, column=col).value
    print(f"  col {col}: {v!r}")

print()
print("=== FILAS 5-19 (TCs) ===")
for row in range(5, min(ws.max_row + 1, 25)):
    fila = []
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=row, column=col).value
        fila.append(str(v) if v is not None else "")
    # Only print ID, Función, Caso, Resultado Esperado, Estado
    print(f"Fila {row}: ID={fila[1]!s:15} Func={fila[3]!s:25} Caso={fila[4]!s:60} Esperado={fila[10]!s:80} Estado={fila[12]}")
