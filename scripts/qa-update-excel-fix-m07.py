#!/usr/bin/env python3
"""Actualiza el Excel de plan de pruebas QA con los TCs aprobados de M07-Portal Cliente."""
import openpyxl
from pathlib import Path
from openpyxl.styles import PatternFill, Font, Alignment

EXCEL_SRC = Path("/home/z/my-project/plan-pruebas-qa-jsadr.xlsx")
EXCEL_OUT = Path("/home/z/my-project/download/plan-pruebas-qa-jsadr-actualizado.xlsx")

# TCs a actualizar: ID, hallazgo, riesgo
TCS = {
    'TC-PORT-003': {
        'hallazgo': '/api/portal/login bloqueaba PIN a los 3 intentos; el estándar exige 5 (igual que /api/portal/auth).',
        'riesgo': 'Medio — atacante podría probar más combinaciones antes del bloqueo; inconsistencia con /api/portal/auth.',
    },
    'TC-PORT-004': {
        'hallazgo': 'Cumple: validación !cliente.activo → HTTP 403 "Cuenta inactiva" antes de check PIN.',
        'riesgo': 'N/A — sin riesgo (cumplimiento verificado en auditoría).',
    },
    'TC-PORT-006': {
        'hallazgo': 'Cumple: tokenExpira verificado en /prestamos, /cuenta-pago, /firmar, /simular, /clave-dinamica/solicitar, /clave-dinamica/validar y /solicitudes-web POST.',
        'riesgo': 'N/A — sin riesgo (todas las rutas validan tokenExpira).',
    },
    'TC-PORT-009': {
        'hallazgo': 'Cumple: intentosOTP++ y bloqueo basado en firma.maxIntentos (schema default 5). HTTP 401 en fallo parcial, 429 OTP_BLOQUEADO al exceder.',
        'riesgo': 'N/A — sin riesgo (cumplimiento verificado).',
    },
    'TC-PORT-011': {
        'hallazgo': 'Cumple: clave-dinamica/validar genera codigoConfirmacion (32 bytes hex), marca OtpRegistro.usado=true vía marcarOtpVerificado, comparación constant-time.',
        'riesgo': 'N/A — sin riesgo (cumplimiento verificado).',
    },
    'TC-PORT-012': {
        'hallazgo': 'No existía endpoint GET /api/portal/mi-estado. Creado route.ts con validación de token (header o query), identificación por tokenSesion, retorno de préstamos + saldos + próximos vencimientos.',
        'riesgo': 'Alto — cliente no podía consultar su estado de cuenta desde el portal; carencia funcional.',
    },
    'TC-PORT-013': {
        'hallazgo': 'POST /api/solicitudes-web retornaba HTTP 200 en éxito. Cambiado a HTTP 201 (estándar REST para creación de recurso).',
        'riesgo': 'Bajo — inconsistencia con estándar REST; no afecta funcionalidad pero confunde a clientes API.',
    },
    'TC-PORT-014': {
        'hallazgo': 'No existía método DELETE en /api/portal/login. Añadido: acepta token (body o header), limpia tokenSesion=null y tokenExpira=null en BD, registra LOGOUT en AccesoPortal.',
        'riesgo': 'Medio — cliente no podía cerrar sesión; token quedaba vivo hasta expirar (2h).',
    },
    'TC-PORT-015': {
        'hallazgo': 'GET /api/portal/[cedula] no validaba token. Añadida validación completa: token requerido, búsqueda por tokenSesion, comparación clienteAutenticado.cedula === cedula URL, HTTP 403 CROSS_CLIENTE_BLOQUEADO si no coinciden, auditoría de intento en AccesoPortal.',
        'riesgo': 'CRÍTICO — cualquier persona podía consultar datos de cualquier cédula sin autenticación; riesgo de fuga de información financiera.',
    },
}

# Cargar Excel
wb = openpyxl.load_workbook(EXCEL_OUT)
ws = wb['9. M07-Portal Cliente']

# Estilos
fill_amarillo = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
fill_verde = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
font_amarilla = Font(color='9C5700', bold=True)
font_verde = Font(color='006100', bold=True)

# Buscar columna de Hallazgo, Riesgo y Estado (normalmente columnas 12, 13 o adicionales)
# Fila 4 son los headers
headers = {}
for col in range(1, ws.max_column + 1):
    v = ws.cell(row=4, column=col).value
    if v:
        headers[str(v).strip().lower()] = col

# Buscar columnas de Hallazgo, Riesgo y Estado
col_hallazgo = None
col_riesgo = None
col_estado = None
col_id = None
for k, v in headers.items():
    if 'hallazgo' in k:
        col_hallazgo = v
    elif 'riesgo' in k:
        col_riesgo = v
    elif 'estado' in k:
        col_estado = v
    elif k == 'id':
        col_id = v

print(f"Columnas encontradas: ID={col_id}, Hallazgo={col_hallazgo}, Riesgo={col_riesgo}, Estado={col_estado}")

# Si no existen columnas de Hallazgo/Riesgo, añadir al final
if col_hallazgo is None:
    col_hallazgo = ws.max_column + 1
    ws.cell(row=4, column=col_hallazgo, value='Hallazgo')
    ws.cell(row=4, column=col_hallazgo).font = Font(bold=True)
    col_riesgo = col_hallazgo + 1
    ws.cell(row=4, column=col_riesgo, value='Riesgo')
    ws.cell(row=4, column=col_riesgo).font = Font(bold=True)
    print(f"Añadidas nuevas columnas: Hallazgo={col_hallazgo}, Riesgo={col_riesgo}")

# Actualizar filas
updated = 0
for row in range(5, ws.max_row + 1):
    tc_id = ws.cell(row=row, column=col_id).value if col_id else None
    if not tc_id:
        continue
    tc_id = str(tc_id).strip()
    if tc_id in TCS:
        info = TCS[tc_id]
        # Estado
        if col_estado:
            cell = ws.cell(row=row, column=col_estado, value='Aprobado')
            cell.fill = fill_verde
            cell.font = font_verde
            cell.alignment = Alignment(horizontal='center')
        # Hallazgo
        ws.cell(row=row, column=col_hallazgo, value=info['hallazgo'])
        ws.cell(row=row, column=col_hallazgo).fill = fill_amarillo
        ws.cell(row=row, column=col_hallazgo).font = font_amarilla
        ws.cell(row=row, column=col_hallazgo).alignment = Alignment(wrap_text=True, vertical='top')
        # Riesgo
        ws.cell(row=row, column=col_riesgo, value=info['riesgo'])
        ws.cell(row=row, column=col_riesgo).fill = fill_amarillo
        ws.cell(row=row, column=col_riesgo).font = font_amarilla
        ws.cell(row=row, column=col_riesgo).alignment = Alignment(wrap_text=True, vertical='top')
        updated += 1
        print(f"  Actualizado {tc_id}")

print(f"\nTotal TCs actualizados: {updated}")

# Guardar
wb.save(EXCEL_OUT)
print(f"\nExcel guardado: {EXCEL_OUT}")
