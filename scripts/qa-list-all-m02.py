"""Lista todos los TCs del módulo M02-Clientes con su estado."""
from openpyxl import load_workbook
wb = load_workbook("/home/z/my-project/upload/plan-pruebas-qa-jsadr.xlsx", data_only=True)
ws = wb["4. M02-Clientes"]
print(f"Sheet: {ws.title}  max_row={ws.max_row}  max_col={ws.max_column}")
print()
for r in range(1, ws.max_row + 1):
    tc_id = ws.cell(row=r, column=2).value
    estado = ws.cell(row=r, column=13).value
    func = ws.cell(row=r, column=4).value
    caso = ws.cell(row=r, column=5).value
    riesgo = ws.cell(row=r, column=11).value
    hallazgo = ws.cell(row=r, column=12).value
    if tc_id and str(tc_id).startswith('TC-CLI'):
        print(f"row {r:3d} | {str(tc_id):18s} | estado={str(estado or '(vacío)'):15s} | riesgo={str(riesgo or '')[:20]:20s} | {str(func or '')[:25]:25s} | {str(caso or '')[:55]}")
        if hallazgo:
            print(f"          HALLAZGO: {str(hallazgo)[:200]}")
