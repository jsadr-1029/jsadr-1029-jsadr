#!/usr/bin/env python3
"""Actualiza hoja M11-Integraciones del Excel con hallazgos/riesgo/estado."""
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment

XLSX = "/home/z/my-project/download/plan-pruebas-qa-jsadr-actualizado.xlsx"
wb = load_workbook(XLSX)
ws = wb["13. M11-Integraciones"]

# Encabezados actuales (fila 4): ID | Módulo | Función | Caso de Prueba | Tipo | Prioridad |
# Precondiciones | Pasos | Datos de Entrada | Resultado Esperado | Criterios de Aceptación | Estado
# Vamos a añadir 3 columnas nuevas: Hallazgo | Riesgo | Estado (reemplaza la col Estado si existe)

# Encontrar columna "Estado" (col 12) y añadir Hallazgo/Riesgo en 13/14
hdr_row = 4
cols = {}
for c in range(1, ws.max_column + 1):
    v = ws.cell(row=hdr_row, column=c).value
    if v:
        cols[str(v).strip()] = c

print("Columnas:", cols)

# Añadir encabezados Hallazgo y Riesgo si no existen
if "Hallazgo" not in cols:
    new_col = ws.max_column + 1
    ws.cell(row=hdr_row, column=new_col, value="Hallazgo").font = Font(bold=True)
    cols["Hallazgo"] = new_col
if "Riesgo" not in cols:
    new_col = ws.max_column + 1
    ws.cell(row=hdr_row, column=new_col, value="Riesgo").font = Font(bold=True)
    cols["Riesgo"] = new_col

# Estilos
green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
green_font = Font(color="006100")
yellow_font = Font(color="9C5700")

# Datos de hallazgos por TC
hallazgos = {
    "TC-INT-001": {
        "hallazgo": "Endpoint /api/email {accion:'probar'} solo llamaba probarSmtp() y NO actualizaba ConexionAPI.probada ni fechaUltimaPrueba. Fix v4.14: ahora invoca verificarCuentaBrevo() (/v3/account) y persiste probada+fechaUltimaPrueba+resultadoUltimaPrueba en ConexionAPI.EMAIL_SMTP.",
        "riesgo": "Medio (Funcional)",
    },
    "TC-INT-002": {
        "hallazgo": "No existía función que llamara a https://api.brevo.com/v3/account para verificar la BREVO_API_KEY. Fix v4.14: creada verificarCuentaBrevo() en src/lib/email.ts que consulta /v3/account con header api-key y devuelve {success, message, cuenta{email,plan,firstName,lastName}}.",
        "riesgo": "Medio (Integración)",
    },
    "TC-INT-003": {
        "hallazgo": "WhatsApp Cloud API integrada en M09 (fix v4.12): src/lib/whatsapp-cloud.ts llama graph.facebook.com/v18.0/{phoneNumberId}/messages con WHATSAPP_TOKEN y devuelve wamid. Verificado.",
        "riesgo": "—",
    },
    "TC-INT-004": {
        "hallazgo": "Endpoint POST /api/configuracion-global/bancolombia/probar ya implementado: requiere ADMIN (requireRole), llama obtenerAccessToken (OAuth2), actualiza ConexionAPI.probada, registra audit log BANCOLOMBIA_TEST_CONEXION. Verificado.",
        "riesgo": "—",
    },
    "TC-INT-005": {
        "hallazgo": "Endpoint POST /api/pagos/bancolombia-webhook ya implementado: verifica HMAC SHA-256 con X-Signature, usa crypto.timingSafeEqual, mapea estados Bancolombia→APLICADO/ANULADO/PENDIENTE, idempotente si pago ya APLICADO, recalcula saldos si aprobado, fallback CIDR IPs Bancolombia. Verificado.",
        "riesgo": "—",
    },
    "TC-INT-006": {
        "hallazgo": "Workflow .github/workflows/deploy-vercel.yml ya implementado: trigger push a main + workflow_dispatch, checkout v4, setup-node v4 node 20, npm install --legacy-peer-deps, vercel pull --production, prisma generate, vercel deploy --prod --yes, concurrency group. Verificado.",
        "riesgo": "—",
    },
    "TC-INT-007": {
        "hallazgo": ".env.example NO documentaba BREVO_API_KEY ni BREVO_SMTP_KEY. Fix v4.14: añadida sección 6.1 'BREVO (Sendinblue) — HTTPS API + SMTP relay' con BREVO_API_KEY (xkeysib-), BREVO_SMTP_KEY (xsmtpsib-), BREVO_FROM_EMAIL, BREVO_FROM_NAME. Código ya las referencia en email.ts y security.ts.",
        "riesgo": "Bajo (Documentación)",
    },
    "TC-INT-008": {
        "hallazgo": "Conexión Neon Postgres verificada: src/lib/db.ts usa PrismaClient singleton globalThis; prisma/schema.prisma datasource provider=postgresql con env('DATABASE_URL'). Verificado.",
        "riesgo": "—",
    },
    "TC-INT-009": {
        "hallazgo": "Conexión SSL verificada: .env y .env.example usan DATABASE_URL con sslmode=require; schema.prisma datasource url env('DATABASE_URL'). Verificado.",
        "riesgo": "—",
    },
    "TC-INT-010": {
        "hallazgo": "Git remote origin configurado a github.com jsadr-1029/jsadr-1029-jsadr; branch main. Verificado.",
        "riesgo": "—",
    },
    "TC-INT-011": {
        "hallazgo": "Secret scanning: .gitignore excluye .env; .env contiene claves xkeysib- pero NO está commited; .env.example es el template sin secrets reales; código valida prefijo xkeysib- en BREVO_API_KEY. Verificado.",
        "riesgo": "—",
    },
    "TC-INT-012": {
        "hallazgo": "Workflow deploy-vercel.yml completo: checkout@v4, setup-node@v4 node 20, npm install --legacy-peer-deps, vercel pull --environment=production, prisma generate, vercel deploy --prod --yes, secrets VERCEL_TOKEN/VERCEL_ORG_ID/VERCEL_PROJECT_ID, concurrency group, GITHUB_STEP_SUMMARY. Verificado.",
        "riesgo": "—",
    },
    "TC-INT-013": {
        "hallazgo": "Build producción: next.config.ts output='standalone', ignoreBuildErrors=false (no degradar TS), reactStrictMode=true, poweredByHeader=false; vercel.json buildCommand='prisma generate && next build', installCommand='npm install --legacy-peer-deps', framework nextjs. Verificado.",
        "riesgo": "—",
    },
    "TC-INT-014": {
        "hallazgo": "Runtime variables verificadas: email.ts usa process.env.BREVO_API_KEY y BREVO_SMTP_KEY; security.ts usa process.env.API_ENCRYPTION_KEY con throw si no definida; whatsapp-cloud.ts usa process.env.WHATSAPP_TOKEN y WHATSAPP_PHONE_NUMBER_ID; db.ts/schema usan process.env.DATABASE_URL. .env.example documenta todas. Verificado.",
        "riesgo": "—",
    },
    "TC-INT-015": {
        "hallazgo": "Pipeline DevOps completo: workflow dispara en push a main, ejecuta prisma generate (sync schema), vercel deploy --prod; vercel.json buildCommand incluye prisma generate; schema.prisma datasource postgresql; DATABASE_URL via Vercel env (no hardcodeada). Verificado.",
        "riesgo": "—",
    },
}

# Recorrer filas y actualizar
for row in range(5, ws.max_row + 1):
    tc_id = ws.cell(row=row, column=2).value
    if not tc_id or not str(tc_id).startswith("TC-INT"):
        continue
    tc_id = str(tc_id).strip()
    info = hallazgos.get(tc_id)
    if not info:
        continue

    # Marcar Estado (col 12) como Aprobado
    estado_col = cols.get("Estado", 12)
    cell = ws.cell(row=row, column=estado_col, value="Aprobado")
    cell.fill = green_fill
    cell.font = green_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

    # Hallazgo (col nueva)
    hcell = ws.cell(row=row, column=cols["Hallazgo"], value=info["hallazgo"])
    hcell.fill = yellow_fill if info["hallazgo"] != "—" else green_fill
    hcell.font = yellow_font if info["hallazgo"] != "—" else green_font
    hcell.alignment = Alignment(wrap_text=True, vertical="top")

    # Riesgo (col nueva)
    rcell = ws.cell(row=row, column=cols["Riesgo"], value=info["riesgo"])
    rcell.fill = yellow_fill if info["riesgo"] != "—" else green_fill
    rcell.font = yellow_font if info["riesgo"] != "—" else green_font
    rcell.alignment = Alignment(horizontal="center", vertical="center")

    # Ajustar altura de fila
    ws.row_dimensions[row].height = max(45, min(120, len(info["hallazgo"]) // 5))

# Ajustar ancho columnas nuevas
ws.column_dimensions[chr(64 + cols["Hallazgo"])].width = 80
ws.column_dimensions[chr(64 + cols["Riesgo"])].width = 22

wb.save(XLSX)
print(f"\n✅ Excel actualizado: {XLSX}")
print(f"   Hoja: 13. M11-Integraciones")
print(f"   {len(hallazgos)} TCs actualizados")
