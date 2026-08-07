"""Lee todos los casos de prueba pendientes del M02-Clientes."""
from openpyxl import load_workbook

WB_PATH = "/home/z/my-project/upload/plan-pruebas-qa-jsadr.xlsx"
wb = load_workbook(WB_PATH, data_only=True)
ws = wb["4. M02-Clientes"]

# Pendientes: rows 6,7,8,9,10,12,14,15,16,17,18,19 (TC-CLI-002 a 006, 008, 010-015)
PENDIENTES = [6,7,8,9,10,12,14,15,16,17,18,19]

for r in PENDIENTES:
    tc_id = ws.cell(row=r, column=2).value
    func = ws.cell(row=r, column=4).value
    caso = ws.cell(row=r, column=5).value
    tipo = ws.cell(row=r, column=6).value
    prio = ws.cell(row=r, column=7).value
    pre = ws.cell(row=r, column=8).value
    pasos = ws.cell(row=r, column=9).value
    datos = ws.cell(row=r, column=10).value
    esperado = ws.cell(row=r, column=11).value
    criterios = ws.cell(row=r, column=12).value
    estado = ws.cell(row=r, column=13).value
    
    print(f"\n{'='*70}")
    print(f"FILA {r} | {tc_id} | Estado: {estado}")
    print(f"{'='*70}")
    print(f"Función: {func}")
    print(f"Caso: {caso}")
    print(f"Tipo: {tipo} | Prioridad: {prio}")
    print(f"Precondiciones: {pre}")
    print(f"Pasos: {pasos}")
    print(f"Datos entrada: {datos}")
    print(f"Resultado esperado: {esperado}")
    print(f"Criterios aceptación: {criterios}")
